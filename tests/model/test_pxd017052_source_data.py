"""Regression tests for the fail-closed PXD017052 source-data audit."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from biointerfaceos.pxd017052_source_data import (
    PXD017052SourceDataError,
    PXD017052SourceDataWorkflow,
)

ROOT = Path(__file__).resolve().parents[2]


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes(), usedforsecurity=False).hexdigest()


def _write_data_1(path: Path, result_units: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "Supplementary Data 2.   3 NP protein groups from MaxQuant")
    headers = ["Protein IDs"]
    headers.extend(f"Intensity {unit}" for unit in result_units)
    headers.extend(f"LFQ intensity {unit}" for unit in result_units)
    for column, value in enumerate(headers, start=1):
        sheet.cell(2, column, value)
    for row in range(3, 847):
        sheet.cell(row, 1, f"P{row}")
    for column in range(len(headers) + 1, 98):
        sheet.cell(2, column, f"field_{column}")
    workbook.save(path)


def _write_source_data(path: Path, replicates: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Figure 3A"
    sheet.cell(2, 1, "Nanoparticles")
    sheet.cell(2, 2, "Assay Replicate")
    for row, record in enumerate(replicates, start=3):
        sheet.cell(row, 1, record["particle"])
        sheet.cell(row, 2, record["assay_replicate"])
    dls = workbook.create_sheet("Figure 2 DLS")
    dls.cell(2, 1, "S-007-008")
    dls.cell(2, 6, "S-003-001")
    dls.cell(2, 11, "S-011-001")
    workbook.save(path)


def _write_pride_readme(path: Path, result_units: list[str]) -> None:
    rows = [
        "ID\tNAME\tURI\tTYPE\tMAPPINGS",
        "68\ttxt3NP.zip\thttps://example.test/txt3NP.zip\tSEARCH\t69,70,71,72,73,74,75,76,77",
    ]
    rows.extend(
        f"{index}\t{unit}.raw\thttps://example.test/{unit}.raw\tRAW\t-"
        for index, unit in enumerate(result_units, start=69)
    )
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def _prepared_workflow(tmp_path: Path) -> PXD017052SourceDataWorkflow:
    registry = json.loads(
        (ROOT / "docs/data/R2_T131_PXD017052_SOURCE_DATA_AUDIT_REGISTRY.json").read_text(
            encoding="utf-8"
        )
    )
    raw_root = tmp_path / "raw"
    raw_root.mkdir()
    schema = registry["workbook_schema"]
    assets = {item["asset_id"]: item for item in registry["assets"]}
    _write_data_1(
        raw_root / assets["supplementary_data_1"]["file_name"],
        schema["supplementary_data_1"]["result_unit_ids"],
    )
    _write_source_data(
        raw_root / assets["source_data"]["file_name"],
        schema["source_data"]["figure_3a"]["particle_replicates"],
    )
    _write_pride_readme(
        raw_root / registry["particle_unit_map"]["pride_readme_file_name"],
        schema["supplementary_data_1"]["result_unit_ids"],
    )
    for asset_id in ("supplementary_information", "supplementary_file_description"):
        (raw_root / assets[asset_id]["file_name"]).write_bytes(asset_id.encode("utf-8"))
    for asset in registry["assets"]:
        path = raw_root / asset["file_name"]
        asset["bytes"] = path.stat().st_size
        asset["sha256"] = _sha256(path)
        asset["md5"] = _md5(path)
        asset["publisher_etag"] = asset["md5"]
    readme = raw_root / registry["particle_unit_map"]["pride_readme_file_name"]
    registry["particle_unit_map"]["pride_readme_bytes"] = readme.stat().st_size
    registry["particle_unit_map"]["pride_readme_sha256"] = _sha256(readme)
    registry["particle_unit_map"]["pride_readme_md5"] = _md5(readme)
    registry_path = tmp_path / "registry.json"
    registry_path.write_text(json.dumps(registry), encoding="utf-8")
    return PXD017052SourceDataWorkflow(
        ROOT,
        registry_path=registry_path,
        raw_root=raw_root,
        output_root=tmp_path / "audit",
    )


def test_audit_verifies_public_assets_without_inferring_particle_labels(tmp_path: Path) -> None:
    workflow = _prepared_workflow(tmp_path)

    summary = workflow.run(strict=True)
    receipt = json.loads(summary.receipt_path.read_text(encoding="utf-8"))

    assert summary.official_asset_count == 4
    assert summary.result_unit_count == 9
    assert summary.result_to_raw_match_count == 9
    assert summary.explicit_raw_to_particle_map_count == 0
    assert receipt["status"] == "VERIFIED_PUBLIC_ASSETS_INCOMPLETE_SOURCE_UNIT_TO_PARTICLE_MAP"
    assert receipt["admission"] == "NOT_ADMITTED"
    assert receipt["model_use"] == "PROHIBITED"
    assert workflow.verify() == summary


def test_audit_requires_strict_mode(tmp_path: Path) -> None:
    workflow = _prepared_workflow(tmp_path)

    with pytest.raises(PXD017052SourceDataError, match="requires --strict"):
        workflow.run()


def test_audit_rejects_raw_asset_tampering(tmp_path: Path) -> None:
    workflow = _prepared_workflow(tmp_path)
    asset = workflow.raw_root / "41467_2020_17033_MOESM2_ESM.docx"
    asset.write_bytes(asset.read_bytes() + b"tampered")

    with pytest.raises(PXD017052SourceDataError, match="byte count differs|checksum differs"):
        workflow.run(strict=True)


def test_audit_rejects_pride_readme_tampering(tmp_path: Path) -> None:
    workflow = _prepared_workflow(tmp_path)
    readme = workflow.raw_root / "README.txt"
    readme.write_text("tampered\n", encoding="utf-8")

    with pytest.raises(PXD017052SourceDataError, match="README checksum differs"):
        workflow.run(strict=True)


def test_audit_rejects_inferred_raw_to_particle_crosswalk(tmp_path: Path) -> None:
    workflow = _prepared_workflow(tmp_path)
    registry = json.loads(workflow.registry_path.read_text(encoding="utf-8"))
    registry["particle_unit_map"]["explicit_raw_to_particle_crosswalk"] = [
        {"raw": "EXP18102_X1066_A_MSB37240A.raw", "particle": "SP-003"}
    ]
    workflow.registry_path.write_text(json.dumps(registry), encoding="utf-8")

    with pytest.raises(PXD017052SourceDataError, match="source-unit mapping boundary"):
        workflow.run(strict=True)
