"""Fail-closed audit of the public PXD017052 source-data route.

The study's CC-BY supplementary assets make the quantitative 3-NP result
table, the three particle records and the PRIDE raw-file names auditable.
They do *not* publish a row-level raw-file-to-particle crosswalk.  This
workflow records that distinction without inferring particle labels from the
order of files, values, or replicate groups.
"""

from __future__ import annotations

import hashlib
import json
import stat
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PXD017052SourceDataError(RuntimeError):
    """Raised when the PXD017052 provenance boundary is weakened or stale."""


@dataclass(frozen=True)
class PXD017052SourceDataSummary:
    """Compact accounting for the audited but non-admitted source route."""

    official_asset_count: int
    result_unit_count: int
    result_to_raw_match_count: int
    explicit_raw_to_particle_map_count: int
    receipt_path: Path


def _canonical(value: Any) -> bytes:
    return (
        json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n"
    ).encode("utf-8")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise PXD017052SourceDataError(f"{label} must be an object")
    return dict(value)


def _list(value: Any, label: str, *, minimum: int = 0) -> list[Any]:
    if not isinstance(value, list) or len(value) < minimum:
        raise PXD017052SourceDataError(f"{label} must contain at least {minimum} items")
    return value


def _string(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise PXD017052SourceDataError(f"{label} must be a non-empty string")
    return value.strip()


def _integer(value: Any, label: str, *, minimum: int = 0) -> int:
    if not isinstance(value, int) or isinstance(value, bool) or value < minimum:
        raise PXD017052SourceDataError(f"{label} must be an integer >= {minimum}")
    return value


class PXD017052SourceDataWorkflow:
    """Verify public publisher assets without creating an inferred unit map."""

    AUDIT_ID = "bioif-r2-pxd017052-source-data-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R2_T131_PXD017052_SOURCE_DATA_AUDIT_REGISTRY.json"
    RAW_RELATIVE = "data/raw/r2_t131_pxd017052"
    OUTPUT_RELATIVE = "reports/review_round_2/pxd017052_source_data/v1.0.0"
    REQUIRED_FALSE = (
        "target_frozen",
        "model_fitted",
        "paired_ablations_run",
        "external_ood_evaluated",
        "negative_controls_run",
        "independent_validation",
        "scientific_submission_ready",
    )
    REQUIRED_REGISTRY_FIELDS = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "development_cutoff",
        "evidence_class",
        "allowed_claim_level",
        "primary_study",
        "source_policy",
        "assets",
        "workbook_schema",
        "material_records",
        "particle_unit_map",
        "decision",
    }
    REQUIRED_PRIMARY_FIELDS = {
        "article_doi",
        "article_url",
        "article_license",
        "pride_accession",
        "pride_file_index_url",
    }
    REQUIRED_POLICY_FIELDS = {
        "cc0_cohort_license",
        "candidate_ccby_policy",
        "prohibit_implicit_source_unit_to_particle_inference",
        "require_explicit_raw_to_particle_crosswalk",
        "require_second_independent_lab_for_target",
        "prohibit_model_use_before_t121_amendment",
    }
    REQUIRED_ASSET_FIELDS = {
        "asset_id",
        "file_name",
        "publisher_url",
        "content_type",
        "bytes",
        "sha256",
        "md5",
        "publisher_etag",
        "role",
        "source_reference",
    }
    EXPECTED_ASSET_IDS = {
        "supplementary_information",
        "supplementary_file_description",
        "supplementary_data_1",
        "source_data",
    }
    EXPECTED_RESULT_UNIT_COUNT = 9
    EXPECTED_MATERIAL_COUNT = 3

    def __init__(
        self,
        root: Path,
        *,
        registry_path: Path | None = None,
        raw_root: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.raw_root = raw_root or self.root / self.RAW_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise PXD017052SourceDataError(f"cannot parse {label}") from exc

    @staticmethod
    def _require(condition: bool, label: str) -> None:
        if not condition:
            raise PXD017052SourceDataError(label)

    @staticmethod
    def _sha256_value(value: Any, label: str) -> str:
        text = _string(value, label).lower()
        if len(text) != 64 or any(character not in "0123456789abcdef" for character in text):
            raise PXD017052SourceDataError(f"{label} must be a SHA-256 digest")
        return text

    @staticmethod
    def _md5_value(value: Any, label: str) -> str:
        text = _string(value, label).lower()
        if len(text) != 32 or any(character not in "0123456789abcdef" for character in text):
            raise PXD017052SourceDataError(f"{label} must be an MD5 digest")
        return text

    def _registry(self) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
        registry = self._json(self.registry_path, "T131 source-data audit registry")
        self._require(
            set(registry) == self.REQUIRED_REGISTRY_FIELDS and registry.get("schema_version") == 1,
            "T131 registry fields are invalid",
        )
        self._require(
            registry.get("audit_id") == self.AUDIT_ID,
            "T131 registry identity is invalid",
        )
        self._require(
            registry.get("development_cutoff") == "2024-12-31T23:59:59+00:00",
            "T131 development cutoff changed",
        )
        self._require(
            registry.get("evidence_class") == "DEVELOPMENT_OBSERVATION"
            and registry.get("allowed_claim_level") == "EXPLORATORY",
            "T131 evidence semantics are unsafe",
        )
        _string(registry.get("evaluated_at"), "T131 evaluated_at")

        primary = _mapping(registry.get("primary_study"), "T131 primary study")
        self._require(set(primary) == self.REQUIRED_PRIMARY_FIELDS, "T131 primary study fields")
        self._require(
            primary.get("article_doi") == "10.1038/s41467-020-17033-7"
            and primary.get("article_license") == "CC-BY-4.0"
            and primary.get("pride_accession") == "PXD017052",
            "T131 primary study identity is invalid",
        )
        for key in ("article_url", "pride_file_index_url"):
            self._require(
                _string(primary.get(key), f"T131 primary study {key}").startswith("https://"),
                "T131 primary locator is invalid",
            )

        policy = _mapping(registry.get("source_policy"), "T131 source policy")
        self._require(set(policy) == self.REQUIRED_POLICY_FIELDS, "T131 policy fields are invalid")
        self._require(
            policy.get("cc0_cohort_license") == "CC0-1.0"
            and policy.get("candidate_ccby_policy") == "EXPLICIT_AMENDMENT_REQUIRED"
            and all(
                policy.get(key) is True
                for key in self.REQUIRED_POLICY_FIELDS
                - {"cc0_cohort_license", "candidate_ccby_policy"}
            ),
            "T131 source policy boundary is weakened",
        )

        assets: dict[str, dict[str, Any]] = {}
        for item in _list(registry.get("assets"), "T131 assets", minimum=4):
            asset = _mapping(item, "T131 asset")
            self._require(set(asset) == self.REQUIRED_ASSET_FIELDS, "T131 asset fields are invalid")
            asset_id = _string(asset.get("asset_id"), "T131 asset identity")
            self._require(asset_id not in assets, "T131 asset identity is duplicated")
            _string(asset.get("file_name"), "T131 asset file name")
            self._require(
                _string(asset.get("publisher_url"), "T131 publisher URL").startswith("https://"),
                "T131 publisher URL is invalid",
            )
            _string(asset.get("content_type"), "T131 asset content type")
            _integer(asset.get("bytes"), "T131 asset bytes", minimum=1)
            self._sha256_value(asset.get("sha256"), "T131 asset SHA-256")
            self._md5_value(asset.get("md5"), "T131 asset MD5")
            self._md5_value(asset.get("publisher_etag"), "T131 publisher ETag")
            _string(asset.get("role"), "T131 asset role")
            _string(asset.get("source_reference"), "T131 asset source reference")
            assets[asset_id] = asset
        self._require(set(assets) == self.EXPECTED_ASSET_IDS, "T131 asset inventory is invalid")

        schema = _mapping(registry.get("workbook_schema"), "T131 workbook schema")
        self._require(
            set(schema) == {"supplementary_data_1", "source_data"},
            "T131 workbook schema fields are invalid",
        )
        data_1 = _mapping(schema.get("supplementary_data_1"), "T131 supplementary data schema")
        source_data = _mapping(schema.get("source_data"), "T131 source data schema")
        self._require(
            data_1.get("file_name") == assets["supplementary_data_1"]["file_name"]
            and data_1.get("sheet_name") == "Sheet1"
            and data_1.get("row_count") == 846
            and data_1.get("column_count") == 97
            and data_1.get("internal_title")
            == "Supplementary Data 2.   3 NP protein groups from MaxQuant",
            "T131 supplementary data schema is invalid",
        )
        result_units = _list(data_1.get("result_unit_ids"), "T131 result units", minimum=1)
        self._require(
            len(result_units) == self.EXPECTED_RESULT_UNIT_COUNT
            and all(isinstance(value, str) and value.strip() for value in result_units)
            and len(set(result_units)) == len(result_units),
            "T131 result unit inventory is invalid",
        )
        self._require(
            source_data.get("file_name") == assets["source_data"]["file_name"],
            "T131 source-data file reference is invalid",
        )
        figure_3a = _mapping(source_data.get("figure_3a"), "T131 Figure 3A schema")
        self._require(
            set(figure_3a) == {"sheet_name", "particle_replicates"}
            and figure_3a.get("sheet_name") == "Figure 3A",
            "T131 Figure 3A schema is invalid",
        )
        particle_replicates = _list(
            figure_3a.get("particle_replicates"), "T131 particle replicates", minimum=9
        )
        normalized_replicates: set[tuple[str, int]] = set()
        for item in particle_replicates:
            row = _mapping(item, "T131 particle replicate")
            self._require(set(row) == {"particle", "assay_replicate"}, "T131 replicate fields")
            normalized_replicates.add(
                (
                    _string(row.get("particle"), "T131 particle identifier"),
                    _integer(row.get("assay_replicate"), "T131 assay replicate", minimum=1),
                )
            )
        self._require(
            len(normalized_replicates) == self.EXPECTED_RESULT_UNIT_COUNT,
            "T131 particle replicate inventory is invalid",
        )
        self._require(
            source_data.get("figure_2_dls_particle_labels")
            == ["S-007-008", "S-003-001", "S-011-001"],
            "T131 Figure 2 DLS identifiers are invalid",
        )

        materials = _list(registry.get("material_records"), "T131 material records", minimum=3)
        material_ids: set[str] = set()
        for item in materials:
            material = _mapping(item, "T131 material record")
            self._require(
                set(material)
                == {
                    "particle",
                    "supplement_label",
                    "source_data_label",
                    "surface_chemistry",
                    "z_average_size_nm",
                    "pdi",
                    "zeta_potential_mv",
                    "source_reference",
                },
                "T131 material record fields are invalid",
            )
            particle = _string(material.get("particle"), "T131 material particle")
            self._require(particle not in material_ids, "T131 material particle is duplicated")
            material_ids.add(particle)
            _string(material.get("supplement_label"), "T131 material supplement label")
            _string(material.get("source_data_label"), "T131 material source-data label")
            _string(material.get("surface_chemistry"), "T131 material surface chemistry")
            for field in ("z_average_size_nm", "pdi", "zeta_potential_mv"):
                if not isinstance(material.get(field), (int, float)) or isinstance(
                    material.get(field), bool
                ):
                    raise PXD017052SourceDataError(f"T131 material {field} is not numeric")
            self._require(
                _string(material.get("source_reference"), "T131 material source reference")
                == "41467_2020_17033_MOESM1_ESM.pdf page 15, Supplementary Table 1",
                "T131 material source reference is invalid",
            )
        self._require(
            material_ids == {"SP-003", "SP-007", "SP-011"}
            and len(materials) == self.EXPECTED_MATERIAL_COUNT,
            "T131 material inventory is invalid",
        )

        unit_map = _mapping(registry.get("particle_unit_map"), "T131 particle-unit map")
        self._require(
            set(unit_map)
            == {
                "pride_result_archive",
                "pride_readme_url",
                "pride_readme_file_name",
                "pride_readme_bytes",
                "pride_readme_sha256",
                "pride_readme_md5",
                "raw_unit_files",
                "result_unit_ids",
                "result_to_raw_status",
                "source_data_particle_replicates",
                "explicit_raw_to_particle_crosswalk",
                "raw_to_particle_status",
                "unit_to_material_map_status",
            },
            "T131 particle-unit map fields are invalid",
        )
        self._require(
            unit_map.get("pride_result_archive") == "txt3NP.zip"
            and _string(unit_map.get("pride_readme_url"), "T131 PRIDE README URL").startswith(
                "https://"
            )
            and unit_map.get("pride_readme_file_name") == "README.txt"
            and _integer(unit_map.get("pride_readme_bytes"), "T131 PRIDE README bytes", minimum=1)
            and self._sha256_value(
                unit_map.get("pride_readme_sha256"), "T131 PRIDE README SHA-256"
            )
            and self._md5_value(unit_map.get("pride_readme_md5"), "T131 PRIDE README MD5"),
            "T131 PRIDE evidence is invalid",
        )
        raw_units = _list(unit_map.get("raw_unit_files"), "T131 raw units", minimum=1)
        self._require(
            len(raw_units) == self.EXPECTED_RESULT_UNIT_COUNT
            and all(isinstance(value, str) and value.endswith(".raw") for value in raw_units)
            and len(set(raw_units)) == len(raw_units),
            "T131 raw unit inventory is invalid",
        )
        self._require(
            unit_map.get("result_unit_ids") == result_units
            and [value[:-4] for value in raw_units] == result_units
            and unit_map.get("result_to_raw_status") == "EXACT_HEADER_TO_RAW_BASENAME_VERIFIED"
            and unit_map.get("source_data_particle_replicates") == particle_replicates
            and unit_map.get("explicit_raw_to_particle_crosswalk") == []
            and unit_map.get("raw_to_particle_status") == "MISSING_IN_PUBLISHED_ASSETS"
            and unit_map.get("unit_to_material_map_status")
            == "INCOMPLETE_NO_EXPLICIT_RAW_TO_PARTICLE_CROSSWALK",
            "T131 source-unit mapping boundary is invalid",
        )

        decision = _mapping(registry.get("decision"), "T131 decision")
        self._require(
            set(decision)
            == {
                "status",
                "admission",
                "cc0_cohort_status",
                "ccby_candidate_cohort_status",
                "model_use",
                "blocked_reasons",
            }
            and decision.get("status")
            == "VERIFIED_PUBLIC_ASSETS_INCOMPLETE_SOURCE_UNIT_TO_PARTICLE_MAP"
            and decision.get("admission") == "NOT_ADMITTED"
            and decision.get("cc0_cohort_status") == "UNCHANGED"
            and decision.get("ccby_candidate_cohort_status") == "NOT_CREATED_INCOMPLETE_MAP"
            and decision.get("model_use") == "PROHIBITED",
            "T131 decision boundary is invalid",
        )
        blocked_reasons = _list(decision.get("blocked_reasons"), "T131 blocked reasons", minimum=3)
        self._require(
            all(isinstance(reason, str) and reason.strip() for reason in blocked_reasons),
            "T131 blocked reasons are invalid",
        )
        return registry, assets

    def _raw_assets(self, assets: Mapping[str, Mapping[str, Any]]) -> list[dict[str, Any]]:
        root = self.raw_root.resolve(strict=False)
        if not root.is_dir():
            raise PXD017052SourceDataError("T131 protected raw-asset directory is missing")
        verified: list[dict[str, Any]] = []
        for asset_id in sorted(assets):
            asset = assets[asset_id]
            path = (root / str(asset["file_name"])).resolve(strict=False)
            if not path.is_relative_to(root) or not path.is_file():
                raise PXD017052SourceDataError(f"T131 raw asset {asset_id} is missing")
            if path.stat().st_size != asset["bytes"]:
                raise PXD017052SourceDataError(f"T131 raw asset {asset_id} byte count differs")
            if _sha256(path) != asset["sha256"]:
                raise PXD017052SourceDataError(f"T131 raw asset {asset_id} checksum differs")
            if hashlib.md5(path.read_bytes(), usedforsecurity=False).hexdigest() != asset["md5"]:
                raise PXD017052SourceDataError(f"T131 raw asset {asset_id} MD5 differs")
            verified.append(
                {
                    "asset_id": asset_id,
                    "file_name": asset["file_name"],
                    "bytes": asset["bytes"],
                    "sha256": asset["sha256"],
                    "publisher_etag": asset["publisher_etag"],
                    "role": asset["role"],
                }
            )
        return verified

    def _pride_readme_raw_units(self, unit_map: Mapping[str, Any]) -> dict[str, Any]:
        """Verify the downloaded official README and extract only its file identity map."""
        readme_path = self.raw_root / str(unit_map["pride_readme_file_name"])
        if not readme_path.is_file():
            raise PXD017052SourceDataError("T131 PRIDE README is missing")
        self._require(
            readme_path.stat().st_size == unit_map["pride_readme_bytes"]
            and _sha256(readme_path) == unit_map["pride_readme_sha256"]
            and hashlib.md5(readme_path.read_bytes(), usedforsecurity=False).hexdigest()
            == unit_map["pride_readme_md5"],
            "T131 PRIDE README checksum differs",
        )
        try:
            rows = [
                line.split("\t") for line in readme_path.read_text(encoding="utf-8").splitlines()
            ]
        except (OSError, UnicodeError) as exc:
            raise PXD017052SourceDataError("T131 PRIDE README cannot be decoded") from exc
        self._require(
            rows and rows[0] == ["ID", "NAME", "URI", "TYPE", "MAPPINGS"],
            "T131 PRIDE README schema differs",
        )
        records = {row[0]: row for row in rows[1:] if len(row) == 5}
        search = records.get("68")
        self._require(
            search is not None
            and search[1] == unit_map["pride_result_archive"]
            and search[3] == "SEARCH"
            and search[4] == "69,70,71,72,73,74,75,76,77",
            "T131 PRIDE search-to-raw mapping differs",
        )
        raw_units = _list(unit_map["raw_unit_files"], "T131 raw units", minimum=1)
        expected_ids = [str(value) for value in range(69, 78)]
        self._require(
            [
                records.get(identifier, [None, None, None, None, None])[1]
                for identifier in expected_ids
            ]
            == raw_units
            and all(
                records[identifier][3] == "RAW" and records[identifier][4] == "-"
                for identifier in expected_ids
            ),
            "T131 PRIDE raw-unit map differs",
        )
        return {
            "file_name": readme_path.name,
            "bytes": readme_path.stat().st_size,
            "sha256": _sha256(readme_path),
            "result_archive": search[1],
            "mapped_raw_unit_count": len(raw_units),
        }

    def _workbook_schema(
        self, registry: Mapping[str, Any], assets: Mapping[str, Mapping[str, Any]]
    ) -> dict[str, Any]:
        schema = _mapping(registry["workbook_schema"], "T131 workbook schema")
        data_1_schema = _mapping(schema["supplementary_data_1"], "T131 data 1 schema")
        source_schema = _mapping(schema["source_data"], "T131 source data schema")
        data_1_path = self.raw_root / str(assets["supplementary_data_1"]["file_name"])
        source_data_path = self.raw_root / str(assets["source_data"]["file_name"])
        try:
            data_1_book = load_workbook(data_1_path, read_only=True, data_only=True)
            source_data_book = load_workbook(source_data_path, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl has several reader-specific exception types.
            raise PXD017052SourceDataError("T131 workbook cannot be parsed") from exc
        self._require(data_1_book.sheetnames == ["Sheet1"], "T131 data-1 sheet inventory differs")
        worksheet = data_1_book["Sheet1"]
        self._require(
            worksheet.max_row == data_1_schema["row_count"]
            and worksheet.max_column == data_1_schema["column_count"]
            and worksheet.cell(1, 1).value == data_1_schema["internal_title"],
            "T131 data-1 dimensions or title differ",
        )
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
        result_units = list(data_1_schema["result_unit_ids"])
        intensity_headers = [
            value
            for value in headers
            if isinstance(value, str) and value.startswith("Intensity EXP")
        ]
        lfq_headers = [
            value
            for value in headers
            if isinstance(value, str) and value.startswith("LFQ intensity EXP")
        ]
        self._require(
            intensity_headers == [f"Intensity {value}" for value in result_units]
            and lfq_headers == [f"LFQ intensity {value}" for value in result_units],
            "T131 quantitative headers do not close the result-to-raw link",
        )
        figure_3a = _mapping(source_schema["figure_3a"], "T131 Figure 3A schema")
        self._require(
            "Figure 2 DLS" in source_data_book.sheetnames
            and figure_3a["sheet_name"] in source_data_book.sheetnames,
            "T131 source-data sheets are missing",
        )
        figure_3a_sheet = source_data_book[str(figure_3a["sheet_name"])]
        self._require(
            [figure_3a_sheet.cell(2, column).value for column in (1, 2)]
            == ["Nanoparticles", "Assay Replicate"],
            "T131 Figure 3A headers differ",
        )
        actual_replicates = [
            {
                "particle": figure_3a_sheet.cell(row, 1).value,
                "assay_replicate": figure_3a_sheet.cell(row, 2).value,
            }
            for row in range(3, 12)
        ]
        self._require(
            actual_replicates == figure_3a["particle_replicates"],
            "T131 Figure 3A particle replicate rows differ",
        )
        dls_sheet = source_data_book["Figure 2 DLS"]
        self._require(
            [dls_sheet.cell(2, column).value for column in (1, 6, 11)]
            == source_schema["figure_2_dls_particle_labels"],
            "T131 Figure 2 DLS labels differ",
        )
        return {
            "supplementary_data_1": {
                "sheet_name": "Sheet1",
                "row_count": worksheet.max_row,
                "column_count": worksheet.max_column,
                "result_unit_count": len(result_units),
                "intensity_header_count": len(intensity_headers),
                "lfq_intensity_header_count": len(lfq_headers),
            },
            "source_data": {
                "figure_3a_particle_replicate_count": len(actual_replicates),
                "figure_2_dls_particle_label_count": 3,
            },
        }

    def run(self, *, strict: bool = False) -> PXD017052SourceDataSummary:
        """Write a strict public-source audit that preserves non-admission."""
        if not strict:
            raise PXD017052SourceDataError("PXD017052 source-data audit requires --strict")
        if self.output_root.exists():
            raise PXD017052SourceDataError("PXD017052 source-data audit already executed")
        registry, assets = self._registry()
        verified_assets = self._raw_assets(assets)
        unit_map = _mapping(registry["particle_unit_map"], "T131 particle-unit map")
        pride_readme = self._pride_readme_raw_units(unit_map)
        workbook_schema = self._workbook_schema(registry, assets)
        decision = _mapping(registry["decision"], "T131 decision")
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "audited_at": registry["evaluated_at"],
            "registry_sha256": _sha256(self.registry_path),
            "status": decision["status"],
            "evidence_class": registry["evidence_class"],
            "allowed_claim_level": registry["allowed_claim_level"],
            "primary_study": registry["primary_study"],
            "verified_raw_assets": verified_assets,
            "pride_readme_verification": pride_readme,
            "official_asset_count": len(verified_assets),
            "workbook_schema_verification": workbook_schema,
            "result_unit_count": len(unit_map["result_unit_ids"]),
            "pride_raw_unit_count": len(unit_map["raw_unit_files"]),
            "result_to_raw_match_count": len(unit_map["result_unit_ids"]),
            "explicit_raw_to_particle_map_count": len(
                unit_map["explicit_raw_to_particle_crosswalk"]
            ),
            "material_record_count": len(registry["material_records"]),
            "source_unit_to_material_map_status": unit_map["unit_to_material_map_status"],
            "admission": decision["admission"],
            "cc0_cohort_status": decision["cc0_cohort_status"],
            "ccby_candidate_cohort_status": decision["ccby_candidate_cohort_status"],
            "model_use": decision["model_use"],
            "blocked_reasons": decision["blocked_reasons"],
            "next_required_evidence": [
                "An official source asset must explicitly map each of the nine PRIDE raw/result "
                "units to SP-003, SP-007, or SP-011; file ordering, signal values and "
                "replicate grouping remain prohibited substitutes.",
                "Only after that crosswalk and an explicit CC-BY cohort amendment may the "
                "route be considered outside the unchanged T129 CC0 cohort.",
                "A second independent laboratory with the same frozen endpoint and a T121 "
                "amendment remain required before any T123 model run.",
            ],
            **{field: False for field in self.REQUIRED_FALSE},
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        report_path = self.output_root / "pxd017052_source_data_report.json"
        report_path.write_bytes(_canonical(report))
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": report["status"],
            "pxd017052_source_data_report_sha256": _sha256(report_path),
            "official_asset_count": report["official_asset_count"],
            "result_unit_count": report["result_unit_count"],
            "pride_raw_unit_count": report["pride_raw_unit_count"],
            "result_to_raw_match_count": report["result_to_raw_match_count"],
            "explicit_raw_to_particle_map_count": report["explicit_raw_to_particle_map_count"],
            "material_record_count": report["material_record_count"],
            "admission": report["admission"],
            "cc0_cohort_status": report["cc0_cohort_status"],
            "ccby_candidate_cohort_status": report["ccby_candidate_cohort_status"],
            "model_use": report["model_use"],
            **{field: False for field in self.REQUIRED_FALSE},
        }
        receipt_path = self.output_root / "pxd017052_source_data_receipt.json"
        receipt_path.write_bytes(_canonical(receipt))
        for path in self.output_root.iterdir():
            path.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        self.output_root.chmod(
            stat.S_IRUSR | stat.S_IXUSR | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH
        )
        return PXD017052SourceDataSummary(
            official_asset_count=len(verified_assets),
            result_unit_count=len(unit_map["result_unit_ids"]),
            result_to_raw_match_count=len(unit_map["result_unit_ids"]),
            explicit_raw_to_particle_map_count=0,
            receipt_path=receipt_path,
        )

    def verify(self) -> PXD017052SourceDataSummary:
        """Verify output hashes and ensure source incompleteness remains explicit."""
        report_path = self.output_root / "pxd017052_source_data_report.json"
        receipt_path = self.output_root / "pxd017052_source_data_receipt.json"
        report = self._json(report_path, "T131 source-data report")
        receipt = self._json(receipt_path, "T131 source-data receipt")
        registry, assets = self._registry()
        verified_assets = self._raw_assets(assets)
        unit_map = _mapping(registry["particle_unit_map"], "T131 particle-unit map")
        pride_readme = self._pride_readme_raw_units(unit_map)
        self._workbook_schema(registry, assets)
        expected_counts = {
            "official_asset_count": 4,
            "result_unit_count": self.EXPECTED_RESULT_UNIT_COUNT,
            "pride_raw_unit_count": self.EXPECTED_RESULT_UNIT_COUNT,
            "result_to_raw_match_count": self.EXPECTED_RESULT_UNIT_COUNT,
            "explicit_raw_to_particle_map_count": 0,
            "material_record_count": self.EXPECTED_MATERIAL_COUNT,
        }
        self._require(
            report.get("audit_id") == self.AUDIT_ID
            and receipt.get("audit_id") == self.AUDIT_ID
            and report.get("status")
            == "VERIFIED_PUBLIC_ASSETS_INCOMPLETE_SOURCE_UNIT_TO_PARTICLE_MAP"
            and receipt.get("status") == report.get("status")
            and receipt.get("pxd017052_source_data_report_sha256") == _sha256(report_path)
            and report.get("registry_sha256") == _sha256(self.registry_path)
            and report.get("verified_raw_assets") == verified_assets
            and report.get("pride_readme_verification") == pride_readme
            and report.get("source_unit_to_material_map_status")
            == "INCOMPLETE_NO_EXPLICIT_RAW_TO_PARTICLE_CROSSWALK"
            and report.get("admission") == "NOT_ADMITTED"
            and receipt.get("admission") == "NOT_ADMITTED"
            and report.get("cc0_cohort_status") == "UNCHANGED"
            and receipt.get("cc0_cohort_status") == "UNCHANGED"
            and report.get("ccby_candidate_cohort_status") == "NOT_CREATED_INCOMPLETE_MAP"
            and receipt.get("ccby_candidate_cohort_status") == "NOT_CREATED_INCOMPLETE_MAP"
            and report.get("model_use") == "PROHIBITED"
            and receipt.get("model_use") == "PROHIBITED"
            and all(
                report.get(key) == value and receipt.get(key) == value
                for key, value in expected_counts.items()
            )
            and all(
                report.get(field) is False and receipt.get(field) is False
                for field in self.REQUIRED_FALSE
            ),
            "T131 source-data receipt is invalid",
        )
        return PXD017052SourceDataSummary(
            official_asset_count=4,
            result_unit_count=self.EXPECTED_RESULT_UNIT_COUNT,
            result_to_raw_match_count=self.EXPECTED_RESULT_UNIT_COUNT,
            explicit_raw_to_particle_map_count=0,
            receipt_path=receipt_path,
        )
