"""Fail-closed audit of real, openly licensed biointerface observations."""

from __future__ import annotations

import hashlib
import json
import math
import stat
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from biointerfaceos.evidence_semantics import (
    AllowedClaimLevel,
    EvidenceClass,
    EvidenceSemanticsError,
    require_metadata,
)


class EmpiricalProvenanceError(RuntimeError):
    """Raised when a claimed empirical observation cannot be traced to open raw data."""


def _canonical(value: Any) -> bytes:
    return (
        json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n"
    ).encode("utf-8")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise EmpiricalProvenanceError(f"{label} must be an object")
    return value


def _string(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise EmpiricalProvenanceError(f"{label} must be a non-empty string")
    return value.strip()


def _integer(value: Any, label: str, *, minimum: int = 1) -> int:
    if not isinstance(value, int) or isinstance(value, bool) or value < minimum:
        raise EmpiricalProvenanceError(f"{label} must be an integer >= {minimum}")
    return value


def _number(value: Any, label: str) -> float:
    if not isinstance(value, int | float) or isinstance(value, bool):
        raise EmpiricalProvenanceError(f"{label} must be numeric")
    result = float(value)
    if not math.isfinite(result):
        raise EmpiricalProvenanceError(f"{label} must be finite")
    return result


@dataclass(frozen=True)
class EmpiricalProvenanceSummary:
    """Compact accounting for a successful empirical-provenance audit."""

    registry_id: str
    source_count: int
    laboratory_count: int
    raw_asset_count: int
    observation_count: int
    receipt_path: Path


class EmpiricalProvenanceWorkflow:
    """Audit source licence, raw assets, and cell-level observation lineage offline."""

    AUDIT_ID = "bioif-empirical-provenance-audit-v1.1.0"
    AUDITED_AT = "2026-08-12T00:00:00+00:00"
    REGISTRY_RELATIVE = "data/empirical/R2_EMPIRICAL_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_2/empirical_provenance/v1.1.0"
    ALLOWED_LICENSES = frozenset({"CC-BY-4.0", "CC0-1.0", "PDDL-1.0"})
    REQUIRED_REGISTRY_FIELDS = {
        "schema_version",
        "registry_id",
        "evidence_class",
        "allowed_claim_level",
        "accessed_at",
        "sources",
        "excluded_sources",
    }
    REQUIRED_SOURCE_FIELDS = {
        "source_id",
        "doi",
        "title",
        "landing_url",
        "license_id",
        "access",
        "study_id",
        "laboratory",
        "affiliation",
        "material",
        "biological_system",
        "protocol_id",
        "protocol_description",
        "raw_assets",
        "rows",
    }
    REQUIRED_ASSET_FIELDS = {"path", "download_url", "sha256", "bytes", "content_type"}
    REQUIRED_ROW_FIELDS = {
        "observation_id",
        "raw_asset",
        "worksheet",
        "row_number",
        "value_column",
        "independent_unit_id",
        "observed_value",
        "endpoint_id",
        "endpoint_name",
        "unit",
    }

    def __init__(
        self,
        root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    def _load_registry(self) -> dict[str, Any]:
        try:
            registry = _mapping(
                json.loads(self.registry_path.read_text(encoding="utf-8")), "empirical registry"
            )
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise EmpiricalProvenanceError("cannot parse empirical registry") from exc
        if set(registry) != self.REQUIRED_REGISTRY_FIELDS or registry.get("schema_version") != 1:
            raise EmpiricalProvenanceError("empirical registry fields or schema are invalid")
        _string(registry.get("registry_id"), "empirical registry ID")
        _string(registry.get("accessed_at"), "empirical registry access time")
        try:
            evidence_class, claim_level = require_metadata(registry, "empirical registry")
        except EvidenceSemanticsError as exc:
            raise EmpiricalProvenanceError(str(exc)) from exc
        if (
            evidence_class is not EvidenceClass.DEVELOPMENT_OBSERVATION
            or claim_level is not AllowedClaimLevel.EXPLORATORY
        ):
            raise EmpiricalProvenanceError("empirical registry evidence class is unsafe")
        if not isinstance(registry["sources"], list) or not registry["sources"]:
            raise EmpiricalProvenanceError("empirical registry must contain an admitted source")
        if not isinstance(registry["excluded_sources"], list):
            raise EmpiricalProvenanceError("empirical registry exclusions are invalid")
        return registry

    def _source_asset(self, source: dict[str, Any], value: Any) -> dict[str, Any]:
        asset = _mapping(value, "empirical raw asset")
        if set(asset) != self.REQUIRED_ASSET_FIELDS:
            raise EmpiricalProvenanceError("empirical raw asset fields are invalid")
        relative = _string(asset.get("path"), "raw asset path")
        path = (self.root / relative).resolve(strict=False)
        if not path.is_relative_to(self.root) or not path.is_file():
            raise EmpiricalProvenanceError(f"raw asset is missing: {relative}")
        if "fixture" in relative.lower() or "synthetic" in relative.lower():
            raise EmpiricalProvenanceError(f"non-empirical raw asset path: {relative}")
        expected_hash = _string(asset.get("sha256"), "raw asset SHA-256").lower()
        if len(expected_hash) != 64 or any(
            character not in "0123456789abcdef" for character in expected_hash
        ):
            raise EmpiricalProvenanceError("raw asset SHA-256 is invalid")
        if _sha256(path) != expected_hash:
            raise EmpiricalProvenanceError(f"raw asset checksum differs: {relative}")
        if path.stat().st_size != _integer(asset.get("bytes"), "raw asset bytes"):
            raise EmpiricalProvenanceError(f"raw asset byte count differs: {relative}")
        download_url = _string(asset.get("download_url"), "raw asset download URL")
        if not download_url.startswith("https://"):
            raise EmpiricalProvenanceError("raw asset must have an HTTPS source")
        if _string(asset.get("content_type"), "raw asset content type") != (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            raise EmpiricalProvenanceError("raw asset content type is not an XLSX workbook")
        return {
            "source_id": source["source_id"],
            "path": relative,
            "sha256": expected_hash,
            "bytes": path.stat().st_size,
            "download_url": download_url,
            "content_type": asset["content_type"],
        }

    @staticmethod
    def _column_index(column: str) -> int:
        value = _string(column, "raw value column")
        if len(value) != 1 or not value.isalpha() or not value.isupper():
            raise EmpiricalProvenanceError("raw value column must be a single uppercase letter")
        return ord(value) - ord("A") + 1

    def _source_rows(
        self, source: dict[str, Any], assets: dict[str, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        rows_value = source["rows"]
        if not isinstance(rows_value, list) or not rows_value:
            raise EmpiricalProvenanceError(f"{source['source_id']} has no observation rows")
        source_text = " ".join(
            str(source[field]) for field in ("source_id", "title", "study_id", "laboratory")
        ).lower()
        if any(token in source_text for token in ("fixture", "synthetic", "mock")):
            raise EmpiricalProvenanceError(f"{source['source_id']} is not a real-source record")
        result: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        for value in rows_value:
            row = _mapping(value, "empirical observation row")
            if set(row) != self.REQUIRED_ROW_FIELDS:
                raise EmpiricalProvenanceError("empirical observation row fields are invalid")
            observation_id = _string(row.get("observation_id"), "observation ID")
            if observation_id in seen_ids:
                raise EmpiricalProvenanceError(f"duplicate observation ID: {observation_id}")
            seen_ids.add(observation_id)
            raw_asset = _string(row.get("raw_asset"), "observation raw asset")
            if raw_asset not in assets:
                raise EmpiricalProvenanceError(
                    f"observation uses an unregistered raw asset: {raw_asset}"
                )
            worksheet = _string(row.get("worksheet"), "observation worksheet")
            row_number = _integer(row.get("row_number"), "observation row number", minimum=2)
            value_column = _string(row.get("value_column"), "observation value column")
            expected_value = _number(row.get("observed_value"), "observation value")
            path = self.root / raw_asset
            workbook: Any | None = None
            try:
                workbook = load_workbook(path, data_only=True, read_only=True)
                sheet = workbook[worksheet]
                observed_unit = sheet.cell(row=row_number, column=1).value
                observed_value = sheet.cell(
                    row=row_number, column=self._column_index(value_column)
                ).value
            except (KeyError, OSError, ValueError) as exc:
                raise EmpiricalProvenanceError(
                    f"cannot locate raw observation {observation_id}"
                ) from exc
            finally:
                if workbook is not None:
                    workbook.close()
            actual_value = _number(observed_value, f"raw value for {observation_id}")
            if not math.isclose(actual_value, expected_value, rel_tol=0.0, abs_tol=1e-12):
                raise EmpiricalProvenanceError(
                    f"raw value differs from registered observation: {observation_id}"
                )
            independent_unit_id = _string(row.get("independent_unit_id"), "independent unit ID")
            if (
                _string(observed_unit, f"raw independent unit for {observation_id}")
                != independent_unit_id
            ):
                raise EmpiricalProvenanceError(
                    f"raw independent unit differs from registry: {observation_id}"
                )
            result.append(
                {
                    "observation_id": observation_id,
                    "source_id": source["source_id"],
                    "study_id": source["study_id"],
                    "laboratory": source["laboratory"],
                    "affiliation": source["affiliation"],
                    "doi": source["doi"],
                    "landing_url": source["landing_url"],
                    "license_id": source["license_id"],
                    "material": source["material"],
                    "biological_system": source["biological_system"],
                    "protocol_id": source["protocol_id"],
                    "protocol_description": source["protocol_description"],
                    "endpoint_id": _string(row.get("endpoint_id"), "endpoint ID"),
                    "endpoint_name": _string(row.get("endpoint_name"), "endpoint name"),
                    "independent_unit_id": independent_unit_id,
                    "raw_asset": raw_asset,
                    "raw_asset_sha256": assets[raw_asset]["sha256"],
                    "raw_locator": f"{worksheet}!{value_column}{row_number}",
                    "raw_value": actual_value,
                    "unit": _string(row.get("unit"), "observation unit"),
                }
            )
        return result

    def _validate_source(self, value: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        source = _mapping(value, "empirical source")
        if set(source) != self.REQUIRED_SOURCE_FIELDS:
            raise EmpiricalProvenanceError("empirical source fields are invalid")
        for field in self.REQUIRED_SOURCE_FIELDS - {"raw_assets", "rows"}:
            _string(source.get(field), f"empirical source {field}")
        if source["license_id"] not in self.ALLOWED_LICENSES:
            raise EmpiricalProvenanceError(f"source license is not reusable: {source['source_id']}")
        if source["access"] != "ANONYMOUS_PUBLIC":
            raise EmpiricalProvenanceError(
                f"source is not anonymously accessible: {source['source_id']}"
            )
        if not source["landing_url"].startswith("https://") or not source["doi"].startswith("10."):
            raise EmpiricalProvenanceError(
                f"source DOI or landing URL is invalid: {source['source_id']}"
            )
        assets_value = source["raw_assets"]
        if not isinstance(assets_value, list) or not assets_value:
            raise EmpiricalProvenanceError(f"{source['source_id']} has no raw assets")
        assets = [self._source_asset(source, asset) for asset in assets_value]
        paths = [str(asset["path"]) for asset in assets]
        if len(paths) != len(set(paths)):
            raise EmpiricalProvenanceError(f"duplicate raw asset in {source['source_id']}")
        asset_lookup = {str(asset["path"]): asset for asset in assets}
        return assets, self._source_rows(source, asset_lookup)

    def _audit_payload(self) -> dict[str, Any]:
        registry = self._load_registry()
        raw_assets: list[dict[str, Any]] = []
        rows: list[dict[str, Any]] = []
        laboratories: set[str] = set()
        source_ids: set[str] = set()
        for source_value in registry["sources"]:
            source = _mapping(source_value, "empirical source")
            source_id = _string(source.get("source_id"), "empirical source ID")
            if source_id in source_ids:
                raise EmpiricalProvenanceError(f"duplicate empirical source ID: {source_id}")
            source_ids.add(source_id)
            assets, source_rows = self._validate_source(source)
            raw_assets.extend(assets)
            rows.extend(source_rows)
            laboratories.add(_string(source.get("laboratory"), "empirical laboratory"))
        if not rows or not raw_assets:
            raise EmpiricalProvenanceError("no real observations were admitted")
        observation_ids = [str(row["observation_id"]) for row in rows]
        if len(observation_ids) != len(set(observation_ids)):
            raise EmpiricalProvenanceError("duplicate observation IDs across sources")
        return {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "audited_at": self.AUDITED_AT,
            "registry_id": registry["registry_id"],
            "evidence_class": EvidenceClass.DEVELOPMENT_OBSERVATION.value,
            "allowed_claim_level": AllowedClaimLevel.EXPLORATORY.value,
            "status": "PASS_EMPIRICAL_PROVENANCE",
            "source_count": len(source_ids),
            "laboratory_count": len(laboratories),
            "raw_asset_count": len(raw_assets),
            "observation_count": len(rows),
            "raw_asset_manifest": sorted(raw_assets, key=lambda asset: str(asset["path"])),
            "row_provenance": rows,
            "excluded_sources": registry["excluded_sources"],
            "empirical_source": True,
            "statistical_conclusions": False,
            "independent_validation": False,
            "scientific_reproduction": False,
            "scientific_submission_ready": False,
        }

    def run(self, *, strict: bool = False) -> EmpiricalProvenanceSummary:
        """Create one immutable audit record; an admission audit is not an outcome analysis."""

        if not strict:
            raise EmpiricalProvenanceError("T120 requires --strict")
        if self.output_root.exists():
            raise EmpiricalProvenanceError("empirical-provenance audit already executed")
        audit = self._audit_payload()
        self.output_root.mkdir(parents=True, exist_ok=False)
        audit_path = self.output_root / "empirical_provenance_audit.json"
        self._write(audit_path, audit)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": audit["status"],
            "registry_id": audit["registry_id"],
            "audit_sha256": _sha256(audit_path),
            "source_count": audit["source_count"],
            "laboratory_count": audit["laboratory_count"],
            "raw_asset_count": audit["raw_asset_count"],
            "observation_count": audit["observation_count"],
            "empirical_source": True,
            "statistical_conclusions": False,
            "independent_validation": False,
            "scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "audit_receipt.json"
        self._write(receipt_path, receipt)
        for path in self.output_root.iterdir():
            path.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        self.output_root.chmod(
            stat.S_IRUSR | stat.S_IXUSR | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH
        )
        return EmpiricalProvenanceSummary(
            registry_id=str(audit["registry_id"]),
            source_count=int(audit["source_count"]),
            laboratory_count=int(audit["laboratory_count"]),
            raw_asset_count=int(audit["raw_asset_count"]),
            observation_count=int(audit["observation_count"]),
            receipt_path=receipt_path,
        )

    def verify(self) -> dict[str, Any]:
        """Verify the immutable audit receipt without reading a protected result namespace."""

        audit_path = self.output_root / "empirical_provenance_audit.json"
        receipt_path = self.output_root / "audit_receipt.json"
        try:
            audit = _mapping(json.loads(audit_path.read_text(encoding="utf-8")), "empirical audit")
            receipt = _mapping(
                json.loads(receipt_path.read_text(encoding="utf-8")), "empirical audit receipt"
            )
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise EmpiricalProvenanceError("empirical-provenance audit output is invalid") from exc
        if (
            receipt.get("audit_id") != self.AUDIT_ID
            or receipt.get("status") != "PASS_EMPIRICAL_PROVENANCE"
            or receipt.get("audit_sha256") != _sha256(audit_path)
            or receipt.get("registry_id") != audit.get("registry_id")
            or receipt.get("observation_count") != audit.get("observation_count")
            or receipt.get("empirical_source") is not True
            or receipt.get("statistical_conclusions") is not False
            or receipt.get("independent_validation") is not False
            or receipt.get("scientific_submission_ready") is not False
        ):
            raise EmpiricalProvenanceError("empirical-provenance audit receipt is invalid")
        return audit
