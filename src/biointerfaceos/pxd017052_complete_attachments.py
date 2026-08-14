"""Correct the T131 source-map scope from the complete publisher attachment set."""

from __future__ import annotations

import hashlib
import json
import stat
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PXD017052CompleteAttachmentsError(RuntimeError):
    """Raised when a T132 correction is incomplete or permits unsafe promotion."""


@dataclass(frozen=True)
class PXD017052CompleteAttachmentsSummary:
    asset_count: int
    unit_map_count: int
    receipt_path: Path


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _json(path: Path, label: str) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise PXD017052CompleteAttachmentsError(f"cannot parse {label}") from exc
    if not isinstance(value, dict):
        raise PXD017052CompleteAttachmentsError(f"{label} must be an object")
    return value


class PXD017052CompleteAttachmentsWorkflow:
    AUDIT_ID = "bioif-r2-pxd017052-complete-attachments-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R2_T132_PXD017052_COMPLETE_ATTACHMENT_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_2/pxd017052_complete_attachments/v1.0.0"
    FALSE_FIELDS = (
        "target_frozen",
        "model_fitted",
        "paired_ablations_run",
        "external_ood_evaluated",
        "negative_controls_run",
        "independent_validation",
        "scientific_submission_ready",
    )

    def __init__(self, root: Path) -> None:
        self.root = root.resolve(strict=True)
        self.registry_path = self.root / self.REGISTRY_RELATIVE
        self.output_root = self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _require(condition: bool, label: str) -> None:
        if not condition:
            raise PXD017052CompleteAttachmentsError(label)

    def _registry(self) -> tuple[dict[str, Any], Path, list[dict[str, Any]]]:
        registry = _json(self.registry_path, "T132 registry")
        self._require(
            registry.get("schema_version") == 1 and registry.get("audit_id") == self.AUDIT_ID,
            "T132 registry identity is invalid",
        )
        raw_root = self.root / str(registry.get("raw_directory"))
        self._require(raw_root.is_dir(), "T132 protected raw directory is missing")
        assets = registry.get("assets")
        if not isinstance(assets, list) or len(assets) != 8:
            raise PXD017052CompleteAttachmentsError("T132 asset inventory is invalid")
        names: set[str] = set()
        for asset in assets:
            self._require(
                isinstance(asset, dict)
                and set(asset) == {"file_name", "bytes", "sha256", "md5", "publisher_etag", "role"},
                "T132 asset fields are invalid",
            )
            name = asset.get("file_name")
            self._require(
                isinstance(name, str) and name not in names and isinstance(asset.get("bytes"), int),
                "T132 asset identity is invalid",
            )
            names.add(name)
        self._require("41467_2020_17033_MOESM8_ESM.xlsx" in names, "T132 unit-map asset is missing")
        unit_map = registry.get("unit_map")
        if not isinstance(unit_map, list) or len(unit_map) != 9:
            raise PXD017052CompleteAttachmentsError("T132 unit map is invalid")
        if not all(
            isinstance(row, dict) and set(row) == {"particle", "result_unit_id", "assay_replicate"} for row in unit_map
        ):
            raise PXD017052CompleteAttachmentsError("T132 unit-map fields are invalid")
        self._require(
            registry.get("decision")
            == {
                "status": "VERIFIED_COMPLETE_UNIT_TO_PARTICLE_MAP_SINGLE_LAB_CCBY",
                "admission": "NOT_ADMITTED_PENDING_CCBY_AMENDMENT_AND_SECOND_LAB",
                "cc0_cohort_status": "UNCHANGED",
                "model_use": "PROHIBITED",
            },
            "T132 decision boundary is invalid",
        )
        return registry, raw_root, [row for row in unit_map if isinstance(row, dict)]

    def _verify_assets(self, registry: dict[str, Any], raw_root: Path) -> list[dict[str, Any]]:
        verified: list[dict[str, Any]] = []
        for asset in registry["assets"]:
            path = raw_root / asset["file_name"]
            self._require(
                path.is_file()
                and path.stat().st_size == asset["bytes"]
                and _sha256(path) == asset["sha256"]
                and hashlib.md5(path.read_bytes(), usedforsecurity=False).hexdigest()
                == asset["md5"]
                == asset["publisher_etag"],
                f"T132 asset differs: {asset['file_name']}",
            )
            verified.append({key: asset[key] for key in ("file_name", "bytes", "sha256", "role")})
        return verified

    def _verify_map(self, registry: dict[str, Any], raw_root: Path, unit_map: list[dict[str, Any]]) -> None:
        base = _json(self.root / registry["base_t131_registry"], "T131 registry")
        base_units = base["particle_unit_map"]["result_unit_ids"]
        self._require(
            {row["result_unit_id"] for row in unit_map} == set(base_units),
            "T132 units do not close T131 result units",
        )
        self._require(
            len({row["result_unit_id"] for row in unit_map}) == 9
            and {(row["particle"], row["assay_replicate"]) for row in unit_map}
            == {
                (particle, replicate)
                for particle in ("SP-003-001", "SP-007-002", "SP-011-001")
                for replicate in (1, 2, 3)
            },
            "T132 particle replicate coverage is invalid",
        )
        worksheet = load_workbook(raw_root / "41467_2020_17033_MOESM8_ESM.xlsx", read_only=True, data_only=True).active
        self._require(
            worksheet.title == "Sheet1"
            and worksheet.max_row == 3266
            and worksheet.max_column == 18
            and worksheet.cell(1, 1).value == "Supplamentary Data 6.   Sample Identifier"
            and worksheet.cell(2, 1).value == "3 NPs (Figure3)"
            and [worksheet.cell(3, column).value for column in (1, 2, 3)] == ["nanoparticles", "File", "rep"],
            "T132 unit-map worksheet schema differs",
        )
        observed = [
            {
                "particle": worksheet.cell(row, 1).value,
                "result_unit_id": worksheet.cell(row, 2).value,
                "assay_replicate": worksheet.cell(row, 3).value,
            }
            for row in range(4, 13)
        ]
        self._require(observed == unit_map, "T132 explicit unit-to-particle map differs")

    def run(self, *, strict: bool = False) -> PXD017052CompleteAttachmentsSummary:
        if not strict:
            raise PXD017052CompleteAttachmentsError("T132 audit requires --strict")
        if self.output_root.exists():
            raise PXD017052CompleteAttachmentsError("T132 audit already executed")
        registry, raw_root, unit_map = self._registry()
        assets = self._verify_assets(registry, raw_root)
        self._verify_map(registry, raw_root, unit_map)
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": registry["decision"]["status"],
            "registry_sha256": _sha256(self.registry_path),
            "base_t131_receipt_sha256": _sha256(self.root / registry["base_t131_receipt"]),
            "verified_extension_assets": assets,
            "extension_asset_count": len(assets),
            "explicit_unit_to_particle_map": unit_map,
            "explicit_unit_to_particle_map_count": len(unit_map),
            **registry["decision"],
            **{field: False for field in self.FALSE_FIELDS},
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        report_path = self.output_root / "complete_attachment_report.json"
        report_path.write_text(json.dumps(report, sort_keys=True, separators=(",", ":")) + "\n", encoding="utf-8")
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": report["status"],
            "complete_attachment_report_sha256": _sha256(report_path),
            "extension_asset_count": 8,
            "explicit_unit_to_particle_map_count": 9,
            **registry["decision"],
            **{field: False for field in self.FALSE_FIELDS},
        }
        receipt_path = self.output_root / "complete_attachment_receipt.json"
        receipt_path.write_text(json.dumps(receipt, sort_keys=True, separators=(",", ":")) + "\n", encoding="utf-8")
        for path in self.output_root.iterdir():
            path.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        self.output_root.chmod(stat.S_IRUSR | stat.S_IXUSR | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH)
        return PXD017052CompleteAttachmentsSummary(8, 9, receipt_path)
