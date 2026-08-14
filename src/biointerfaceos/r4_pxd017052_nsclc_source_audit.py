"""Audit the 141-subject NSCLC Supplementary Data 5 source.

The source is a public, paper-attached matrix from the Proteograph study.  It
contains five nanoparticle corona conditions and a depleted-plasma control for
each of 141 individual plasma samples.  This audit keeps the cohort separate
from the original nine-assay PXD017052 source map, preserves ``NA`` values,
and admits only the 34 identifiers that resolve uniquely into the frozen R3
target universe.
"""

from __future__ import annotations

import csv
import json
import math
from collections import Counter
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R4PXD017052NSCLCSourceAuditError(RuntimeError):
    """Raised when the 141-subject source contract is not reproducible."""


@dataclass(frozen=True)
class R4PXD017052NSCLCSourceAuditSummary:
    """Accounting for the paper-attached biological cohort."""

    source_asset_count: int
    protein_row_count: int
    biological_unit_count: int
    measurement_batch_count: int
    rank_qualified_measurement_batch_count: int
    shared_canonical_protein_count: int
    source_cell_count: int
    positive_source_cell_count: int
    receipt_path: Path


class R4PXD017052NSCLCSourceAuditWorkflow:
    """Create a traceable NP-corona cell map for the 141 individual samples."""

    AUDIT_ID = "bioif-r4-pxd017052-nsclc-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R4_T180_PXD017052_NSCLC_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_4/pxd017052_nsclc_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R4_PXD017052_NSCLC_source_cell_map.csv"
    STATUS = "ADMITTED_R4_BIOLOGICAL_COHORT_SOURCE_EXPLORATORY_ONLY"
    SOURCE_CELL_FIELDS = (
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "canonical_accession",
        "biological_unit_id",
        "clinical_group",
        "particle",
        "measurement_role",
        "measurement_batch_id",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    )
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "article",
        "source_scope",
        "source_assets",
        "reference_assets",
        "worksheet_contract",
        "quantification_contract",
        "admission_minimums",
        "claim_boundary",
    }

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError, TypeError) as exc:
            raise R4PXD017052NSCLCSourceAuditError(f"cannot parse {label}") from exc

    @staticmethod
    def _number(value: Any) -> float | None:
        if value is None or value == "":
            return None
        if isinstance(value, str) and value.strip().upper() == "NA":
            return None
        if isinstance(value, bool) or not isinstance(value, int | float):
            raise R4PXD017052NSCLCSourceAuditError("source abundance must be numeric, NA or blank")
        number = float(value)
        if not math.isfinite(number) or number < 0:
            raise R4PXD017052NSCLCSourceAuditError("source abundance must be finite and non-negative")
        return number

    def _under(self, base: Path, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise R4PXD017052NSCLCSourceAuditError(f"{label} must use POSIX path separators")
        pure = PurePosixPath(relative_path)
        path = (base / Path(*pure.parts)).resolve(strict=False)
        if (
            pure.is_absolute()
            or not pure.parts
            or ".." in pure.parts
            or not path.is_relative_to(base)
            or not path.is_file()
        ):
            raise R4PXD017052NSCLCSourceAuditError(f"{label} is missing or escapes its root")
        return path

    def _checked_assets(self, registry: dict[str, Any]) -> tuple[Path, dict[str, Path]]:
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise R4PXD017052NSCLCSourceAuditError("registry fields are invalid")
        if (
            registry.get("audit_id") != self.AUDIT_ID
            or registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise R4PXD017052NSCLCSourceAuditError("registry identity or evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "evaluated_at")
        article = _mapping(registry.get("article"), "article")
        if article != {
            "pmcid": "PMC7376165",
            "doi": "10.1038/s41467-020-17033-7",
            "title": "Rapid, deep and precise profiling of the plasma proteome with multi-nanoparticle protein corona",
            "publication_year": 2020,
            "license": "CC-BY-4.0",
            "full_text_locator": "https://europepmc.org/articles/PMC7376165",
        }:
            raise R4PXD017052NSCLCSourceAuditError("article declaration is invalid")
        scope = _mapping(registry.get("source_scope"), "source scope")
        if scope != {
            "source_id": "PXD017052_SEER_BROAD_NSCLC_COHORT",
            "laboratory_anchor": "Seer, Inc. / Broad Institute of MIT and Harvard",
            "biofluid": "individual human plasma samples",
            "measurement_design": "five nanoparticle corona conditions plus depleted-plasma control per subject",
            "analysis_role": "SEPARATE_R4_BIOLOGICAL_COHORT_OOD_CANDIDATE",
            "prohibited_interpretations": [
                "not an independent laboratory anchor",
                "not a protected lockbox evaluator",
                "not a no-author scientific reproduction",
                "not a clinical diagnostic validation",
            ],
        }:
            raise R4PXD017052NSCLCSourceAuditError("source scope is invalid")
        assets = registry.get("source_assets")
        if not isinstance(assets, list) or len(assets) != 1:
            raise R4PXD017052NSCLCSourceAuditError("source assets are invalid")
        asset_paths: dict[str, Path] = {}
        for item in assets:
            item = _mapping(item, "source asset")
            if set(item) != {"asset_id", "relative_path", "sha256", "expected_bytes"}:
                raise R4PXD017052NSCLCSourceAuditError("source asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "source asset ID")
            path = self._under(self.assets_root, _string(item.get("relative_path"), asset_id), asset_id)
            if (
                asset_id in asset_paths
                or path.stat().st_size != item.get("expected_bytes")
                or _sha256(path) != _checksum(item.get("sha256"), asset_id)
            ):
                raise R4PXD017052NSCLCSourceAuditError("source asset checksum differs")
            asset_paths[asset_id] = path
        if set(asset_paths) != {"supplementary_data_5_nsclc"}:
            raise R4PXD017052NSCLCSourceAuditError("source asset roster is invalid")
        refs = registry.get("reference_assets")
        if not isinstance(refs, list) or len(refs) != 3:
            raise R4PXD017052NSCLCSourceAuditError("reference assets are invalid")
        ref_paths: dict[str, Path] = {}
        for item in refs:
            item = _mapping(item, "reference asset")
            if set(item) != {"asset_id", "relative_path", "sha256"}:
                raise R4PXD017052NSCLCSourceAuditError("reference asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "reference asset ID")
            path = self._under(self.root, _string(item.get("relative_path"), asset_id), asset_id)
            if asset_id in ref_paths or _sha256(path) != _checksum(item.get("sha256"), asset_id):
                raise R4PXD017052NSCLCSourceAuditError("reference asset checksum differs")
            ref_paths[asset_id] = path
        if set(ref_paths) != {
            "r3_common_target_ledger",
            "r3_feature_table",
            "source_identifier_resolution",
        }:
            raise R4PXD017052NSCLCSourceAuditError("reference asset roster is incomplete")
        if registry.get("worksheet_contract") != {
            "worksheet": "Sheet1",
            "expected_rows": 2586,
            "expected_columns": 847,
            "header_row": 2,
            "protein_row_start": 3,
            "subject_count": 141,
            "nanoparticle_condition_count": 5,
            "depleted_plasma_control_count": 141,
        }:
            raise R4PXD017052NSCLCSourceAuditError("worksheet contract differs")
        if registry.get("quantification_contract") != {
            "source_protein_column": "prot_group",
            "source_identifier_resolution": "exact source_identifier_resolution row with UNIQUE_HUMAN_CANONICAL_ACCESSION; no ambiguous group splitting",  # noqa: E501
            "rank_eligibility": "strictly positive finite author-reported value within one subject-particle measurement batch",  # noqa: E501
            "na_policy": "retain NA as AUTHOR_NA and exclude from rank; no imputation",
            "zero_policy": "retain explicit numeric zero as AUTHOR_EXPLICIT_ZERO and exclude from rank",
            "depleted_plasma_policy": "retain as provenance control but exclude from NP-corona analysis ledger",
            "raw_scale_cross_study_use": "PROHIBITED",
        }:
            raise R4PXD017052NSCLCSourceAuditError("quantification contract differs")
        if registry.get("admission_minimums") != {
            "biological_unit_count": 141,
            "measurement_batch_count": 705,
            "rank_qualified_measurement_batch_count": 600,
            "shared_canonical_protein_count": 30,
            "minimum_positive_shared_proteins_per_measurement_batch": 10,
        }:
            raise R4PXD017052NSCLCSourceAuditError("admission minimums differ")
        return asset_paths["supplementary_data_5_nsclc"], ref_paths

    @staticmethod
    def _mapping_tables(ledger_path: Path, resolution_path: Path) -> tuple[set[str], dict[str, str]]:
        with ledger_path.open(newline="", encoding="utf-8") as stream:
            ledger = list(csv.DictReader(stream))
        targets = {row["canonical_accession"] for row in ledger if row.get("common_rank_target_member") == "true"}
        if len(targets) != 99:
            raise R4PXD017052NSCLCSourceAuditError("frozen R3 target ledger differs")
        with resolution_path.open(newline="", encoding="utf-8") as stream:
            resolution = list(csv.DictReader(stream))
        mapping: dict[str, str] = {}
        for row in resolution:
            if (
                row.get("source_id") != "PXD017052_SEER_BROAD"
                or row.get("resolution_status") != "UNIQUE_HUMAN_CANONICAL_ACCESSION"
            ):
                continue
            accession = row.get("resolved_canonical_accession", "")
            if accession in targets:
                if row["source_identifier"] in mapping and mapping[row["source_identifier"]] != accession:
                    raise R4PXD017052NSCLCSourceAuditError("source identifier has conflicting mappings")
                mapping[row["source_identifier"]] = accession
        if not mapping:
            raise R4PXD017052NSCLCSourceAuditError("source identifier mapping is empty")
        return targets, mapping

    def _cells(self, workbook_path: Path, references: dict[str, Path]) -> tuple[list[dict[str, Any]], dict[str, int]]:
        _, identifier_map = self._mapping_tables(
            references["r3_common_target_ledger"], references["source_identifier_resolution"]
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if workbook.sheetnames != ["Sheet1"]:
            raise R4PXD017052NSCLCSourceAuditError("workbook sheets differ")
        sheet = workbook["Sheet1"]
        if (sheet.max_row, sheet.max_column) != (2586, 847):
            raise R4PXD017052NSCLCSourceAuditError("workbook dimensions differ")
        header = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
        if header[0] != "prot_group" or any(not isinstance(value, str) or not value for value in header[1:]):
            raise R4PXD017052NSCLCSourceAuditError("workbook header differs")
        target_rows: dict[str, tuple[int, str]] = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=3, max_col=1, values_only=True), start=3):
            identifier = str(row[0] or "")
            accession = identifier_map.get(identifier)
            if accession is None:
                continue
            if accession in target_rows:
                raise R4PXD017052NSCLCSourceAuditError("multiple source rows collapse to one target")
            target_rows[accession] = (row_number, identifier)
        if len(target_rows) != 34:
            raise R4PXD017052NSCLCSourceAuditError("NSCLC cohort target coverage differs")
        cells: list[dict[str, Any]] = []
        batch_counts: Counter[str] = Counter()
        positive_counts: Counter[str] = Counter()
        subject_ids: set[str] = set()
        protein_rows = sheet.max_row - 2
        for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            identifier = str(row[0] or "")
            accession = identifier_map.get(identifier)
            if accession not in target_rows or target_rows[accession][0] != row_number:
                continue
            for column_index, sample_label in enumerate(header[1:], start=1):
                subject_label, particle = sample_label.rsplit("_", 1)
                if particle == "DP":
                    continue
                biological_unit_id, clinical_group = subject_label.split("_", 1)
                value = self._number(row[column_index])
                state = "AUTHOR_NA" if value is None else ("AUTHOR_EXPLICIT_ZERO" if value == 0 else "POSITIVE_FINITE")
                batch_id = f"R4_PXD017052_NSCLC:{subject_label}:{particle}"
                subject_ids.add(biological_unit_id)
                batch_counts[batch_id] += 1
                if state == "POSITIVE_FINITE":
                    positive_counts[batch_id] += 1
                cells.append(
                    {
                        "source_id": "PXD017052_SEER_BROAD_NSCLC_COHORT",
                        "laboratory_anchor": "Seer, Inc. / Broad Institute of MIT and Harvard",
                        "source_asset_id": "supplementary_data_5_nsclc",
                        "source_worksheet": "Sheet1",
                        "source_row": str(row_number),
                        "source_coordinate": f"{get_column_letter(column_index + 1)}{row_number}",
                        "source_identifier": identifier,
                        "canonical_accession": accession,
                        "biological_unit_id": biological_unit_id,
                        "clinical_group": clinical_group,
                        "particle": particle,
                        "measurement_role": "NP_CORONA",
                        "measurement_batch_id": batch_id,
                        "author_quantity_type": "LOG2_MEDIAN_NORMALIZED_PROTEIN_GROUP_INTENSITY",
                        "author_numeric_value": "" if value is None else format(value, ".17g"),
                        "author_value_state": state,
                        "rank_target_eligible": "true" if state == "POSITIVE_FINITE" else "false",
                    }
                )
        qualified = {batch for batch, count in positive_counts.items() if count >= 10}
        if (
            len(subject_ids) != 141
            or len(batch_counts) != 705
            or len(qualified) != 666
            or len({row["canonical_accession"] for row in cells}) != 34
            or len(cells) != 23970
            or sum(row["rank_target_eligible"] == "true" for row in cells) != 17330
        ):
            raise R4PXD017052NSCLCSourceAuditError("source-cell accounting differs")
        return cells, {
            "protein_rows": protein_rows,
            "biological_units": len(subject_ids),
            "measurement_batches": len(batch_counts),
            "qualified_batches": len(qualified),
            "shared_targets": len({row["canonical_accession"] for row in cells}),
            "source_cells": len(cells),
            "positive_cells": sum(row["rank_target_eligible"] == "true" for row in cells),
            "na_cells": sum(row["author_value_state"] == "AUTHOR_NA" for row in cells),
        }

    def run(self, *, strict: bool = False) -> R4PXD017052NSCLCSourceAuditSummary:
        if not strict:
            raise R4PXD017052NSCLCSourceAuditError("R4 PXD017052 NSCLC source audit requires --strict")
        if self.output_root.exists():
            raise R4PXD017052NSCLCSourceAuditError("R4 PXD017052 NSCLC source audit already executed")
        registry = self._json(self.registry_path, "R4 PXD017052 NSCLC registry")
        workbook_path, references = self._checked_assets(registry)
        cells, totals = self._cells(workbook_path, references)
        derived = self.assets_root / self.DERIVED_RELATIVE
        if derived.exists():
            raise R4PXD017052NSCLCSourceAuditError("derived source cell map already exists")
        derived.parent.mkdir(parents=True, exist_ok=True)
        with derived.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=self.SOURCE_CELL_FIELDS)
            writer.writeheader()
            writer.writerows(cells)
        self.output_root.mkdir(parents=True, exist_ok=False)
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "source_asset": {
                item["asset_id"]: {
                    "relative_path": item["relative_path"],
                    "sha256": _sha256(workbook_path),
                }
                for item in registry["source_assets"]
            },
            "reference_assets": {
                asset_id: {
                    "relative_path": path.relative_to(self.root).as_posix(),
                    "sha256": _sha256(path),
                }
                for asset_id, path in references.items()
            },
            "source_cell_map": {"relative_path": self.DERIVED_RELATIVE, "sha256": _sha256(derived)},
            **totals,
            "claim_boundary": registry["claim_boundary"],
        }
        report_path = self.output_root / "pxd017052_nsclc_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report": {"relative_path": report_path.name, "sha256": _sha256(report_path)},
            "source_cell_map": report["source_cell_map"],
            "biological_unit_count": totals["biological_units"],
            "measurement_batch_count": totals["measurement_batches"],
            "rank_qualified_measurement_batch_count": totals["qualified_batches"],
            "shared_canonical_protein_count": totals["shared_targets"],
            "source_cell_count": totals["source_cells"],
            "positive_source_cell_count": totals["positive_cells"],
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "pxd017052_nsclc_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R4PXD017052NSCLCSourceAuditSummary(
            len(registry["source_assets"]),
            totals["protein_rows"],
            totals["biological_units"],
            totals["measurement_batches"],
            totals["qualified_batches"],
            totals["shared_targets"],
            totals["source_cells"],
            totals["positive_cells"],
            receipt_path,
        )

    def verify(self) -> R4PXD017052NSCLCSourceAuditSummary:
        registry = self._json(self.registry_path, "R4 PXD017052 NSCLC registry")
        workbook_path, references = self._checked_assets(registry)
        report_path = self.output_root / "pxd017052_nsclc_source_audit_report.json"
        receipt_path = self.output_root / "pxd017052_nsclc_source_audit_receipt.json"
        report = self._json(report_path, "R4 PXD017052 NSCLC source audit report")
        receipt = self._json(receipt_path, "R4 PXD017052 NSCLC source audit receipt")
        if (
            report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report", {}).get("sha256") != _sha256(report_path)
        ):
            raise R4PXD017052NSCLCSourceAuditError("audit receipt differs")
        cell_map = self._under(
            self.assets_root,
            _string(report.get("source_cell_map", {}).get("relative_path"), "source cell map path"),
            "source cell map",
        )
        if report["source_cell_map"].get("sha256") != _sha256(cell_map):
            raise R4PXD017052NSCLCSourceAuditError("source cell map checksum differs")
        with cell_map.open(newline="", encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        if (set(rows[0]) if rows else set()) != set(self.SOURCE_CELL_FIELDS):
            raise R4PXD017052NSCLCSourceAuditError("source cell map fields differ")
        expected = (
            int(report["protein_rows"]),
            int(report["biological_units"]),
            int(report["measurement_batches"]),
            int(report["qualified_batches"]),
            int(report["shared_targets"]),
            int(report["source_cells"]),
            int(report["positive_cells"]),
        )
        if (
            len(rows) != expected[-2]
            or sum(row["rank_target_eligible"] == "true" for row in rows) != expected[-1]
            or report.get("reference_assets")
            != {
                asset_id: {
                    "relative_path": path.relative_to(self.root).as_posix(),
                    "sha256": _sha256(path),
                }
                for asset_id, path in references.items()
            }
            or _sha256(workbook_path) != registry["source_assets"][0]["sha256"]
        ):
            raise R4PXD017052NSCLCSourceAuditError("source audit accounting differs")
        return R4PXD017052NSCLCSourceAuditSummary(len(registry["source_assets"]), *expected, receipt_path)
