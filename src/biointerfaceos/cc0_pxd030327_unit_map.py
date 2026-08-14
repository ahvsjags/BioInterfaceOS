"""Verify the corrected PXD030327 run map without promoting it to a model target."""

from __future__ import annotations

import csv
import hashlib
import json
import stat
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class CC0PXD030327UnitMapError(RuntimeError):
    """Raised when corrected source mapping could weaken the T129 target gate."""


@dataclass(frozen=True)
class CC0PXD030327UnitMapSummary:
    """Non-result accounting for the PXD030327 supplemental source audit."""

    status: str
    unexcluded_unit_count: int
    matrix_run_count: int
    unmapped_matrix_column_count: int
    receipt_path: Path


def _canonical(value: Any) -> bytes:
    return (json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n").encode("utf-8")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise CC0PXD030327UnitMapError(f"{label} must be an object")
    return dict(value)


def _string(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise CC0PXD030327UnitMapError(f"{label} must be a non-empty string")
    return value.strip()


def _integer(value: Any, label: str, *, minimum: int = 0) -> int:
    if not isinstance(value, int) or isinstance(value, bool) or value < minimum:
        raise CC0PXD030327UnitMapError(f"{label} must be an integer >= {minimum}")
    return value


class CC0PXD030327UnitMapWorkflow:
    """Audit source-mapped units while requiring a non-admission decision."""

    AUDIT_ID = "bioif-r2-cc0-pxd030327-unit-map-v1.0.0"
    MANIFEST_RELATIVE = "docs/data/R2_T129_PXD030327_UNIT_MAP_MANIFEST.json"
    OUTPUT_RELATIVE = "reports/review_round_2/cc0_pxd030327_unit_map/v1.0.0"
    REQUIRED_MANIFEST_FIELDS = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "source",
        "assets",
        "analysis_unit_contract",
        "target_admission",
    }
    REQUIRED_ASSET_FIELDS = {
        "asset_id",
        "local_relative_path",
        "download_url",
        "expected_bytes",
        "sha256",
    }
    REQUIRED_SOURCE_FIELDS = {
        "accession",
        "laboratory",
        "license_id",
        "project_api_url",
        "landing_url",
        "publication_date",
    }
    REQUIRED_ADMISSION_FIELDS = {
        "numeric_material_covariate_status",
        "source_ratio_interpretation",
        "categorical_np_label_status",
        "cross_laboratory_endpoint_status",
        "admission",
        "model_use",
    }

    def __init__(
        self,
        root: Path,
        *,
        manifest_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.manifest_path = manifest_path or self.root / self.MANIFEST_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    def _relative_path(self, relative: Any, label: str) -> Path:
        path = (self.root / _string(relative, label)).resolve(strict=False)
        if not path.is_relative_to(self.root) or not path.is_file():
            raise CC0PXD030327UnitMapError(f"{label} is missing or outside the repository")
        return path

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise CC0PXD030327UnitMapError(f"cannot parse {label}") from exc

    @staticmethod
    def _require(condition: bool, label: str) -> None:
        if not condition:
            raise CC0PXD030327UnitMapError(label)

    def _manifest(self) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
        manifest = self._json(self.manifest_path, "PXD030327 unit-map manifest")
        self._require(
            set(manifest) == self.REQUIRED_MANIFEST_FIELDS
            and manifest.get("schema_version") == 1
            and manifest.get("audit_id") == self.AUDIT_ID
            and manifest.get("evidence_class") == "DEVELOPMENT_OBSERVATION"
            and manifest.get("allowed_claim_level") == "EXPLORATORY",
            "PXD030327 unit-map manifest identity is invalid",
        )
        source = _mapping(manifest.get("source"), "PXD030327 source")
        publication_date = source.get("publication_date")
        self._require(
            set(source) == self.REQUIRED_SOURCE_FIELDS
            and source.get("accession") == "PXD030327"
            and source.get("license_id") == "CC0-1.0"
            and source.get("project_api_url") == "https://www.ebi.ac.uk/pride/ws/archive/v2/projects/PXD030327"
            and source.get("landing_url") == "https://www.ebi.ac.uk/pride/archive/projects/PXD030327"
            and isinstance(publication_date, str)
            and publication_date <= "2024-12-31",
            "PXD030327 source identity is invalid",
        )
        assets = manifest.get("assets")
        if not isinstance(assets, list) or len(assets) != 3:
            raise CC0PXD030327UnitMapError("PXD030327 source asset inventory is invalid")
        assets_by_id: dict[str, dict[str, Any]] = {}
        for value in assets:
            asset = _mapping(value, "PXD030327 source asset")
            self._require(
                set(asset) == self.REQUIRED_ASSET_FIELDS,
                "PXD030327 source asset fields are invalid",
            )
            asset_id = _string(asset.get("asset_id"), "PXD030327 source asset id")
            self._require(
                asset_id not in assets_by_id,
                "PXD030327 source asset ids must be unique",
            )
            self._require(
                _string(asset.get("download_url"), "PXD030327 source download URL").startswith(
                    "https://ftp.pride.ebi.ac.uk/"
                ),
                "PXD030327 source download URL is not official",
            )
            digest = _string(asset.get("sha256"), "PXD030327 source SHA-256").lower()
            self._require(
                len(digest) == 64 and all(character in "0123456789abcdef" for character in digest),
                "PXD030327 source SHA-256 is invalid",
            )
            _integer(asset.get("expected_bytes"), "PXD030327 source asset bytes", minimum=1)
            assets_by_id[asset_id] = asset
        self._require(
            set(assets_by_id) == {"sample_table", "seven_np_matrix", "ten_plate_matrix"},
            "PXD030327 required source assets are missing",
        )
        admission = _mapping(manifest.get("target_admission"), "PXD030327 target admission")
        self._require(
            set(admission) == self.REQUIRED_ADMISSION_FIELDS
            and admission.get("numeric_material_covariate_status")
            == "MISSING_SOURCE_MATCHED_MATERIAL_OR_SIZE_COVARIATE"
            and admission.get("source_ratio_interpretation")
            == "SOURCE_DEFINED_NUMERIC_EXPOSURE_NOT_MATERIAL_OR_SIZE_COVARIATE"
            and admission.get("categorical_np_label_status") == "PROHIBITED_AS_PREDICTIVE_IDENTITY_FEATURE"
            and admission.get("cross_laboratory_endpoint_status") == "SINGLE_LAB_ONLY_NO_COMMON_ENDPOINT"
            and admission.get("admission") == "NOT_ADMITTED"
            and admission.get("model_use") == "PROHIBITED",
            "PXD030327 target admission boundary is weakened",
        )
        return manifest, assets_by_id

    def _verified_assets(self, assets: dict[str, dict[str, Any]]) -> dict[str, Path]:
        paths: dict[str, Path] = {}
        for asset_id, asset in assets.items():
            path = self._relative_path(asset.get("local_relative_path"), f"PXD030327 asset {asset_id}")
            self._require(
                path.stat().st_size == asset["expected_bytes"] and _sha256(path) == asset["sha256"],
                f"PXD030327 asset {asset_id} does not match the declared source bytes",
            )
            paths[asset_id] = path
        return paths

    @staticmethod
    def _matrix_runs(path: Path, expected_metadata: list[Any]) -> set[str]:
        try:
            with path.open(encoding="utf-8", newline="") as handle:
                header = next(csv.reader(handle, delimiter="\t"))
        except (OSError, UnicodeError, StopIteration, csv.Error) as exc:
            raise CC0PXD030327UnitMapError("cannot read PXD030327 protein matrix") from exc
        if header[: len(expected_metadata)] != expected_metadata:
            raise CC0PXD030327UnitMapError("PXD030327 matrix metadata columns are invalid")
        runs = {
            column.replace("\\", "/").rsplit("/", 1)[-1].removesuffix(".d")
            for column in header[len(expected_metadata) :]
        }
        if not runs or len(runs) != len(header) - len(expected_metadata):
            raise CC0PXD030327UnitMapError("PXD030327 matrix run columns are invalid")
        return runs

    @staticmethod
    def _source_rows(path: Path, contract: dict[str, Any]) -> tuple[set[str], dict[str, Any]]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook[_string(contract.get("worksheet"), "PXD030327 worksheet")]
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()
        except (KeyError, OSError, ValueError) as exc:
            raise CC0PXD030327UnitMapError("cannot read PXD030327 source design") from exc
        if not rows:
            raise CC0PXD030327UnitMapError("PXD030327 source design is empty")
        required_columns = contract.get("required_columns")
        if not isinstance(required_columns, list) or rows[0] != tuple(required_columns):
            raise CC0PXD030327UnitMapError("PXD030327 source design columns are invalid")
        positions = {name: index for index, name in enumerate(required_columns)}
        source_rows = [row for row in rows[1:] if row[positions["Run"]] is not None]
        included_rows = [row for row in source_rows if row[positions["Remove from analysis"]] is False]
        included_runs = {str(row[positions["Run"]]) for row in included_rows}
        if len(included_runs) != len(included_rows):
            raise CC0PXD030327UnitMapError("PXD030327 unexcluded source runs are not unique")
        observed = {
            "source_row_count": len(source_rows),
            "unexcluded_unit_count": len(included_rows),
            "excluded_row_count": len(source_rows) - len(included_rows),
            "source_np_labels": sorted({str(row[positions["NP"]]) for row in included_rows}),
            "source_ratio_values": sorted(
                {row[positions["P/NP ratio"]] for row in included_rows},
                key=lambda value: (1, 0.0, value) if isinstance(value, str) else (0, float(value), ""),
            ),
            "source_replicate_values": sorted({row[positions["Replicate"]] for row in included_rows}),
        }
        return included_runs, observed

    def _audit_inputs(self, manifest: dict[str, Any], paths: dict[str, Path]) -> tuple[dict[str, Any], dict[str, Any]]:
        contract = _mapping(manifest.get("analysis_unit_contract"), "PXD030327 unit contract")
        included_runs, observed = self._source_rows(paths["sample_table"], contract)
        expected_metadata = contract.get("matrix_metadata_columns")
        if not isinstance(expected_metadata, list) or len(expected_metadata) != 5:
            raise CC0PXD030327UnitMapError("PXD030327 matrix metadata contract is invalid")
        matrix_runs = {
            asset_id: self._matrix_runs(paths[asset_id], expected_metadata)
            for asset_id in ("seven_np_matrix", "ten_plate_matrix")
        }
        for key in (
            "source_row_count",
            "unexcluded_unit_count",
            "excluded_row_count",
            "source_np_labels",
            "source_ratio_values",
            "source_replicate_values",
        ):
            if contract.get(key) != observed[key]:
                raise CC0PXD030327UnitMapError(f"PXD030327 source design {key} is stale")
        counts = {
            "matrix_run_column_counts": {key: len(value) for key, value in matrix_runs.items()},
            "unexcluded_matrix_match_counts": {key: len(value & included_runs) for key, value in matrix_runs.items()},
            "unmapped_matrix_column_counts": {key: len(value - included_runs) for key, value in matrix_runs.items()},
            "unique_matrix_run_count": len(set().union(*matrix_runs.values())),
            "unexcluded_units_missing_from_matrices": len(included_runs - set().union(*matrix_runs.values())),
        }
        if any(contract.get(key) != value for key, value in counts.items()):
            raise CC0PXD030327UnitMapError("PXD030327 matrix-to-unit mapping is stale")
        return observed | counts, {
            asset_id: {"path": str(path.relative_to(self.root)), "sha256": _sha256(path)}
            for asset_id, path in paths.items()
        }

    def run(self, *, strict: bool = False) -> CC0PXD030327UnitMapSummary:
        """Write an immutable no-model receipt for corrected source mapping."""
        if not strict:
            raise CC0PXD030327UnitMapError("PXD030327 unit-map audit requires --strict")
        if self.output_root.exists():
            raise CC0PXD030327UnitMapError("PXD030327 unit-map audit already executed")
        manifest, assets = self._manifest()
        paths = self._verified_assets(assets)
        observed, source_assets = self._audit_inputs(manifest, paths)
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "evaluated_at": manifest["evaluated_at"],
            "manifest_sha256": _sha256(self.manifest_path),
            "status": "VERIFIED_SINGLE_LAB_UNIT_MAP_NOT_ADMITTED",
            "source": manifest["source"],
            "verified_source_assets": source_assets,
            "analysis_unit_observations": observed,
            "target_admission": manifest["target_admission"],
            "model_fitted": False,
            "paired_ablations_run": False,
            "external_ood_evaluated": False,
            "negative_controls_run": False,
            "independent_validation": False,
            "scientific_submission_ready": False,
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        report_path = self.output_root / "unit_map_correction_report.json"
        receipt_path = self.output_root / "unit_map_correction_receipt.json"
        report_path.write_bytes(_canonical(report))
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": report["status"],
            "unit_map_correction_report_sha256": _sha256(report_path),
            "unexcluded_unit_count": observed["unexcluded_unit_count"],
            "unique_matrix_run_count": observed["unique_matrix_run_count"],
            "unmapped_matrix_column_count": sum(observed["unmapped_matrix_column_counts"].values()),
            "admission": "NOT_ADMITTED",
            "model_use": "PROHIBITED",
            "model_fitted": False,
            "paired_ablations_run": False,
            "external_ood_evaluated": False,
            "negative_controls_run": False,
            "independent_validation": False,
            "scientific_submission_ready": False,
        }
        receipt_path.write_bytes(_canonical(receipt))
        for path in self.output_root.iterdir():
            path.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        self.output_root.chmod(stat.S_IRUSR | stat.S_IXUSR | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH)
        return CC0PXD030327UnitMapSummary(
            status=report["status"],
            unexcluded_unit_count=observed["unexcluded_unit_count"],
            matrix_run_count=observed["unique_matrix_run_count"],
            unmapped_matrix_column_count=sum(observed["unmapped_matrix_column_counts"].values()),
            receipt_path=receipt_path,
        )

    def verify(self) -> CC0PXD030327UnitMapSummary:
        """Verify the immutable supplemental receipt without recalculating a target."""
        report_path = self.output_root / "unit_map_correction_report.json"
        receipt_path = self.output_root / "unit_map_correction_receipt.json"
        report = self._json(report_path, "PXD030327 unit-map report")
        receipt = self._json(receipt_path, "PXD030327 unit-map receipt")
        observations = _mapping(report.get("analysis_unit_observations"), "PXD030327 observations")
        admission = _mapping(report.get("target_admission"), "PXD030327 admission")
        required_false = (
            "model_fitted",
            "paired_ablations_run",
            "external_ood_evaluated",
            "negative_controls_run",
            "independent_validation",
            "scientific_submission_ready",
        )
        if (
            report.get("audit_id") != self.AUDIT_ID
            or receipt.get("audit_id") != self.AUDIT_ID
            or report.get("status") != "VERIFIED_SINGLE_LAB_UNIT_MAP_NOT_ADMITTED"
            or receipt.get("status") != report.get("status")
            or receipt.get("unit_map_correction_report_sha256") != _sha256(report_path)
            or receipt.get("admission") != "NOT_ADMITTED"
            or receipt.get("model_use") != "PROHIBITED"
            or admission.get("admission") != "NOT_ADMITTED"
            or admission.get("model_use") != "PROHIBITED"
            or any(report.get(field) is not False or receipt.get(field) is not False for field in required_false)
            or receipt.get("unexcluded_unit_count") != observations.get("unexcluded_unit_count")
            or receipt.get("unique_matrix_run_count") != observations.get("unique_matrix_run_count")
            or receipt.get("unmapped_matrix_column_count")
            != sum(_mapping(observations.get("unmapped_matrix_column_counts"), "PXD030327 unmapped columns").values())
        ):
            raise CC0PXD030327UnitMapError("PXD030327 unit-map receipt is invalid")
        return CC0PXD030327UnitMapSummary(
            status=report["status"],
            unexcluded_unit_count=_integer(
                observations.get("unexcluded_unit_count"), "PXD030327 unit count", minimum=1
            ),
            matrix_run_count=_integer(
                observations.get("unique_matrix_run_count"), "PXD030327 matrix run count", minimum=1
            ),
            unmapped_matrix_column_count=_integer(
                receipt.get("unmapped_matrix_column_count"),
                "PXD030327 unmapped matrix column count",
                minimum=0,
            ),
            receipt_path=receipt_path,
        )
