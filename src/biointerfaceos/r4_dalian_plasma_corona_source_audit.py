"""Audit the CC0 PXD060795 human-plasma corona workbook for R4 sensitivity work.

The source is intentionally kept outside the primary R4 OOD endpoint: six
corona batches pass the shared-target threshold, but the frozen endpoint
requires twelve.  The audit therefore records usable quantitative evidence
without upgrading a small, pooled/unspecified source into independent
biological validation.
"""

from __future__ import annotations

import csv
import json
import math
from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R4DalianPlasmaCoronaSourceAuditError(RuntimeError):
    """Raised when the PXD060795 source contract is not reproducible."""


@dataclass(frozen=True)
class R4DalianPlasmaCoronaSourceAuditSummary:
    """Accounting summary for the small-n independent source."""

    source_asset_count: int
    protein_row_count: int
    all_measurement_batch_count: int
    corona_measurement_batch_count: int
    rank_qualified_measurement_batch_count: int
    shared_canonical_protein_count: int
    source_cell_count: int
    candidate_positive_source_cell_count: int
    receipt_path: Path


class R4DalianPlasmaCoronaSourceAuditWorkflow:
    """Create a no-imputation source-cell map for PXD060795."""

    AUDIT_ID = "bioif-r4-dalian-plasma-corona-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R4_T162_PXD060795_DALIAN_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_4/dalian_plasma_corona_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R4_PXD060795_dalian_plasma_corona_source_cell_map.csv"
    STATUS = "ADMITTED_R4_SMALL_N_INDEPENDENT_SOURCE_PENDING_SENSITIVITY_PROTOCOL"
    SOURCE_CELL_FIELDS = (
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "canonical_accession",
        "measurement_batch_id",
        "biological_unit_id",
        "condition_label",
        "analysis_candidate_eligible",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    )

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise R4DalianPlasmaCoronaSourceAuditError(f"cannot parse {label}") from exc

    @staticmethod
    def _numeric(value: Any) -> float | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        if isinstance(value, bool) or not isinstance(value, int | float):
            raise R4DalianPlasmaCoronaSourceAuditError("author abundance must be numeric or blank")
        value = float(value)
        if not math.isfinite(value) or value < 0:
            raise R4DalianPlasmaCoronaSourceAuditError("author abundance must be finite and non-negative")
        return value

    def _under(self, root: Path, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise R4DalianPlasmaCoronaSourceAuditError(f"{label} must use POSIX separators")
        pure = PurePosixPath(relative_path)
        if pure.is_absolute() or not pure.parts or ".." in pure.parts:
            raise R4DalianPlasmaCoronaSourceAuditError(f"{label} escapes its root")
        path = (root / Path(*pure.parts)).resolve(strict=False)
        if not path.is_relative_to(root) or not path.is_file():
            raise R4DalianPlasmaCoronaSourceAuditError(f"{label} is missing or outside its root")
        return path

    def _registry(self) -> tuple[dict[str, Any], Path, Path]:
        registry = self._json(self.registry_path, "R4 Dalian source registry")
        expected = {
            "schema_version",
            "audit_id",
            "evaluated_at",
            "evidence_class",
            "allowed_claim_level",
            "dataset",
            "source_scope",
            "source_assets",
            "r3_reference_asset",
            "worksheet_contract",
            "quantification_contract",
            "admission_minimums",
            "claim_boundary",
        }
        if set(registry) != expected or registry.get("schema_version") != 1 or registry.get("audit_id") != self.AUDIT_ID:
            raise R4DalianPlasmaCoronaSourceAuditError("registry fields are invalid")
        if registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION" or registry.get("allowed_claim_level") != "EXPLORATORY":
            raise R4DalianPlasmaCoronaSourceAuditError("registry evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "evaluated_at")
        _string(registry.get("claim_boundary"), "claim_boundary")
        dataset = _mapping(registry["dataset"], "dataset")
        if dataset.get("accession") != "PXD060795" or dataset.get("license") != "CC0":
            raise R4DalianPlasmaCoronaSourceAuditError("dataset declaration is invalid")
        scope = _mapping(registry["source_scope"], "source scope")
        if scope.get("source_id") != "PXD060795_DALIAN_PLA_MICRO_NANOPLASTIC_HUMAN_PLASMA_CORONA":
            raise R4DalianPlasmaCoronaSourceAuditError("source scope is invalid")
        assets = registry["source_assets"]
        if not isinstance(assets, list) or len(assets) != 1:
            raise R4DalianPlasmaCoronaSourceAuditError("source assets are invalid")
        item = _mapping(assets[0], "source asset")
        if set(item) != {"asset_id", "relative_path", "sha256", "expected_bytes"} or item["asset_id"] != "search_result_workbook":
            raise R4DalianPlasmaCoronaSourceAuditError("source asset fields are invalid")
        asset = self._under(self.assets_root, _string(item["relative_path"], "source asset path"), "source asset")
        if asset.stat().st_size != item["expected_bytes"] or _sha256(asset) != _checksum(item["sha256"], "source asset"):
            raise R4DalianPlasmaCoronaSourceAuditError("source asset checksum differs")
        reference = _mapping(registry["r3_reference_asset"], "R3 reference asset")
        feature_path = self._under(self.root, _string(reference["relative_path"], "R3 reference asset"), "R3 reference asset")
        if _sha256(feature_path) != _checksum(reference["sha256"], "R3 reference asset"):
            raise R4DalianPlasmaCoronaSourceAuditError("R3 reference asset checksum differs")
        return registry, asset, feature_path

    @staticmethod
    def _features(path: Path) -> set[str]:
        with path.open(newline="", encoding="utf-8-sig") as stream:
            accessions = {row["canonical_accession"] for row in csv.DictReader(stream)}
        if not accessions:
            raise R4DalianPlasmaCoronaSourceAuditError("R3 feature table is empty")
        return accessions

    def _cells(self, asset: Path, feature_path: Path, registry: dict[str, Any]) -> list[dict[str, str]]:
        contract = registry["worksheet_contract"]
        workbook = load_workbook(asset, read_only=True, data_only=True)
        if workbook.sheetnames != [contract["worksheet"]]:
            raise R4DalianPlasmaCoronaSourceAuditError("workbook sheets differ")
        sheet = workbook.active
        if sheet.max_row != contract["expected_rows_after_header"] + 1 or sheet.max_column != contract["expected_columns"]:
            raise R4DalianPlasmaCoronaSourceAuditError("workbook dimensions differ")
        header = next(sheet.iter_rows(min_row=contract["header_row"], max_row=contract["header_row"], values_only=True))
        measurement_columns = contract["measurement_columns"]
        indices = [(index, label) for index, label in enumerate(header) if label in measurement_columns]
        if [label for _, label in indices] != measurement_columns or header[3] != contract["accession_column"]:
            raise R4DalianPlasmaCoronaSourceAuditError("workbook header differs")
        features = self._features(feature_path)
        rows: list[dict[str, str]] = []
        direct_rows: set[int] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source_id = row[3]
            if not isinstance(source_id, str) or source_id.strip() not in features:
                continue
            accession = source_id.strip()
            direct_rows.add(row_number)
            for column_index, label in indices:
                condition = label.rsplit(", ", 1)[1]
                value = self._numeric(row[column_index])
                candidate = bool(contract["condition_candidate_policy"][condition])
                state = "SOURCE_BLANK" if value is None else ("NUMERIC_ZERO" if value == 0 else "POSITIVE")
                rows.append(
                    {
                        "source_id": registry["source_scope"]["source_id"],
                        "laboratory_anchor": registry["source_scope"]["laboratory_anchor"],
                        "source_asset_id": "search_result_workbook",
                        "source_worksheet": contract["worksheet"],
                        "source_row": str(row_number),
                        "source_coordinate": f"{get_column_letter(column_index + 1)}{row_number}",
                        "source_identifier": accession,
                        "canonical_accession": accession,
                        "measurement_batch_id": f"R4_PXD060795_{label.split(': ', 2)[1].replace(': ', '_').replace(', ', '_')}",
                        "biological_unit_id": "POOLED_OR_UNSPECIFIED_HUMAN_PLASMA",
                        "condition_label": condition,
                        "analysis_candidate_eligible": "true" if candidate else "false",
                        "author_quantity_type": "PROTEOME_DISCOVERER_NORMALIZED_ABUNDANCE",
                        "author_numeric_value": "" if value is None else str(value),
                        "author_value_state": state,
                        "rank_target_eligible": "true" if candidate and value is not None and value > 0 else "false",
                    }
                )
        if len(direct_rows) != contract["expected_direct_shared_rows"] or len({row["canonical_accession"] for row in rows}) != contract["expected_shared_canonical_accessions"]:
            raise R4DalianPlasmaCoronaSourceAuditError("shared R3 source-row accounting differs")
        if len(rows) != contract["expected_direct_shared_rows"] * len(measurement_columns):
            raise R4DalianPlasmaCoronaSourceAuditError("source-cell accounting differs")
        by_batch: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in rows:
            by_batch[row["measurement_batch_id"]].append(row)
        candidates = {batch: values for batch, values in by_batch.items() if values[0]["analysis_candidate_eligible"] == "true"}
        qualified = {batch: values for batch, values in candidates.items() if sum(row["rank_target_eligible"] == "true" for row in values) >= 10}
        expected = contract["expected_positive_shared_by_measurement_column"]
        observed = [sum(row["rank_target_eligible"] == "true" for row in by_batch[batch]) for batch in by_batch]
        if observed != expected or len(by_batch) != 9 or len(candidates) != 6 or len(qualified) != 6:
            raise R4DalianPlasmaCoronaSourceAuditError("measurement-batch accounting differs")
        if sum(row["rank_target_eligible"] == "true" for row in rows) != contract["expected_candidate_positive_cells"]:
            raise R4DalianPlasmaCoronaSourceAuditError("candidate positive source-cell accounting differs")
        return rows

    def run(self, *, strict: bool = False) -> R4DalianPlasmaCoronaSourceAuditSummary:
        if not strict:
            raise R4DalianPlasmaCoronaSourceAuditError("R4 Dalian source audit requires --strict")
        if self.output_root.exists():
            raise R4DalianPlasmaCoronaSourceAuditError("R4 Dalian source audit already executed")
        registry, asset, feature_path = self._registry()
        rows = self._cells(asset, feature_path, registry)
        derived = self.assets_root / self.DERIVED_RELATIVE
        if derived.exists():
            raise R4DalianPlasmaCoronaSourceAuditError("derived source cell map already exists")
        derived.parent.mkdir(parents=True, exist_ok=True)
        with derived.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=self.SOURCE_CELL_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
        self.output_root.mkdir(parents=True, exist_ok=False)
        batches = {row["measurement_batch_id"] for row in rows}
        candidates = {row["measurement_batch_id"] for row in rows if row["analysis_candidate_eligible"] == "true"}
        qualified = {
            batch for batch in candidates if sum(row["rank_target_eligible"] == "true" for row in rows if row["measurement_batch_id"] == batch) >= 10
        }
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "source_asset": {"relative_path": "Search_result.xlsx", "sha256": _sha256(asset)},
            "r3_reference_asset": {"relative_path": registry["r3_reference_asset"]["relative_path"], "sha256": _sha256(feature_path)},
            "source_cell_map": {"relative_path": self.DERIVED_RELATIVE, "sha256": _sha256(derived)},
            "protein_row_count": 547,
            "all_measurement_batch_count": len(batches),
            "corona_measurement_batch_count": len(candidates),
            "rank_qualified_measurement_batch_count": len(qualified),
            "shared_canonical_protein_count": len({row["canonical_accession"] for row in rows}),
            "source_cell_count": len(rows),
            "candidate_positive_source_cell_count": sum(row["rank_target_eligible"] == "true" for row in rows),
            "independent_laboratory_anchor_count_contributed": 1,
            "primary_ood_minimum_met": False,
            "claim_boundary": registry["claim_boundary"],
        }
        report_path = self.output_root / "dalian_plasma_corona_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report": {"relative_path": report_path.name, "sha256": _sha256(report_path)},
            "source_cell_map": report["source_cell_map"],
            "primary_ood_minimum_met": False,
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "dalian_plasma_corona_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R4DalianPlasmaCoronaSourceAuditSummary(1, 547, len(batches), len(candidates), len(qualified), len({row["canonical_accession"] for row in rows}), len(rows), sum(row["rank_target_eligible"] == "true" for row in rows), receipt_path)

    def verify(self) -> R4DalianPlasmaCoronaSourceAuditSummary:
        registry, asset, feature_path = self._registry()
        report_path = self.output_root / "dalian_plasma_corona_source_audit_report.json"
        receipt_path = self.output_root / "dalian_plasma_corona_source_audit_receipt.json"
        report = self._json(report_path, "R4 Dalian audit report")
        receipt = self._json(receipt_path, "R4 Dalian audit receipt")
        if report.get("status") != self.STATUS or receipt.get("status") != self.STATUS or receipt.get("report", {}).get("sha256") != _sha256(report_path):
            raise R4DalianPlasmaCoronaSourceAuditError("audit receipt differs")
        cell_map = self._under(self.assets_root, report["source_cell_map"]["relative_path"], "source cell map")
        if report["source_cell_map"].get("sha256") != _sha256(cell_map):
            raise R4DalianPlasmaCoronaSourceAuditError("source cell map checksum differs")
        with cell_map.open(newline="", encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        expected = self._cells(asset, feature_path, registry)
        if rows != expected:
            raise R4DalianPlasmaCoronaSourceAuditError("source cell map differs")
        summary = R4DalianPlasmaCoronaSourceAuditSummary(1, 547, 9, 6, 6, 27, len(rows), sum(row["rank_target_eligible"] == "true" for row in rows), receipt_path)
        if report.get("source_cell_count") != summary.source_cell_count or report.get("primary_ood_minimum_met") is not False:
            raise R4DalianPlasmaCoronaSourceAuditError("audit accounting differs")
        return summary
