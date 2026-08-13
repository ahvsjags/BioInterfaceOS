"""Audit a CC-BY full-text protein-corona multi-core data package.

This workflow admits one narrowly defined, real-data benchmark: technical
between-core LC-MS/MS heterogeneity for identical protein-corona aliquots.
It intentionally does *not* admit a biological cross-study prediction target,
because the source varies measurement centres rather than biological cohorts
or nanoparticle conditions.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
from collections import Counter
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook


class FulltextMulticoreAuditError(RuntimeError):
    """Raised when a full-text data package is incomplete or over-promoted."""


def _canonical(value: Any) -> bytes:
    return (
        json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False) + "\n"
    ).encode("utf-8")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise FulltextMulticoreAuditError(f"{label} must be an object")
    return dict(value)


def _list(value: Any, label: str, *, minimum: int = 0) -> list[Any]:
    if not isinstance(value, list) or len(value) < minimum:
        raise FulltextMulticoreAuditError(f"{label} must contain at least {minimum} items")
    return value


def _string(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise FulltextMulticoreAuditError(f"{label} must be a non-empty string")
    return value.strip()


def _checksum(value: Any, label: str) -> str:
    checksum = _string(value, label)
    if len(checksum) != 64 or any(character not in "0123456789abcdef" for character in checksum):
        raise FulltextMulticoreAuditError(f"{label} must be a lowercase SHA-256")
    return checksum


def _finite(value: Any, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, int | float):
        raise FulltextMulticoreAuditError(f"{label} must be finite numeric")
    numeric = float(value)
    if not math.isfinite(numeric):
        raise FulltextMulticoreAuditError(f"{label} must be finite numeric")
    return numeric


def _csv_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".17g")
    return str(value)


@dataclass(frozen=True)
class FulltextMulticoreAuditSummary:
    """Accounting for a narrow, technical real-data benchmark."""

    source_asset_count: int
    semiquantitative_core_count: int
    analysis_unit_count: int
    replicate_source_cell_count: int
    numeric_replicate_value_count: int
    status: str
    receipt_path: Path


class FulltextMulticoreAuditWorkflow:
    """Create an auditable source-to-cell map for the CC-BY multi-core study."""

    AUDIT_ID = "bioif-r3-fulltext-multicore-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R3_T144_FULLTEXT_MULTICORE_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_3/fulltext_multicore_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R3_PMC9633814_semiquantitative_source_cell_map.csv"
    STATUS = "ADMITTED_TECHNICAL_CROSS_CORE_BENCHMARK_ONLY"
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "source",
        "source_assets",
        "semiquantitative_table",
        "benchmark_scope",
    }
    REQUIRED_SOURCE = {
        "source_id",
        "title",
        "doi",
        "pmcid",
        "license",
        "article_locator",
        "supplementary_package_locator",
    }
    REQUIRED_ASSET = {"asset_id", "relative_path", "sha256", "expected_bytes"}
    REQUIRED_TABLE = {
        "asset_id",
        "worksheet",
        "header_row",
        "expected_columns",
        "expected_data_rows",
        "expected_facility_row_counts",
        "expected_replicate_source_cell_count",
        "expected_numeric_replicate_value_count",
        "expected_non_numeric_replicate_markers",
    }
    REQUIRED_SCOPE = {
        "benchmark_id",
        "admission_status",
        "measurement_scope",
        "analysis_unit",
        "primary_uses",
        "prohibited_uses",
        "model_status",
        "scientific_submission_ready",
    }
    EXPECTED_COLUMNS = (
        "Protein IDs",
        "Sequence coverage (%)",
        "PSM",
        "Number of peptides",
        "Number of unique peptides",
        "Gene names",
        "Replicate 1",
        "Replicate 2",
        "Replicate 3",
        "CV (%)",
        "Core facility code",
    )

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=True)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    def _asset_path(self, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise FulltextMulticoreAuditError(f"{label} must use a POSIX relative path")
        pure_path = PurePosixPath(relative_path)
        if pure_path.is_absolute() or not pure_path.parts or ".." in pure_path.parts:
            raise FulltextMulticoreAuditError(f"{label} escapes the source asset root")
        path = (self.assets_root / Path(*pure_path.parts)).resolve(strict=False)
        if not path.is_relative_to(self.assets_root) or not path.is_file():
            raise FulltextMulticoreAuditError(f"{label} is missing or outside the source asset root")
        return path

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise FulltextMulticoreAuditError(f"cannot parse {label}") from exc

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path]]:
        registry = self._json(self.registry_path, "R3 full-text multicore registry")
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise FulltextMulticoreAuditError("R3 full-text multicore registry fields are invalid")
        if registry.get("audit_id") != self.AUDIT_ID:
            raise FulltextMulticoreAuditError("R3 full-text multicore registry identity is invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise FulltextMulticoreAuditError("R3 full-text multicore evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "R3 full-text multicore evaluated_at")

        source = _mapping(registry.get("source"), "R3 full-text multicore source")
        if set(source) != self.REQUIRED_SOURCE:
            raise FulltextMulticoreAuditError("R3 full-text multicore source fields are invalid")
        if (
            source.get("source_id") != "PMC9633814"
            or source.get("doi") != "10.1038/s41467-022-34438-8"
            or source.get("pmcid") != "PMC9633814"
            or source.get("license") != "CC-BY-4.0"
        ):
            raise FulltextMulticoreAuditError("R3 full-text multicore source identity or licence is invalid")
        for field in ("title", "article_locator", "supplementary_package_locator"):
            value = _string(source.get(field), f"R3 full-text multicore source {field}")
            if field.endswith("locator") and not value.startswith("https://"):
                raise FulltextMulticoreAuditError("R3 full-text multicore source locator is invalid")

        assets: dict[str, Path] = {}
        for asset_value in _list(registry.get("source_assets"), "R3 source assets", minimum=2):
            asset = _mapping(asset_value, "R3 source asset")
            if set(asset) != self.REQUIRED_ASSET:
                raise FulltextMulticoreAuditError("R3 source asset fields are invalid")
            asset_id = _string(asset.get("asset_id"), "R3 source asset id")
            if asset_id in assets:
                raise FulltextMulticoreAuditError("R3 source asset id is duplicated")
            path = self._asset_path(_string(asset.get("relative_path"), "R3 source asset path"), asset_id)
            if path.stat().st_size != asset.get("expected_bytes"):
                raise FulltextMulticoreAuditError(f"R3 source asset size mismatch: {asset_id}")
            if _sha256(path) != _checksum(asset.get("sha256"), f"R3 source asset hash {asset_id}"):
                raise FulltextMulticoreAuditError(f"R3 source asset checksum mismatch: {asset_id}")
            assets[asset_id] = path

        table = _mapping(registry.get("semiquantitative_table"), "R3 semiquantitative table")
        if set(table) != self.REQUIRED_TABLE:
            raise FulltextMulticoreAuditError("R3 semiquantitative table fields are invalid")
        if table.get("asset_id") not in assets or table.get("worksheet") != "Final Prot":
            raise FulltextMulticoreAuditError("R3 semiquantitative asset map is invalid")
        if table.get("header_row") != 5 or tuple(table.get("expected_columns", [])) != self.EXPECTED_COLUMNS:
            raise FulltextMulticoreAuditError("R3 semiquantitative column map is invalid")
        if not isinstance(table.get("expected_data_rows"), int) or table["expected_data_rows"] < 1:
            raise FulltextMulticoreAuditError("R3 semiquantitative row count is invalid")
        counts = _mapping(table.get("expected_facility_row_counts"), "R3 facility row counts")
        if not counts or any(not isinstance(value, int) or value < 1 for value in counts.values()):
            raise FulltextMulticoreAuditError("R3 facility row counts are invalid")
        for field in (
            "expected_replicate_source_cell_count",
            "expected_numeric_replicate_value_count",
        ):
            if not isinstance(table.get(field), int) or table[field] < 1:
                raise FulltextMulticoreAuditError("R3 replicate count is invalid")
        markers = _mapping(
            table.get("expected_non_numeric_replicate_markers"), "R3 replicate markers"
        )
        if any(
            not isinstance(marker, str)
            or not marker.strip()
            or not isinstance(count, int)
            or count < 1
            for marker, count in markers.items()
        ):
            raise FulltextMulticoreAuditError("R3 replicate marker count is invalid")

        scope = _mapping(registry.get("benchmark_scope"), "R3 benchmark scope")
        if set(scope) != self.REQUIRED_SCOPE:
            raise FulltextMulticoreAuditError("R3 benchmark scope fields are invalid")
        if (
            scope.get("benchmark_id") != "R3-TECHNICAL-CORONA-MULTICORE-001"
            or scope.get("admission_status") != self.STATUS
            or scope.get("measurement_scope") != "INDEPENDENT_PROTEOMICS_CORE_HETEROGENEITY"
            or scope.get("analysis_unit") != "protein_by_core_facility_source_row"
            or scope.get("model_status") != "NOT_FITTED"
            or scope.get("scientific_submission_ready") is not False
        ):
            raise FulltextMulticoreAuditError("R3 benchmark scope is over-promoted or invalid")
        for field in ("primary_uses", "prohibited_uses"):
            if any(not isinstance(item, str) or not item.strip() for item in _list(scope.get(field), field, minimum=2)):
                raise FulltextMulticoreAuditError("R3 benchmark scope list is invalid")
        return registry, assets

    @staticmethod
    def _table_rows(
        table_path: Path, table: dict[str, Any]
    ) -> tuple[list[dict[str, str]], Counter[str], int, Counter[str]]:
        try:
            workbook = load_workbook(table_path, read_only=True, data_only=True)
            if table["worksheet"] not in workbook.sheetnames:
                raise FulltextMulticoreAuditError("R3 worksheet is missing")
            worksheet = workbook[table["worksheet"]]
            all_rows = worksheet.iter_rows(values_only=True)
            for _ in range(table["header_row"] - 1):
                next(all_rows)
            header = tuple(_csv_cell(value).strip() for value in next(all_rows))
            if header != FulltextMulticoreAuditWorkflow.EXPECTED_COLUMNS:
                raise FulltextMulticoreAuditError("R3 worksheet header does not match source map")
            extracted: list[dict[str, str]] = []
            facilities: Counter[str] = Counter()
            numeric_replicate_values = 0
            non_numeric_markers: Counter[str] = Counter()
            for row_number, values in enumerate(all_rows, start=table["header_row"] + 1):
                if len(values) != len(header):
                    raise FulltextMulticoreAuditError("R3 worksheet data row width is invalid")
                record = {header[index]: _csv_cell(value) for index, value in enumerate(values)}
                facility = record["Core facility code"].strip()
                if not facility:
                    raise FulltextMulticoreAuditError("R3 source row lacks a core facility code")
                for replicate in ("Replicate 1", "Replicate 2", "Replicate 3"):
                    source_value = values[header.index(replicate)]
                    if isinstance(source_value, int | float) and not isinstance(source_value, bool):
                        _finite(source_value, f"R3 source row {row_number} {replicate}")
                        numeric_replicate_values += 1
                    elif isinstance(source_value, str) and source_value.strip() == "NA":
                        non_numeric_markers["NA"] += 1
                    else:
                        raise FulltextMulticoreAuditError(
                            f"R3 source row {row_number} {replicate} has an unregistered marker"
                        )
                extracted.append(
                    {
                        "analysis_unit_id": f"PMC9633814:Final Prot:{row_number}",
                        "source_article_doi": "10.1038/s41467-022-34438-8",
                        "source_pmcid": "PMC9633814",
                        "source_license": "CC-BY-4.0",
                        "source_asset_id": table["asset_id"],
                        "source_worksheet": table["worksheet"],
                        "source_row": str(row_number),
                        "source_cell_range": f"Final Prot!A{row_number}:K{row_number}",
                        "core_facility_code": facility,
                        "protein_ids": record["Protein IDs"],
                        "gene_names": record["Gene names"],
                        "sequence_coverage_percent": record["Sequence coverage (%)"],
                        "peptide_spectrum_matches": record["PSM"],
                        "number_of_peptides": record["Number of peptides"],
                        "number_of_unique_peptides": record["Number of unique peptides"],
                        "replicate_1": record["Replicate 1"],
                        "replicate_2": record["Replicate 2"],
                        "replicate_3": record["Replicate 3"],
                        "coefficient_of_variation_percent": record["CV (%)"],
                    }
                )
                facilities[facility] += 1
        except (OSError, StopIteration) as exc:
            raise FulltextMulticoreAuditError("R3 worksheet cannot be read") from exc
        finally:
            try:
                workbook.close()
            except UnboundLocalError:
                pass
        return extracted, facilities, numeric_replicate_values, non_numeric_markers

    def run(self, *, strict: bool = False) -> FulltextMulticoreAuditSummary:
        if not strict:
            raise FulltextMulticoreAuditError("R3 full-text multicore audit requires --strict")
        if self.output_root.exists():
            raise FulltextMulticoreAuditError("R3 full-text multicore audit already executed")
        registry, assets = self._registry()
        table = _mapping(registry["semiquantitative_table"], "R3 semiquantitative table")
        rows, facilities, numeric_replicates, non_numeric_markers = self._table_rows(
            assets[table["asset_id"]], table
        )
        expected_facilities = _mapping(
            table["expected_facility_row_counts"], "R3 facility row counts"
        )
        if len(rows) != table["expected_data_rows"] or dict(sorted(facilities.items())) != expected_facilities:
            raise FulltextMulticoreAuditError("R3 source row or facility counts do not match registry")
        replicate_source_cells = len(rows) * 3
        if (
            replicate_source_cells != table["expected_replicate_source_cell_count"]
            or numeric_replicates != table["expected_numeric_replicate_value_count"]
            or dict(sorted(non_numeric_markers.items()))
            != table["expected_non_numeric_replicate_markers"]
        ):
            raise FulltextMulticoreAuditError("R3 source replicate count does not match registry")

        derived_path = self.assets_root / self.DERIVED_RELATIVE
        if derived_path.exists():
            raise FulltextMulticoreAuditError("R3 derived source-cell map already exists")
        derived_path.parent.mkdir(parents=True, exist_ok=True)
        with derived_path.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(rows[0]), lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        derived_sha256 = _sha256(derived_path)
        scope = _mapping(registry["benchmark_scope"], "R3 benchmark scope")
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "evaluated_at": registry["evaluated_at"],
            "registry_sha256": _sha256(self.registry_path),
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "source": registry["source"],
            "verified_source_asset_count": len(assets),
            "source_asset_sha256": {asset_id: _sha256(path) for asset_id, path in assets.items()},
            "technical_scope": scope,
            "semiquantitative_core_count": len(facilities),
            "core_facility_row_counts": dict(sorted(facilities.items())),
            "analysis_unit_count": len(rows),
            "replicate_source_cell_count": replicate_source_cells,
            "numeric_replicate_value_count": numeric_replicates,
            "non_numeric_replicate_markers": dict(sorted(non_numeric_markers.items())),
            "source_to_cell_map": {
                "location": self.DERIVED_RELATIVE,
                "sha256": derived_sha256,
                "coordinate_definition": "Every output row retains its original Final Prot worksheet row and A:K cell range.",
                "transformation": "openpyxl read_only/data_only extraction; no imputation, pooling, relabelling, or biological-label inference.",
            },
            "status": self.STATUS,
            "target_status": "TECHNICAL_BENCHMARK_ONLY",
            "model_fitted": False,
            "paired_ablations_run": False,
            "external_ood_evaluated": False,
            "independent_validation": False,
            "scientific_submission_ready": False,
        }
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report_sha256": hashlib.sha256(_canonical(report)).hexdigest(),
            "verified_source_asset_count": len(assets),
            "semiquantitative_core_count": len(facilities),
            "analysis_unit_count": len(rows),
            "replicate_source_cell_count": replicate_source_cells,
            "numeric_replicate_value_count": numeric_replicates,
            "target_status": "TECHNICAL_BENCHMARK_ONLY",
            "model_fitted": False,
            "scientific_submission_ready": False,
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        self._write(self.output_root / "fulltext_multicore_audit_report.json", report)
        self._write(self.output_root / "fulltext_multicore_audit_receipt.json", receipt)
        return FulltextMulticoreAuditSummary(
            source_asset_count=len(assets),
            semiquantitative_core_count=len(facilities),
            analysis_unit_count=len(rows),
            replicate_source_cell_count=replicate_source_cells,
            numeric_replicate_value_count=numeric_replicates,
            status=self.STATUS,
            receipt_path=self.output_root / "fulltext_multicore_audit_receipt.json",
        )

    def verify(self) -> FulltextMulticoreAuditSummary:
        report_path = self.output_root / "fulltext_multicore_audit_report.json"
        receipt_path = self.output_root / "fulltext_multicore_audit_receipt.json"
        report = self._json(report_path, "R3 full-text multicore report")
        receipt = self._json(receipt_path, "R3 full-text multicore receipt")
        source_map = _mapping(report.get("source_to_cell_map"), "R3 source-to-cell map")
        derived_path = self.assets_root / _string(source_map.get("location"), "R3 source map location")
        if (
            report.get("audit_id") != self.AUDIT_ID
            or receipt.get("audit_id") != self.AUDIT_ID
            or report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report_sha256") != _sha256(report_path)
            or report.get("target_status") != "TECHNICAL_BENCHMARK_ONLY"
            or report.get("model_fitted") is not False
            or report.get("scientific_submission_ready") is not False
            or not derived_path.is_file()
            or source_map.get("sha256") != _sha256(derived_path)
        ):
            raise FulltextMulticoreAuditError("R3 full-text multicore audit receipt is invalid")
        return FulltextMulticoreAuditSummary(
            source_asset_count=int(receipt["verified_source_asset_count"]),
            semiquantitative_core_count=int(receipt["semiquantitative_core_count"]),
            analysis_unit_count=int(receipt["analysis_unit_count"]),
            replicate_source_cell_count=int(receipt["replicate_source_cell_count"]),
            numeric_replicate_value_count=int(receipt["numeric_replicate_value_count"]),
            status=self.STATUS,
            receipt_path=receipt_path,
        )
