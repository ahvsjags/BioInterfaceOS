"""Create a source-to-cell map for the CC-BY PXD017052 supplementary matrix."""

from __future__ import annotations

import contextlib
import csv
import hashlib
import json
import math
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.fulltext_multicore_audit import (
    _canonical,
    _checksum,
    _csv_cell,
    _list,
    _mapping,
    _sha256,
    _string,
)


class PXD017052SourceCellAuditError(RuntimeError):
    """Raised when PXD017052 row/cell provenance is invalid or over-promoted."""


@dataclass(frozen=True)
class PXD017052SourceCellAuditSummary:
    """Accounting for a source-native, cell-level PXD017052 table map."""

    source_asset_count: int
    protein_row_count: int
    result_unit_count: int
    analysis_unit_count: int
    source_blank_count: int
    status: str
    receipt_path: Path


class PXD017052SourceCellAuditWorkflow:
    """Preserve every PXD017052 LFQ cell with its published particle crosswalk."""

    AUDIT_ID = "bioif-r3-pxd017052-source-cell-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R3_T147_PXD017052_SOURCE_CELL_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_3/pxd017052_source_cell_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R3_PXD017052_lfq_source_cell_map.csv"
    STATUS = "ADMITTED_SOURCE_NATIVE_CCBY_WITH_EXPLICIT_UNIT_TO_PARTICLE_MAP"
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "source",
        "source_assets",
        "table",
        "unit_to_particle_map",
        "scope",
    }
    REQUIRED_SOURCE = {"doi", "pmcid", "license", "article_locator"}
    REQUIRED_ASSET = {"asset_id", "relative_path", "sha256", "expected_bytes"}
    REQUIRED_TABLE = {
        "asset_id",
        "worksheet",
        "title_cell",
        "title_value",
        "header_row",
        "protein_identifier_column",
        "lfq_header_prefix",
        "expected_protein_rows",
        "expected_lfq_cells",
        "expected_numeric_lfq_cells",
        "expected_source_blank_lfq_cells",
        "expected_explicit_zero_lfq_cells",
    }
    REQUIRED_MAP = {"particle", "result_unit_id", "assay_replicate"}
    REQUIRED_SCOPE = {
        "admission_status",
        "data_semantics",
        "prohibited_use",
        "model_status",
        "scientific_submission_ready",
    }

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=True)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise PXD017052SourceCellAuditError(f"cannot parse {label}") from exc

    def _asset_path(self, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise PXD017052SourceCellAuditError(f"{label} must use a POSIX relative path")
        pure_path = PurePosixPath(relative_path)
        if pure_path.is_absolute() or not pure_path.parts or ".." in pure_path.parts:
            raise PXD017052SourceCellAuditError(f"{label} escapes the source asset root")
        path = (self.assets_root / Path(*pure_path.parts)).resolve(strict=False)
        if not path.is_relative_to(self.assets_root) or not path.is_file():
            raise PXD017052SourceCellAuditError(f"{label} is missing or outside the source asset root")
        return path

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path], dict[str, dict[str, Any]]]:
        registry = self._json(self.registry_path, "R3 PXD017052 source-cell registry")
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise PXD017052SourceCellAuditError("R3 PXD017052 registry fields are invalid")
        if registry.get("audit_id") != self.AUDIT_ID:
            raise PXD017052SourceCellAuditError("R3 PXD017052 registry identity is invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise PXD017052SourceCellAuditError("R3 PXD017052 evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "R3 PXD017052 evaluated_at")
        source = _mapping(registry.get("source"), "R3 PXD017052 source")
        if set(source) != self.REQUIRED_SOURCE or (
            source.get("doi") != "10.1038/s41467-020-17033-7"
            or source.get("pmcid") != "PMC7376165"
            or source.get("license") != "CC-BY-4.0"
            or not _string(source.get("article_locator"), "R3 PXD017052 article locator").startswith("https://")
        ):
            raise PXD017052SourceCellAuditError("R3 PXD017052 source identity or licence is invalid")
        assets: dict[str, Path] = {}
        for value in _list(registry.get("source_assets"), "R3 PXD017052 assets", minimum=2):
            asset = _mapping(value, "R3 PXD017052 asset")
            if set(asset) != self.REQUIRED_ASSET:
                raise PXD017052SourceCellAuditError("R3 PXD017052 asset fields are invalid")
            asset_id = _string(asset.get("asset_id"), "R3 PXD017052 asset id")
            if asset_id in assets:
                raise PXD017052SourceCellAuditError("R3 PXD017052 asset id is duplicated")
            path = self._asset_path(_string(asset.get("relative_path"), asset_id), asset_id)
            if path.stat().st_size != asset.get("expected_bytes") or _sha256(path) != _checksum(
                asset.get("sha256"), asset_id
            ):
                raise PXD017052SourceCellAuditError(f"R3 PXD017052 asset differs: {asset_id}")
            assets[asset_id] = path
        table = _mapping(registry.get("table"), "R3 PXD017052 table")
        if set(table) != self.REQUIRED_TABLE or table.get("asset_id") not in assets:
            raise PXD017052SourceCellAuditError("R3 PXD017052 table fields are invalid")
        for field in (
            "worksheet",
            "title_cell",
            "title_value",
            "protein_identifier_column",
            "lfq_header_prefix",
        ):
            _string(table.get(field), f"R3 PXD017052 table {field}")
        for field in self.REQUIRED_TABLE - {
            "asset_id",
            "worksheet",
            "title_cell",
            "title_value",
            "protein_identifier_column",
            "lfq_header_prefix",
        }:
            if not isinstance(table.get(field), int) or table[field] < 0:
                raise PXD017052SourceCellAuditError("R3 PXD017052 table count is invalid")
        map_by_unit: dict[str, dict[str, Any]] = {}
        for value in _list(registry.get("unit_to_particle_map"), "R3 PXD017052 unit map", minimum=9):
            row = _mapping(value, "R3 PXD017052 unit map row")
            if set(row) != self.REQUIRED_MAP:
                raise PXD017052SourceCellAuditError("R3 PXD017052 unit-map fields are invalid")
            unit = _string(row.get("result_unit_id"), "R3 PXD017052 result unit")
            if unit in map_by_unit or not _string(row.get("particle"), "R3 PXD017052 particle"):
                raise PXD017052SourceCellAuditError("R3 PXD017052 unit-map identity is invalid")
            if row.get("assay_replicate") not in {1, 2, 3}:
                raise PXD017052SourceCellAuditError("R3 PXD017052 replicate is invalid")
            map_by_unit[unit] = row
        if len(map_by_unit) != 9:
            raise PXD017052SourceCellAuditError("R3 PXD017052 unit-map coverage is invalid")
        scope = _mapping(registry.get("scope"), "R3 PXD017052 scope")
        if set(scope) != self.REQUIRED_SCOPE or (
            scope.get("admission_status") != self.STATUS
            or scope.get("data_semantics") != "SOURCE_NATIVE_LFQ_CELLS"
            or scope.get("model_status") != "NOT_FITTED"
            or scope.get("scientific_submission_ready") is not False
        ):
            raise PXD017052SourceCellAuditError("R3 PXD017052 scope is over-promoted or invalid")
        if any(
            not isinstance(item, str) or not item.strip()
            for item in _list(scope.get("prohibited_use"), "R3 PXD017052 prohibited use", minimum=2)
        ):
            raise PXD017052SourceCellAuditError("R3 PXD017052 scope list is invalid")
        return registry, assets, map_by_unit

    @staticmethod
    def _source_cells(
        workbook_path: Path, table: dict[str, Any], unit_map: dict[str, dict[str, Any]]
    ) -> tuple[list[dict[str, str]], int, int, int]:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            if table["worksheet"] not in workbook.sheetnames:
                raise PXD017052SourceCellAuditError("R3 PXD017052 worksheet is missing")
            worksheet = workbook[table["worksheet"]]
            if worksheet[table["title_cell"]].value != table["title_value"]:
                raise PXD017052SourceCellAuditError("R3 PXD017052 worksheet title differs")
            header = [_csv_cell(cell.value).strip() for cell in worksheet[table["header_row"]]]
            try:
                protein_column = header.index(table["protein_identifier_column"])
            except ValueError as exc:
                raise PXD017052SourceCellAuditError("R3 PXD017052 protein column is missing") from exc
            lfq_columns = {
                value.removeprefix(table["lfq_header_prefix"]): index
                for index, value in enumerate(header)
                if value.startswith(table["lfq_header_prefix"])
            }
            if set(lfq_columns) != set(unit_map):
                raise PXD017052SourceCellAuditError("R3 PXD017052 LFQ headers do not close the unit map")
            rows: list[dict[str, str]] = []
            source_blank_count = 0
            explicit_zero_count = 0
            numeric_count = 0
            for source_row, values in enumerate(
                worksheet.iter_rows(min_row=table["header_row"] + 1, values_only=True),
                start=table["header_row"] + 1,
            ):
                protein_identifier = _csv_cell(values[protein_column]).strip()
                if not protein_identifier:
                    raise PXD017052SourceCellAuditError("R3 PXD017052 row lacks protein identifier")
                for result_unit_id, column_index in sorted(lfq_columns.items()):
                    value = values[column_index]
                    value_state = "SOURCE_BLANK" if value is None else "NUMERIC"
                    if value is None:
                        source_blank_count += 1
                    elif isinstance(value, int | float) and not isinstance(value, bool) and math.isfinite(float(value)):
                        numeric_count += 1
                        if float(value) == 0.0:
                            explicit_zero_count += 1
                    else:
                        raise PXD017052SourceCellAuditError("R3 PXD017052 LFQ value is not numeric or blank")
                    unit = unit_map[result_unit_id]
                    source_column = get_column_letter(column_index + 1)
                    rows.append(
                        {
                            "analysis_unit_id": f"PXD017052:Sheet1:{source_row}:{result_unit_id}",
                            "source_article_doi": "10.1038/s41467-020-17033-7",
                            "source_pmcid": "PMC7376165",
                            "source_license": "CC-BY-4.0",
                            "source_asset_id": table["asset_id"],
                            "source_worksheet": table["worksheet"],
                            "source_row": str(source_row),
                            "source_cell": f"Sheet1!{source_column}{source_row}",
                            "protein_source_identifier": protein_identifier,
                            "result_unit_id": result_unit_id,
                            "particle": str(unit["particle"]),
                            "assay_replicate": str(unit["assay_replicate"]),
                            "author_reported_lfq": _csv_cell(value),
                            "author_value_state": value_state,
                        }
                    )
        except OSError as exc:
            raise PXD017052SourceCellAuditError("R3 PXD017052 workbook cannot be read") from exc
        finally:
            with contextlib.suppress(UnboundLocalError):
                workbook.close()
        return rows, numeric_count, source_blank_count, explicit_zero_count

    def run(self, *, strict: bool = False) -> PXD017052SourceCellAuditSummary:
        if not strict:
            raise PXD017052SourceCellAuditError("R3 PXD017052 source-cell audit requires --strict")
        if self.output_root.exists():
            raise PXD017052SourceCellAuditError("R3 PXD017052 source-cell audit already executed")
        registry, assets, unit_map = self._registry()
        table = _mapping(registry["table"], "R3 PXD017052 table")
        rows, numeric_count, blank_count, zero_count = self._source_cells(assets[table["asset_id"]], table, unit_map)
        expected = (
            table["expected_lfq_cells"],
            table["expected_numeric_lfq_cells"],
            table["expected_source_blank_lfq_cells"],
            table["expected_explicit_zero_lfq_cells"],
        )
        observed = (len(rows), numeric_count, blank_count, zero_count)
        if len(rows) // len(unit_map) != table["expected_protein_rows"] or observed != expected:
            raise PXD017052SourceCellAuditError("R3 PXD017052 source-cell counts do not match registry")
        source_map = self.assets_root / self.DERIVED_RELATIVE
        if source_map.exists():
            raise PXD017052SourceCellAuditError("R3 PXD017052 source-cell map already exists")
        source_map.parent.mkdir(parents=True, exist_ok=True)
        with source_map.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(rows[0]), lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        scope = _mapping(registry["scope"], "R3 PXD017052 scope")
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "evaluated_at": registry["evaluated_at"],
            "registry_sha256": _sha256(self.registry_path),
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "source": registry["source"],
            "verified_source_asset_count": len(assets),
            "source_asset_sha256": {asset_id: _sha256(path) for asset_id, path in assets.items()},
            "technical_scope": scope,
            "protein_row_count": table["expected_protein_rows"],
            "result_unit_count": len(unit_map),
            "analysis_unit_count": len(rows),
            "numeric_lfq_cell_count": numeric_count,
            "source_blank_lfq_cell_count": blank_count,
            "explicit_zero_lfq_cell_count": zero_count,
            "source_to_cell_map": {
                "location": self.DERIVED_RELATIVE,
                "sha256": _sha256(source_map),
                "coordinate_definition": "Every output record retains its original LFQ cell address and explicit published unit-to-particle mapping.",  # noqa: E501
                "transformation": "openpyxl read_only/data_only extraction; no imputation, zero-to-missing conversion, abundance scaling or protein remapping.",  # noqa: E501
            },
            "status": self.STATUS,
            "target_status": "SOURCE_NATIVE_CCBY_ONLY",
            "model_fitted": False,
            "paired_ablations_run": False,
            "external_ood_evaluated": False,
            "independent_validation": False,
            "scientific_submission_ready": False,
        }
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report_sha256": hashlib.sha256(_canonical(report)).hexdigest(),
            "verified_source_asset_count": len(assets),
            "protein_row_count": table["expected_protein_rows"],
            "result_unit_count": len(unit_map),
            "analysis_unit_count": len(rows),
            "source_blank_lfq_cell_count": blank_count,
            "target_status": "SOURCE_NATIVE_CCBY_ONLY",
            "model_fitted": False,
            "scientific_submission_ready": False,
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        self._write(self.output_root / "pxd017052_source_cell_audit_report.json", report)
        self._write(self.output_root / "pxd017052_source_cell_audit_receipt.json", receipt)
        return PXD017052SourceCellAuditSummary(
            source_asset_count=len(assets),
            protein_row_count=table["expected_protein_rows"],
            result_unit_count=len(unit_map),
            analysis_unit_count=len(rows),
            source_blank_count=blank_count,
            status=self.STATUS,
            receipt_path=self.output_root / "pxd017052_source_cell_audit_receipt.json",
        )

    def verify(self) -> PXD017052SourceCellAuditSummary:
        report_path = self.output_root / "pxd017052_source_cell_audit_report.json"
        receipt_path = self.output_root / "pxd017052_source_cell_audit_receipt.json"
        report = self._json(report_path, "R3 PXD017052 source-cell report")
        receipt = self._json(receipt_path, "R3 PXD017052 source-cell receipt")
        source_map = _mapping(report.get("source_to_cell_map"), "R3 PXD017052 source map")
        source_map_path = self.assets_root / _string(source_map.get("location"), "R3 source map location")
        if (
            report.get("audit_id") != self.AUDIT_ID
            or receipt.get("audit_id") != self.AUDIT_ID
            or report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report_sha256") != _sha256(report_path)
            or report.get("target_status") != "SOURCE_NATIVE_CCBY_ONLY"
            or report.get("model_fitted") is not False
            or report.get("scientific_submission_ready") is not False
            or not source_map_path.is_file()
            or source_map.get("sha256") != _sha256(source_map_path)
        ):
            raise PXD017052SourceCellAuditError("R3 PXD017052 source-cell receipt is invalid")
        return PXD017052SourceCellAuditSummary(
            source_asset_count=int(receipt["verified_source_asset_count"]),
            protein_row_count=int(receipt["protein_row_count"]),
            result_unit_count=int(receipt["result_unit_count"]),
            analysis_unit_count=int(receipt["analysis_unit_count"]),
            source_blank_count=int(receipt["source_blank_lfq_cell_count"]),
            status=self.STATUS,
            receipt_path=receipt_path,
        )
