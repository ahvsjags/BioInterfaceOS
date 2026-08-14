"""Audit the CC-BY silver-nanoparticle human-plasma LFQ source at cell level."""

from __future__ import annotations

import csv
import json
import math
from collections import Counter
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R3SilverPlasmaSourceAuditError(RuntimeError):
    """Raised when the silver-plasma source cannot be traced to its raw cells."""


@dataclass(frozen=True)
class R3SilverPlasmaSourceAuditSummary:
    """Accounting for the one real external-laboratory source audit."""

    source_asset_count: int
    protein_row_count: int
    analysis_measurement_batch_count: int
    source_cell_count: int
    positive_source_cell_count: int
    receipt_path: Path


class R3SilverPlasmaSourceAuditWorkflow:
    """Make every admitted LFQ rank candidate traceable to its workbook cell."""

    AUDIT_ID = "bioif-r3-silver-plasma-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R3_T154_SILVER_PLASMA_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_3/silver_plasma_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R3_PMC6592156_silver_plasma_source_cell_map.csv"
    STATUS = "ADMITTED_REAL_HUMAN_PLASMA_EXTERNAL_OOD_SOURCE_PENDING_PROTOCOL_FREEZE"
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "article",
        "source_scope",
        "source_assets",
        "worksheet_contracts",
        "quantification_contract",
        "admission_minimums",
        "claim_boundary",
    }
    REQUIRED_ARTICLE = {
        "pmcid",
        "doi",
        "title",
        "publication_year",
        "license",
        "full_text_locator",
        "supplementary_locator",
    }
    REQUIRED_SCOPE = {
        "source_id",
        "laboratory_anchor",
        "biofluid",
        "nanoparticle",
        "analysis_role",
        "prohibited_interpretations",
    }
    REQUIRED_ASSET = {"asset_id", "relative_path", "sha256", "expected_bytes"}
    REQUIRED_WORKSHEET = {
        "worksheet",
        "expected_protein_rows",
        "condition_kind",
        "condition_count",
        "replicate_labels",
        "excluded_column_label",
    }
    REQUIRED_QUANTIFICATION = {
        "source_accession_column",
        "author_quantity_type",
        "rank_eligibility",
        "numeric_zero_policy",
        "blank_policy",
        "control_policy",
        "raw_scale_cross_study_use",
    }
    REQUIRED_MINIMUMS = {
        "analysis_measurement_batch_count",
        "minimum_positive_proteins_per_measurement_batch",
        "independent_laboratory_anchor_count",
    }
    SOURCE_CELL_FIELDS = [
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "condition_kind",
        "condition_label",
        "replicate_label",
        "measurement_batch_id",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    ]

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = (output_root or self.root / self.OUTPUT_RELATIVE).resolve(strict=False)

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            value = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise R3SilverPlasmaSourceAuditError(f"cannot parse {label}") from exc
        try:
            return _mapping(value, label)
        except Exception as exc:
            raise R3SilverPlasmaSourceAuditError(f"cannot parse {label}") from exc

    def _asset_path(self, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise R3SilverPlasmaSourceAuditError(f"{label} must use POSIX path separators")
        pure = PurePosixPath(relative_path)
        if pure.is_absolute() or not pure.parts or ".." in pure.parts:
            raise R3SilverPlasmaSourceAuditError(f"{label} escapes source asset root")
        path = (self.assets_root / Path(*pure.parts)).resolve(strict=False)
        if not path.is_relative_to(self.assets_root) or not path.is_file():
            raise R3SilverPlasmaSourceAuditError(f"{label} is missing or outside source asset root")
        return path

    @staticmethod
    def _normalize(value: Any) -> str:
        if not isinstance(value, str):
            return ""
        return " ".join(value.replace("˚", "°").split())

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path]]:
        registry = self._json(self.registry_path, "R3 silver-plasma source registry")
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise R3SilverPlasmaSourceAuditError("silver-plasma registry fields are invalid")
        if (
            registry.get("audit_id") != self.AUDIT_ID
            or registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise R3SilverPlasmaSourceAuditError("silver-plasma registry evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "silver-plasma registry evaluated_at")
        _string(registry.get("claim_boundary"), "silver-plasma registry claim boundary")
        article = _mapping(registry.get("article"), "silver-plasma article")
        if set(article) != self.REQUIRED_ARTICLE or article != {
            "pmcid": "PMC6592156",
            "doi": "10.1039/c8en01054d",
            "title": "Protein corona formed on silver nanoparticles in blood plasma is highly selective and resistant to physicochemical changes of the solution",
            "publication_year": 2019,
            "license": "CC-BY-3.0",
            "full_text_locator": "https://europepmc.org/articles/PMC6592156",
            "supplementary_locator": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC6592156/supplementaryFiles",
        }:
            raise R3SilverPlasmaSourceAuditError("silver-plasma article declaration is invalid")
        scope = _mapping(registry.get("source_scope"), "silver-plasma source scope")
        if set(scope) != self.REQUIRED_SCOPE or (
            scope.get("source_id") != "PMC6592156_SILVER_NANOPARTICLE_HUMAN_PLASMA"
            or scope.get("laboratory_anchor")
            != "University of Southern Denmark / Russian Academy of Sciences study"
            or scope.get("biofluid") != "human blood plasma"
            or scope.get("nanoparticle") != "60 nm silver nanoparticles"
            or scope.get("analysis_role") != "EXTERNAL_LAB_OOD_CANDIDATE_ONLY"
            or not isinstance(scope.get("prohibited_interpretations"), list)
            or len(scope["prohibited_interpretations"]) != 3
        ):
            raise R3SilverPlasmaSourceAuditError("silver-plasma source scope is invalid")
        assets = registry.get("source_assets")
        if not isinstance(assets, list) or len(assets) != 2:
            raise R3SilverPlasmaSourceAuditError("silver-plasma source assets are invalid")
        asset_paths: dict[str, Path] = {}
        for asset in assets:
            item = _mapping(asset, "silver-plasma source asset")
            if set(item) != self.REQUIRED_ASSET:
                raise R3SilverPlasmaSourceAuditError("silver-plasma source asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "silver-plasma source asset ID")
            path = self._asset_path(_string(item.get("relative_path"), asset_id), asset_id)
            expected_bytes = item.get("expected_bytes")
            if (
                isinstance(expected_bytes, bool)
                or not isinstance(expected_bytes, int)
                or expected_bytes <= 0
                or path.stat().st_size != expected_bytes
                or _sha256(path) != _checksum(item.get("sha256"), asset_id)
                or asset_id in asset_paths
            ):
                raise R3SilverPlasmaSourceAuditError("silver-plasma source asset checksum differs")
            asset_paths[asset_id] = path
        if set(asset_paths) != {"supplementary_package_zip", "supplementary_data_2_lfq_xlsx"}:
            raise R3SilverPlasmaSourceAuditError("silver-plasma source assets are incomplete")
        sheets = registry.get("worksheet_contracts")
        if not isinstance(sheets, list) or len(sheets) != 2:
            raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet contracts are invalid")
        expected_sheets = {
            "pH": (387, "PH"),
            "Temperature": (512, "TEMPERATURE"),
        }
        for item in sheets:
            contract = _mapping(item, "silver-plasma worksheet contract")
            if set(contract) != self.REQUIRED_WORKSHEET:
                raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet fields are invalid")
            name = _string(contract.get("worksheet"), "silver-plasma worksheet")
            expected = expected_sheets.pop(name, None)
            if (
                expected is None
                or contract.get("expected_protein_rows") != expected[0]
                or contract.get("condition_kind") != expected[1]
                or contract.get("condition_count") != 5
                or contract.get("replicate_labels") != ["Rep1", "Rep2", "Rep3"]
                or contract.get("excluded_column_label") != "Control"
            ):
                raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet contract is invalid")
        if expected_sheets:
            raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet contracts are incomplete")
        quantification = _mapping(registry.get("quantification_contract"), "silver-plasma quantification")
        if set(quantification) != self.REQUIRED_QUANTIFICATION or quantification != {
            "source_accession_column": "Uniprot Accession",
            "author_quantity_type": "LOG_CONVERTED_LFQ_ABUNDANCE",
            "rank_eligibility": "strictly positive finite author-reported value in one Rep1/Rep2/Rep3 column",
            "numeric_zero_policy": "retain as NUMERIC_ZERO_SEMANTICS_UNSPECIFIED and exclude from rank; never impute",
            "blank_policy": "retain as SOURCE_BLANK and exclude from rank; never impute",
            "control_policy": "exclude Control columns from modelable measurement batches; retain their existence in the audit report",
            "raw_scale_cross_study_use": "PROHIBITED",
        }:
            raise R3SilverPlasmaSourceAuditError("silver-plasma quantification contract is invalid")
        minimums = _mapping(registry.get("admission_minimums"), "silver-plasma admission minimums")
        if set(minimums) != self.REQUIRED_MINIMUMS or minimums != {
            "analysis_measurement_batch_count": 30,
            "minimum_positive_proteins_per_measurement_batch": 10,
            "independent_laboratory_anchor_count": 1,
        }:
            raise R3SilverPlasmaSourceAuditError("silver-plasma admission minimums are invalid")
        return registry, asset_paths

    @staticmethod
    def _column_letter(index: int) -> str:
        value = index
        letters = ""
        while value:
            value, remainder = divmod(value - 1, 26)
            letters = chr(ord("A") + remainder) + letters
        return letters

    def _source_rows(self, registry: Mapping[str, Any], workbook_path: Path) -> tuple[list[dict[str, str]], dict[str, Any]]:
        contracts = {
            item["worksheet"]: item for item in registry["worksheet_contracts"]
        }
        source_scope = registry["source_scope"]
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        mapped_rows: list[dict[str, str]] = []
        seen: set[tuple[str, str]] = set()
        source_cell_counts: Counter[str] = Counter()
        positive_counts: Counter[str] = Counter()
        zero_counts: Counter[str] = Counter()
        blank_counts: Counter[str] = Counter()
        control_column_count = 0
        try:
            if set(workbook.sheetnames) != set(contracts):
                raise R3SilverPlasmaSourceAuditError("silver-plasma workbook worksheets differ")
            for sheet_name in ("pH", "Temperature"):
                sheet = workbook[sheet_name]
                contract = contracts[sheet_name]
                if sheet.max_row - 2 != contract["expected_protein_rows"] or sheet.max_column != 32:
                    raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet dimensions differ")
                row_one = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                row_two = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
                if self._normalize(row_one[0]) != "Uniprot Accession" or self._normalize(row_one[1]) != "GeneID":
                    raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet header is invalid")
                condition = ""
                batch_columns: list[tuple[int, str, str]] = []
                for zero_index in range(2, 22):
                    if row_one[zero_index] is not None:
                        condition = self._normalize(row_one[zero_index])
                    label = self._normalize(row_two[zero_index])
                    if label in contract["replicate_labels"]:
                        if not condition:
                            raise R3SilverPlasmaSourceAuditError("silver-plasma condition header is missing")
                        batch_columns.append((zero_index + 1, condition, label))
                    elif label == contract["excluded_column_label"]:
                        control_column_count += 1
                    elif label:
                        raise R3SilverPlasmaSourceAuditError("silver-plasma measurement-column label is unknown")
                if len(batch_columns) != contract["condition_count"] * len(contract["replicate_labels"]):
                    raise R3SilverPlasmaSourceAuditError("silver-plasma batch-column count differs")
                for source_row in range(3, sheet.max_row + 1):
                    values = [cell.value for cell in sheet[source_row]]
                    accession = self._normalize(values[0])
                    if not accession:
                        raise R3SilverPlasmaSourceAuditError("silver-plasma row lacks a UniProt accession")
                    identity = (sheet_name, accession)
                    if identity in seen:
                        raise R3SilverPlasmaSourceAuditError("silver-plasma worksheet repeats a UniProt accession")
                    seen.add(identity)
                    for column, condition_label, replicate_label in batch_columns:
                        value = values[column - 1]
                        batch_id = (
                            f"PMC6592156:{contract['condition_kind']}:{condition_label}:{replicate_label}"
                        )
                        state: str
                        numeric_value = ""
                        eligible = "false"
                        if value is None or (isinstance(value, str) and not value.strip()):
                            state = "SOURCE_BLANK"
                        elif isinstance(value, bool) or not isinstance(value, (int, float)):
                            raise R3SilverPlasmaSourceAuditError("silver-plasma value is non-numeric")
                        elif not math.isfinite(float(value)):
                            raise R3SilverPlasmaSourceAuditError("silver-plasma value is non-finite")
                        else:
                            numeric_value = format(float(value), ".17g")
                            if float(value) > 0.0:
                                state = "POSITIVE_LOG_LFQ"
                                eligible = "true"
                                positive_counts[batch_id] += 1
                            elif float(value) == 0.0:
                                state = "NUMERIC_ZERO_SEMANTICS_UNSPECIFIED"
                                zero_counts[batch_id] += 1
                            else:
                                raise R3SilverPlasmaSourceAuditError("silver-plasma log LFQ value is negative")
                        if state == "SOURCE_BLANK":
                            blank_counts[batch_id] += 1
                        source_cell_counts[batch_id] += 1
                        mapped_rows.append(
                            {
                                "source_id": source_scope["source_id"],
                                "laboratory_anchor": source_scope["laboratory_anchor"],
                                "source_asset_id": "supplementary_data_2_lfq_xlsx",
                                "source_worksheet": sheet_name,
                                "source_row": str(source_row),
                                "source_coordinate": f"{sheet_name}!{self._column_letter(column)}{source_row}",
                                "source_identifier": accession,
                                "condition_kind": contract["condition_kind"],
                                "condition_label": condition_label,
                                "replicate_label": replicate_label,
                                "measurement_batch_id": batch_id,
                                "author_quantity_type": "LOG_CONVERTED_LFQ_ABUNDANCE",
                                "author_numeric_value": numeric_value,
                                "author_value_state": state,
                                "rank_target_eligible": eligible,
                            }
                        )
        finally:
            workbook.close()
        expected_batches = registry["admission_minimums"]["analysis_measurement_batch_count"]
        if len(source_cell_counts) != expected_batches or control_column_count != 10:
            raise R3SilverPlasmaSourceAuditError("silver-plasma batch/control accounting is invalid")
        minimum_positive = registry["admission_minimums"]["minimum_positive_proteins_per_measurement_batch"]
        if any(count < minimum_positive for count in positive_counts.values()):
            raise R3SilverPlasmaSourceAuditError("silver-plasma batch has insufficient positive proteins")
        summary = {
            "protein_row_count": len(seen),
            "analysis_measurement_batch_count": len(source_cell_counts),
            "control_column_count": control_column_count,
            "source_cell_count": len(mapped_rows),
            "positive_source_cell_count": sum(positive_counts.values()),
            "zero_source_cell_count": sum(zero_counts.values()),
            "blank_source_cell_count": sum(blank_counts.values()),
            "batch_accounting": [
                {
                    "measurement_batch_id": batch_id,
                    "source_cell_count": source_cell_counts[batch_id],
                    "positive_source_cell_count": positive_counts[batch_id],
                    "numeric_zero_source_cell_count": zero_counts[batch_id],
                    "blank_source_cell_count": blank_counts[batch_id],
                }
                for batch_id in sorted(source_cell_counts)
            ],
        }
        return mapped_rows, summary

    def run(self, *, strict: bool = False) -> R3SilverPlasmaSourceAuditSummary:
        if not strict:
            raise R3SilverPlasmaSourceAuditError("silver-plasma source audit requires --strict")
        if self.output_root.exists():
            raise R3SilverPlasmaSourceAuditError("silver-plasma source audit already executed")
        registry, assets = self._registry()
        mapped_rows, summary = self._source_rows(registry, assets["supplementary_data_2_lfq_xlsx"])
        source_map_path = self.assets_root / self.DERIVED_RELATIVE
        if source_map_path.exists():
            with source_map_path.open("r", encoding="utf-8", newline="") as stream:
                existing_rows = list(csv.DictReader(stream))
            if existing_rows != mapped_rows:
                raise R3SilverPlasmaSourceAuditError(
                    "silver-plasma existing source cell map differs from reacquired source"
                )
        else:
            source_map_path.parent.mkdir(parents=True, exist_ok=True)
            with source_map_path.open("w", encoding="utf-8", newline="") as stream:
                writer = csv.DictWriter(
                    stream, fieldnames=self.SOURCE_CELL_FIELDS, lineterminator="\n"
                )
                writer.writeheader()
                writer.writerows(mapped_rows)
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "evaluated_at": registry["evaluated_at"],
            "registry_sha256": _sha256(self.registry_path),
            "evidence_class": registry["evidence_class"],
            "allowed_claim_level": registry["allowed_claim_level"],
            "article": registry["article"],
            "source_scope": registry["source_scope"],
            "quantification_contract": registry["quantification_contract"],
            "source_assets": [
                {
                    **asset,
                    "observed_bytes": assets[asset["asset_id"]].stat().st_size,
                }
                for asset in registry["source_assets"]
            ],
            "source_cell_map": {
                "location": self.DERIVED_RELATIVE,
                "sha256": _sha256(source_map_path),
            },
            **summary,
            "status": self.STATUS,
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "claim_boundary": registry["claim_boundary"],
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        report_path = self.output_root / "silver_plasma_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report_sha256": _sha256(report_path),
            "source_asset_count": len(assets),
            "protein_row_count": summary["protein_row_count"],
            "analysis_measurement_batch_count": summary["analysis_measurement_batch_count"],
            "source_cell_count": summary["source_cell_count"],
            "positive_source_cell_count": summary["positive_source_cell_count"],
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "silver_plasma_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R3SilverPlasmaSourceAuditSummary(
            source_asset_count=len(assets),
            protein_row_count=summary["protein_row_count"],
            analysis_measurement_batch_count=summary["analysis_measurement_batch_count"],
            source_cell_count=summary["source_cell_count"],
            positive_source_cell_count=summary["positive_source_cell_count"],
            receipt_path=receipt_path,
        )

    def verify(self) -> R3SilverPlasmaSourceAuditSummary:
        report_path = self.output_root / "silver_plasma_source_audit_report.json"
        receipt_path = self.output_root / "silver_plasma_source_audit_receipt.json"
        report = self._json(report_path, "silver-plasma source audit report")
        receipt = self._json(receipt_path, "silver-plasma source audit receipt")
        source_map = _mapping(report.get("source_cell_map"), "silver-plasma source cell map")
        path = self.assets_root / _string(source_map.get("location"), "silver-plasma source cell map path")
        if (
            report.get("audit_id") != self.AUDIT_ID
            or receipt.get("audit_id") != self.AUDIT_ID
            or report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report_sha256") != _sha256(report_path)
            or not path.is_file()
            or source_map.get("sha256") != _sha256(path)
            or receipt.get("model_fitted") is not False
            or receipt.get("independent_validation") is not False
            or receipt.get("external_scientific_reproduction") is not False
            or receipt.get("scientific_submission_ready") is not False
        ):
            raise R3SilverPlasmaSourceAuditError("silver-plasma source audit receipt is invalid")
        return R3SilverPlasmaSourceAuditSummary(
            source_asset_count=int(receipt["source_asset_count"]),
            protein_row_count=int(receipt["protein_row_count"]),
            analysis_measurement_batch_count=int(receipt["analysis_measurement_batch_count"]),
            source_cell_count=int(receipt["source_cell_count"]),
            positive_source_cell_count=int(receipt["positive_source_cell_count"]),
            receipt_path=receipt_path,
        )
