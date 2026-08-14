"""Audit a CC-BY clinical plasma-proteomics source without silently merging targets.

The Edinburgh source used nanoparticle enrichment before LC--MS/MS, but its
measurement estimand (source-local protein abundance before/after controlled
inhalation) is not the frozen R3 corona-rank target.  This module therefore
only creates a byte-traceable, separately labelled R4 candidate ledger.
"""

from __future__ import annotations

import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R4EdinburghClinicalSourceAuditError(RuntimeError):
    """Raised when the public clinical source cannot be traced to raw cells."""


@dataclass(frozen=True)
class R4EdinburghClinicalSourceAuditSummary:
    """Compact accounting for the separately scoped external-source candidate."""

    source_asset_count: int
    protein_row_count: int
    measurement_batch_count: int
    shared_canonical_protein_count: int
    source_cell_count: int
    positive_source_cell_count: int
    receipt_path: Path


class R4EdinburghClinicalSourceAuditWorkflow:
    """Create a no-imputation cell map for the CC-BY Edinburgh source."""

    AUDIT_ID = "bioif-r4-edinburgh-clinical-plasma-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R4_T156_EDINBURGH_CLINICAL_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_4/edinburgh_clinical_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R4_EDINBURGH_DS7545_source_cell_map.csv"
    STATUS = "ADMITTED_SEPARATE_R4_CLINICAL_EXTERNAL_CANDIDATE_ONLY"
    SOURCE_CELL_FIELDS = (
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "canonical_accession",
        "sample_phase",
        "measurement_batch_id",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    )
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "article",
        "dataset",
        "source_scope",
        "source_assets",
        "r3_reference_assets",
        "worksheet_contracts",
        "quantification_contract",
        "admission_minimums",
        "claim_boundary",
    }

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise R4EdinburghClinicalSourceAuditError(f"cannot parse {label}") from exc

    @staticmethod
    def _number(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, bool) or not isinstance(value, int | float):
            raise R4EdinburghClinicalSourceAuditError("source abundance must be numeric or blank")
        numeric = float(value)
        if not math.isfinite(numeric):
            raise R4EdinburghClinicalSourceAuditError("source abundance must be finite")
        return numeric

    def _under(self, root: Path, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise R4EdinburghClinicalSourceAuditError(f"{label} must use POSIX path separators")
        pure = PurePosixPath(relative_path)
        if pure.is_absolute() or not pure.parts or ".." in pure.parts:
            raise R4EdinburghClinicalSourceAuditError(f"{label} escapes its root")
        path = (root / Path(*pure.parts)).resolve(strict=False)
        if not path.is_relative_to(root) or not path.is_file():
            raise R4EdinburghClinicalSourceAuditError(f"{label} is missing or outside its root")
        return path

    def _checked_assets(self, registry: dict[str, Any]) -> tuple[Path, dict[str, Path], dict[str, Path]]:
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise R4EdinburghClinicalSourceAuditError("registry fields are invalid")
        if registry.get("audit_id") != self.AUDIT_ID:
            raise R4EdinburghClinicalSourceAuditError("registry audit ID is invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise R4EdinburghClinicalSourceAuditError("registry evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "evaluated_at")
        _string(registry.get("claim_boundary"), "claim_boundary")

        article = _mapping(registry.get("article"), "article")
        expected_article = {
            "pmcid": "PMC11106005",
            "doi": "10.1038/s41565-023-01572-3",
            "title": "First-in-human controlled inhalation of thin graphene oxide nanosheets to study acute cardiorespiratory responses",  # noqa: E501
            "publication_year": 2024,
            "license": "CC-BY-4.0",
            "full_text_locator": "https://europepmc.org/articles/PMC11106005",
        }
        if article != expected_article:
            raise R4EdinburghClinicalSourceAuditError("article declaration is invalid")
        dataset = _mapping(registry.get("dataset"), "dataset")
        expected_dataset = {
            "doi": "10.7488/ds/7545",
            "handle": "10283/8569",
            "license": "CC-BY-4.0",
            "repository_locator": "https://datashare.ed.ac.uk/handle/10283/8569",
        }
        if dataset != expected_dataset:
            raise R4EdinburghClinicalSourceAuditError("dataset declaration is invalid")
        scope = _mapping(registry.get("source_scope"), "source scope")
        if (
            set(scope)
            != {
                "source_id",
                "laboratory_anchor",
                "biofluid",
                "nanoparticle_enrichment",
                "analysis_role",
                "prohibited_interpretations",
            }
            or scope.get("source_id") != "EDINBURGH_DS7545_HUMAN_PLASMA_NANOOMICS"
            or scope.get("laboratory_anchor") != "University of Edinburgh-led controlled human exposure study"
            or scope.get("biofluid") != "human plasma"
            or scope.get("nanoparticle_enrichment") != "lipid-nanoparticle NanoOmics plasma enrichment"
            or scope.get("analysis_role") != "SEPARATE_R4_EXTERNAL_OOD_CANDIDATE_ONLY"
            or not isinstance(scope.get("prohibited_interpretations"), list)
            or len(scope["prohibited_interpretations"]) != 4
        ):
            raise R4EdinburghClinicalSourceAuditError("source scope is invalid")

        assets = registry.get("source_assets")
        if not isinstance(assets, list) or len(assets) != 2:
            raise R4EdinburghClinicalSourceAuditError("source assets are invalid")
        asset_paths: dict[str, Path] = {}
        for item in assets:
            item = _mapping(item, "source asset")
            if set(item) != {
                "asset_id",
                "relative_path",
                "sha256",
                "expected_bytes",
                "repository_md5",
            }:
                raise R4EdinburghClinicalSourceAuditError("source asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "source asset ID")
            path = self._under(self.assets_root, _string(item.get("relative_path"), asset_id), asset_id)
            if (
                asset_id in asset_paths
                or not isinstance(item.get("expected_bytes"), int)
                or isinstance(item.get("expected_bytes"), bool)
                or path.stat().st_size != item["expected_bytes"]
                or _sha256(path) != _checksum(item.get("sha256"), asset_id)
                or len(_string(item.get("repository_md5"), asset_id)) != 32
            ):
                raise R4EdinburghClinicalSourceAuditError("source asset checksum differs")
            asset_paths[asset_id] = path
        if set(asset_paths) != {"blood_proteomics_xlsx", "dataset_readme"}:
            raise R4EdinburghClinicalSourceAuditError("source assets are incomplete")

        references = registry.get("r3_reference_assets")
        if not isinstance(references, list) or len(references) != 3:
            raise R4EdinburghClinicalSourceAuditError("R3 reference assets are invalid")
        reference_paths: dict[str, Path] = {}
        for item in references:
            item = _mapping(item, "R3 reference asset")
            if set(item) != {"asset_id", "relative_path", "sha256"}:
                raise R4EdinburghClinicalSourceAuditError("R3 reference asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "R3 reference asset ID")
            path = self._under(self.root, _string(item.get("relative_path"), asset_id), asset_id)
            if asset_id in reference_paths or _sha256(path) != _checksum(item.get("sha256"), asset_id):
                raise R4EdinburghClinicalSourceAuditError("R3 reference asset checksum differs")
            reference_paths[asset_id] = path
        if set(reference_paths) != {"feature_table", "fasta_batch_0001", "fasta_batch_0002"}:
            raise R4EdinburghClinicalSourceAuditError("R3 reference assets are incomplete")

        contracts = registry.get("worksheet_contracts")
        expected_contracts = {
            "sGO 0h v 6h": (702, 25, 23, 19),
            "usGO 0h vs 6h": (782, 28, 26, 21),
        }
        if not isinstance(contracts, list) or len(contracts) != len(expected_contracts):
            raise R4EdinburghClinicalSourceAuditError("worksheet contracts are invalid")
        for item in contracts:
            item = _mapping(item, "worksheet contract")
            if set(item) != {
                "worksheet",
                "expected_rows",
                "expected_columns",
                "sample_column_count",
                "unambiguous_shared_protein_rows",
            }:
                raise R4EdinburghClinicalSourceAuditError("worksheet contract fields are invalid")
            worksheet_name = _string(item.get("worksheet"), "worksheet contract name")
            expected = expected_contracts.pop(worksheet_name, None)
            if (
                expected is None
                or tuple(
                    item[key]
                    for key in (
                        "expected_rows",
                        "expected_columns",
                        "sample_column_count",
                        "unambiguous_shared_protein_rows",
                    )
                )
                != expected
            ):
                raise R4EdinburghClinicalSourceAuditError("worksheet contract differs")
        if expected_contracts:
            raise R4EdinburghClinicalSourceAuditError("worksheet contracts are incomplete")
        quantification = _mapping(registry.get("quantification_contract"), "quantification contract")
        if quantification != {
            "source_accession_column": "Accession",
            "author_quantity_type": "NORMALIZED_ABUNDANCE",
            "entry_name_mapping": "map only a source row with exactly one R3 FASTA entry-name match; exclude all zero-or-many-match rows",  # noqa: E501
            "rank_eligibility": "strictly positive finite author-reported abundance within one sample column",
            "numeric_zero_policy": "retain as NUMERIC_ZERO and exclude from rank; never impute",
            "blank_policy": "retain as SOURCE_BLANK and exclude from rank; never impute",
            "raw_scale_cross_study_use": "PROHIBITED",
        }:
            raise R4EdinburghClinicalSourceAuditError("quantification contract is invalid")
        if _mapping(registry.get("admission_minimums"), "admission minimums") != {
            "measurement_batch_count": 49,
            "minimum_positive_shared_proteins_per_measurement_batch": 10,
            "independent_laboratory_anchor_count": 1,
        }:
            raise R4EdinburghClinicalSourceAuditError("admission minimums are invalid")
        return asset_paths["blood_proteomics_xlsx"], asset_paths, reference_paths

    @staticmethod
    def _reference_mapping(feature_path: Path, fasta_paths: list[Path]) -> tuple[set[str], dict[str, str]]:
        with feature_path.open(newline="", encoding="utf-8") as stream:
            features = {row["canonical_accession"] for row in csv.DictReader(stream)}
        entry_to_accession: dict[str, str] = {}
        for path in fasta_paths:
            for line in path.read_text(encoding="utf-8").splitlines():
                if not line.startswith(">"):
                    continue
                parts = line[1:].split("|", 2)
                if len(parts) != 3 or not parts[1] or not parts[2].split():
                    raise R4EdinburghClinicalSourceAuditError("R3 FASTA header is invalid")
                entry_name = parts[2].split()[0]
                if entry_name in entry_to_accession and entry_to_accession[entry_name] != parts[1]:
                    raise R4EdinburghClinicalSourceAuditError("R3 FASTA entry name is ambiguous")
                entry_to_accession[entry_name] = parts[1]
        if not features or not entry_to_accession:
            raise R4EdinburghClinicalSourceAuditError("R3 reference assets are empty")
        return features, entry_to_accession

    def _cells(
        self, workbook_path: Path, reference_paths: dict[str, Path]
    ) -> tuple[list[dict[str, Any]], dict[str, int]]:
        features, entry_to_accession = self._reference_mapping(
            reference_paths["feature_table"],
            [reference_paths["fasta_batch_0001"], reference_paths["fasta_batch_0002"]],
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        expected = {"sGO 0h v 6h": (702, 25), "usGO 0h vs 6h": (782, 28)}
        if set(workbook.sheetnames) != {"sGO 0h v 6h", "usGO 0h vs 6h", "identification"}:
            raise R4EdinburghClinicalSourceAuditError("workbook sheets differ")
        cells: list[dict[str, Any]] = []
        protein_rows = 0
        per_sheet_shared: dict[str, int] = {}
        for sheet_name, (row_count, column_count) in expected.items():
            sheet = workbook[sheet_name]
            if sheet.max_row != row_count or sheet.max_column != column_count:
                raise R4EdinburghClinicalSourceAuditError("workbook dimensions differ")
            phase_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            header = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
            if header[0:2] != ("Accession", "Description") or any(
                not isinstance(value, str) or not value for value in header[2:]
            ):
                raise R4EdinburghClinicalSourceAuditError("workbook sample header differs")
            seen_accessions: set[str] = set()
            shared_rows = 0
            protein_rows += sheet.max_row - 3
            for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                identifiers = [token.strip() for token in str(row[0] or "").split(";") if token.strip()]
                mapped = {entry_to_accession[token] for token in identifiers if token in entry_to_accession}
                if len(mapped) != 1:
                    continue
                canonical_accession = next(iter(mapped))
                if canonical_accession not in features:
                    raise R4EdinburghClinicalSourceAuditError("mapped entry is absent from R3 feature table")
                if canonical_accession in seen_accessions:
                    raise R4EdinburghClinicalSourceAuditError("a measurement batch would collapse multiple source rows")
                seen_accessions.add(canonical_accession)
                shared_rows += 1
                current_phase = ""
                for column_index in range(2, sheet.max_column):
                    numeric = self._number(row[column_index])
                    if numeric is None:
                        state, eligible, rendered = "SOURCE_BLANK", False, ""
                    elif numeric == 0.0:
                        state, eligible, rendered = "NUMERIC_ZERO", False, "0"
                    elif numeric > 0.0:
                        state, eligible, rendered = "POSITIVE_FINITE", True, format(numeric, ".17g")
                    else:
                        state, eligible, rendered = (
                            "NEGATIVE_FINITE",
                            False,
                            format(numeric, ".17g"),
                        )
                    declared_phase = str(phase_row[column_index] or "").strip()
                    if declared_phase:
                        current_phase = declared_phase
                    if current_phase not in {"before exposure", "after exposure"}:
                        raise R4EdinburghClinicalSourceAuditError("sample phase is invalid")
                    sample_label = header[column_index]
                    cells.append(
                        {
                            "source_id": "EDINBURGH_DS7545_HUMAN_PLASMA_NANOOMICS",
                            "laboratory_anchor": "University of Edinburgh-led controlled human exposure study",
                            "source_asset_id": "blood_proteomics_xlsx",
                            "source_worksheet": sheet_name,
                            "source_row": str(row_number),
                            "source_coordinate": f"{get_column_letter(column_index + 1)}{row_number}",
                            "source_identifier": str(row[0]),
                            "canonical_accession": canonical_accession,
                            "sample_phase": current_phase,
                            "measurement_batch_id": f"R4_EDINBURGH_{'SGO' if sheet_name.startswith('sGO') else 'USGO'}_{sample_label}",  # noqa: E501
                            "author_quantity_type": "NORMALIZED_ABUNDANCE",
                            "author_numeric_value": rendered,
                            "author_value_state": state,
                            "rank_target_eligible": "true" if eligible else "false",
                        }
                    )
            per_sheet_shared[sheet_name] = shared_rows
        if per_sheet_shared != {"sGO 0h v 6h": 19, "usGO 0h vs 6h": 21}:
            raise R4EdinburghClinicalSourceAuditError("shared R3 source-row count differs")
        batches = {row["measurement_batch_id"] for row in cells}
        canonical = {row["canonical_accession"] for row in cells}
        positive_by_batch = {
            batch: sum(row["rank_target_eligible"] == "true" for row in cells if row["measurement_batch_id"] == batch)
            for batch in batches
        }
        if len(batches) != 49 or len(canonical) != 23 or min(positive_by_batch.values()) < 10:
            raise R4EdinburghClinicalSourceAuditError("external candidate admission minimum is not met")
        return cells, {
            "protein_rows": protein_rows,
            "batches": len(batches),
            "canonical": len(canonical),
            "positive": sum(row["rank_target_eligible"] == "true" for row in cells),
        }

    def run(self, *, strict: bool = False) -> R4EdinburghClinicalSourceAuditSummary:
        if not strict:
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh source audit requires --strict")
        if self.output_root.exists():
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh source audit already executed")
        registry = self._json(self.registry_path, "R4 Edinburgh registry")
        workbook_path, asset_paths, reference_paths = self._checked_assets(registry)
        cells, totals = self._cells(workbook_path, reference_paths)
        derived = self.assets_root / self.DERIVED_RELATIVE
        if derived.exists():
            raise R4EdinburghClinicalSourceAuditError("derived R4 Edinburgh source map already exists")
        derived.parent.mkdir(parents=True, exist_ok=True)
        with derived.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=self.SOURCE_CELL_FIELDS)
            writer.writeheader()
            writer.writerows(cells)
        self.output_root.mkdir(parents=True, exist_ok=False)
        source_asset_references = {
            item["asset_id"]: {
                "relative_path": item["relative_path"],
                "sha256": _sha256(asset_paths[item["asset_id"]]),
            }
            for item in registry["source_assets"]
        }
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "source_assets": source_asset_references,
            "r3_reference_assets": {
                asset_id: {
                    "relative_path": path.relative_to(self.root).as_posix(),
                    "sha256": _sha256(path),
                }
                for asset_id, path in reference_paths.items()
            },
            "source_cell_map": {"relative_path": self.DERIVED_RELATIVE, "sha256": _sha256(derived)},
            "protein_row_count": totals["protein_rows"],
            "measurement_batch_count": totals["batches"],
            "shared_canonical_protein_count": totals["canonical"],
            "source_cell_count": len(cells),
            "positive_source_cell_count": totals["positive"],
            "claim_boundary": registry["claim_boundary"],
        }
        report_path = self.output_root / "edinburgh_clinical_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report": {"relative_path": report_path.name, "sha256": _sha256(report_path)},
            "source_cell_map": report["source_cell_map"],
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "edinburgh_clinical_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R4EdinburghClinicalSourceAuditSummary(
            len(asset_paths),
            totals["protein_rows"],
            totals["batches"],
            totals["canonical"],
            len(cells),
            totals["positive"],
            receipt_path,
        )

    def verify(self) -> R4EdinburghClinicalSourceAuditSummary:
        registry = self._json(self.registry_path, "R4 Edinburgh registry")
        _, asset_paths, reference_paths = self._checked_assets(registry)
        report_path = self.output_root / "edinburgh_clinical_source_audit_report.json"
        receipt_path = self.output_root / "edinburgh_clinical_source_audit_receipt.json"
        report = self._json(report_path, "R4 Edinburgh audit report")
        receipt = self._json(receipt_path, "R4 Edinburgh audit receipt")
        if (
            report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report", {}).get("sha256") != _sha256(report_path)
        ):
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh audit receipt differs")
        cell_map = self._under(
            self.assets_root,
            _string(report.get("source_cell_map", {}).get("relative_path"), "source cell map path"),
            "source cell map",
        )
        if report["source_cell_map"].get("sha256") != _sha256(cell_map):
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh source cell map differs")
        with cell_map.open(newline="", encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        if (set(rows[0]) if rows else set()) != set(self.SOURCE_CELL_FIELDS):
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh source cell map fields differ")
        summary = R4EdinburghClinicalSourceAuditSummary(
            len(asset_paths),
            int(report["protein_row_count"]),
            int(report["measurement_batch_count"]),
            int(report["shared_canonical_protein_count"]),
            int(report["source_cell_count"]),
            int(report["positive_source_cell_count"]),
            receipt_path,
        )
        if (
            len(rows) != summary.source_cell_count
            or sum(row["rank_target_eligible"] == "true" for row in rows) != summary.positive_source_cell_count
        ):
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh source cell counts differ")
        if report.get("r3_reference_assets") != {
            asset_id: {
                "relative_path": path.relative_to(self.root).as_posix(),
                "sha256": _sha256(path),
            }
            for asset_id, path in reference_paths.items()
        }:
            raise R4EdinburghClinicalSourceAuditError("R4 Edinburgh R3 references differ")
        return summary
