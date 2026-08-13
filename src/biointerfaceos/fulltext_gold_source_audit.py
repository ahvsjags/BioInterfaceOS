"""Audit CC-BY human-plasma gold-nanoparticle supplementary tables.

The source tables are positive/rank-only evidence.  This workflow preserves
the author-reported zero values but never converts an omitted row into a
non-detection label.
"""

from __future__ import annotations

import csv
import hashlib
import json
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

from biointerfaceos.fulltext_multicore_audit import (
    _canonical,
    _checksum,
    _csv_cell,
    _finite,
    _list,
    _mapping,
    _sha256,
    _string,
)


class FulltextGoldSourceAuditError(RuntimeError):
    """Raised when a gold-nanoparticle full-text source is over-promoted."""


@dataclass(frozen=True)
class FulltextGoldSourceAuditSummary:
    """Accounting for the source-native, rank-only gold data package."""

    source_asset_count: int
    table_count: int
    analysis_unit_count: int
    explicit_zero_count: int
    status: str
    receipt_path: Path


class FulltextGoldSourceAuditWorkflow:
    """Build source-to-cell records for one CC-BY human-plasma GNP study."""

    AUDIT_ID = "bioif-r3-fulltext-gold-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R3_T146_FULLTEXT_GOLD_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_3/fulltext_gold_source_audit/v1.0.0"
    DERIVED_RELATIVE = "PMC7788026_derived/R3_PMC7788026_source_cell_map.csv"
    STATUS = "ADMITTED_SOURCE_NATIVE_RANK_ONLY_EXTERNAL_CANDIDATE"
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "source",
        "source_assets",
        "source_tables",
        "scope",
    }
    REQUIRED_SOURCE = {
        "source_id",
        "doi",
        "pmcid",
        "license",
        "article_locator",
        "supplementary_package_locator",
    }
    REQUIRED_ASSET = {"asset_id", "relative_path", "sha256", "expected_bytes"}
    REQUIRED_TABLE = {
        "asset_id",
        "worksheet",
        "header_row",
        "condition_id",
        "protein_identifier_column",
        "quantification_column",
        "expected_data_rows",
        "expected_explicit_zero_count",
    }
    REQUIRED_SCOPE = {
        "admission_status",
        "data_semantics",
        "permitted_use",
        "prohibited_use",
        "model_status",
        "scientific_submission_ready",
    }

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=True)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise FulltextGoldSourceAuditError(f"cannot parse {label}") from exc

    def _asset_path(self, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise FulltextGoldSourceAuditError(f"{label} must use a POSIX relative path")
        pure_path = PurePosixPath(relative_path)
        if pure_path.is_absolute() or not pure_path.parts or ".." in pure_path.parts:
            raise FulltextGoldSourceAuditError(f"{label} escapes the source asset root")
        path = (self.assets_root / Path(*pure_path.parts)).resolve(strict=False)
        if not path.is_relative_to(self.assets_root) or not path.is_file():
            raise FulltextGoldSourceAuditError(f"{label} is missing or outside the source asset root")
        return path

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path], list[dict[str, Any]]]:
        registry = self._json(self.registry_path, "R3 full-text gold source registry")
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise FulltextGoldSourceAuditError("R3 full-text gold registry fields are invalid")
        if registry.get("audit_id") != self.AUDIT_ID:
            raise FulltextGoldSourceAuditError("R3 full-text gold registry identity is invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise FulltextGoldSourceAuditError("R3 full-text gold evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "R3 full-text gold evaluated_at")

        source = _mapping(registry.get("source"), "R3 full-text gold source")
        if set(source) != self.REQUIRED_SOURCE:
            raise FulltextGoldSourceAuditError("R3 full-text gold source fields are invalid")
        if (
            source.get("source_id") != "PMC7788026"
            or source.get("doi") != "10.1186/s12645-020-00071-6"
            or source.get("pmcid") != "PMC7788026"
            or source.get("license") != "CC-BY-4.0"
        ):
            raise FulltextGoldSourceAuditError("R3 full-text gold source identity or licence is invalid")
        for field in ("article_locator", "supplementary_package_locator"):
            if not _string(source.get(field), f"R3 full-text gold {field}").startswith("https://"):
                raise FulltextGoldSourceAuditError("R3 full-text gold source locator is invalid")

        assets: dict[str, Path] = {}
        for value in _list(registry.get("source_assets"), "R3 full-text gold assets", minimum=4):
            asset = _mapping(value, "R3 full-text gold asset")
            if set(asset) != self.REQUIRED_ASSET:
                raise FulltextGoldSourceAuditError("R3 full-text gold asset fields are invalid")
            asset_id = _string(asset.get("asset_id"), "R3 full-text gold asset id")
            if asset_id in assets:
                raise FulltextGoldSourceAuditError("R3 full-text gold asset id is duplicated")
            path = self._asset_path(_string(asset.get("relative_path"), asset_id), asset_id)
            if path.stat().st_size != asset.get("expected_bytes"):
                raise FulltextGoldSourceAuditError(f"R3 full-text gold asset size mismatch: {asset_id}")
            if _sha256(path) != _checksum(asset.get("sha256"), asset_id):
                raise FulltextGoldSourceAuditError(f"R3 full-text gold asset checksum mismatch: {asset_id}")
            assets[asset_id] = path

        tables: list[dict[str, Any]] = []
        condition_ids: set[str] = set()
        for value in _list(registry.get("source_tables"), "R3 full-text gold tables", minimum=3):
            table = _mapping(value, "R3 full-text gold table")
            if set(table) != self.REQUIRED_TABLE:
                raise FulltextGoldSourceAuditError("R3 full-text gold table fields are invalid")
            if table.get("asset_id") not in assets or table.get("header_row") not in {1, 2}:
                raise FulltextGoldSourceAuditError("R3 full-text gold asset/table map is invalid")
            condition_id = _string(table.get("condition_id"), "R3 full-text gold condition id")
            if condition_id in condition_ids:
                raise FulltextGoldSourceAuditError("R3 full-text gold condition id is duplicated")
            condition_ids.add(condition_id)
            for field in ("worksheet", "protein_identifier_column", "quantification_column"):
                _string(table.get(field), f"R3 full-text gold table {field}")
            for field in ("expected_data_rows", "expected_explicit_zero_count"):
                if not isinstance(table.get(field), int) or table[field] < 0:
                    raise FulltextGoldSourceAuditError("R3 full-text gold table count is invalid")
            tables.append(table)

        scope = _mapping(registry.get("scope"), "R3 full-text gold scope")
        if set(scope) != self.REQUIRED_SCOPE:
            raise FulltextGoldSourceAuditError("R3 full-text gold scope fields are invalid")
        if (
            scope.get("admission_status") != self.STATUS
            or scope.get("data_semantics") != "SOURCE_NATIVE_REPORTED_PROTEIN_ROWS"
            or scope.get("model_status") != "NOT_FITTED"
            or scope.get("scientific_submission_ready") is not False
        ):
            raise FulltextGoldSourceAuditError("R3 full-text gold scope is over-promoted or invalid")
        for field in ("permitted_use", "prohibited_use"):
            if any(
                not isinstance(item, str) or not item.strip()
                for item in _list(scope.get(field), field, minimum=2)
            ):
                raise FulltextGoldSourceAuditError("R3 full-text gold scope list is invalid")
        return registry, assets, tables

    @staticmethod
    def _extract_table(path: Path, table: dict[str, Any]) -> tuple[list[dict[str, str]], int]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if table["worksheet"] not in workbook.sheetnames:
                raise FulltextGoldSourceAuditError("R3 full-text gold worksheet is missing")
            sheet = workbook[table["worksheet"]]
            source_rows = sheet.iter_rows(values_only=True)
            for _ in range(table["header_row"] - 1):
                next(source_rows)
            header = [_csv_cell(value).strip() for value in next(source_rows)]
            try:
                protein_index = header.index(table["protein_identifier_column"])
                quantity_index = header.index(table["quantification_column"])
            except ValueError as exc:
                raise FulltextGoldSourceAuditError("R3 full-text gold column is missing") from exc
            cell_width = len(header)
            rows: list[dict[str, str]] = []
            explicit_zeros = 0
            for source_row, values in enumerate(source_rows, start=table["header_row"] + 1):
                if len(values) != cell_width:
                    raise FulltextGoldSourceAuditError("R3 full-text gold data row width is invalid")
                protein_identifier = _csv_cell(values[protein_index]).strip()
                if not protein_identifier:
                    raise FulltextGoldSourceAuditError("R3 full-text gold data row lacks protein identity")
                quantity = _finite(values[quantity_index], f"R3 full-text gold row {source_row}")
                if quantity == 0.0:
                    explicit_zeros += 1
                rows.append(
                    {
                        "analysis_unit_id": f"PMC7788026:{table['worksheet']}:{source_row}",
                        "source_article_doi": "10.1186/s12645-020-00071-6",
                        "source_pmcid": "PMC7788026",
                        "source_license": "CC-BY-4.0",
                        "source_asset_id": table["asset_id"],
                        "source_worksheet": table["worksheet"],
                        "source_row": str(source_row),
                        "source_cell_range": f"{table['worksheet']}!A{source_row}:{chr(64 + cell_width)}{source_row}",
                        "source_condition_id": table["condition_id"],
                        "protein_source_identifier": protein_identifier,
                        "quantification_column": table["quantification_column"],
                        "author_reported_quantification": _csv_cell(values[quantity_index]),
                    }
                )
        except (OSError, StopIteration) as exc:
            raise FulltextGoldSourceAuditError("R3 full-text gold table cannot be read") from exc
        finally:
            try:
                workbook.close()
            except UnboundLocalError:
                pass
        return rows, explicit_zeros

    def run(self, *, strict: bool = False) -> FulltextGoldSourceAuditSummary:
        if not strict:
            raise FulltextGoldSourceAuditError("R3 full-text gold source audit requires --strict")
        if self.output_root.exists():
            raise FulltextGoldSourceAuditError("R3 full-text gold source audit already executed")
        registry, assets, tables = self._registry()
        rows: list[dict[str, str]] = []
        table_counts: dict[str, int] = {}
        zero_counts: dict[str, int] = {}
        for table in tables:
            extracted, zero_count = self._extract_table(assets[table["asset_id"]], table)
            if len(extracted) != table["expected_data_rows"] or zero_count != table[
                "expected_explicit_zero_count"
            ]:
                raise FulltextGoldSourceAuditError("R3 full-text gold table count does not match registry")
            table_counts[table["condition_id"]] = len(extracted)
            zero_counts[table["condition_id"]] = zero_count
            rows.extend(extracted)
        source_map = self.assets_root / self.DERIVED_RELATIVE
        if source_map.exists():
            raise FulltextGoldSourceAuditError("R3 full-text gold source-cell map already exists")
        source_map.parent.mkdir(parents=True, exist_ok=True)
        with source_map.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(rows[0]), lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)
        scope = _mapping(registry["scope"], "R3 full-text gold scope")
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "evaluated_at": registry["evaluated_at"],
            "registry_sha256": _sha256(self.registry_path),
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "source": registry["source"],
            "verified_source_asset_count": len(assets),
            "source_asset_sha256": {asset_id: _sha256(path) for asset_id, path in assets.items()},
            "technical_scope": scope,
            "source_condition_row_counts": table_counts,
            "analysis_unit_count": len(rows),
            "explicit_zero_count": sum(zero_counts.values()),
            "source_to_cell_map": {
                "location": self.DERIVED_RELATIVE,
                "sha256": _sha256(source_map),
                "coordinate_definition": "Every output row retains the original worksheet row and full source-row cell range.",
                "transformation": "openpyxl read_only/data_only extraction; no abundance rescaling, protein remapping, non-detection inference or rank transformation.",
            },
            "status": self.STATUS,
            "target_status": "SOURCE_NATIVE_RANK_ONLY",
            "model_fitted": False,
            "paired_ablations_run": False,
            "external_ood_evaluated": False,
            "independent_validation": False,
            "scientific_submission_ready": False,
        }
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report_sha256": hashlib.sha256(_canonical(report)).hexdigest(),
            "verified_source_asset_count": len(assets),
            "table_count": len(tables),
            "analysis_unit_count": len(rows),
            "explicit_zero_count": sum(zero_counts.values()),
            "target_status": "SOURCE_NATIVE_RANK_ONLY",
            "model_fitted": False,
            "scientific_submission_ready": False,
        }
        self.output_root.mkdir(parents=True, exist_ok=False)
        self._write(self.output_root / "fulltext_gold_source_audit_report.json", report)
        self._write(self.output_root / "fulltext_gold_source_audit_receipt.json", receipt)
        return FulltextGoldSourceAuditSummary(
            source_asset_count=len(assets),
            table_count=len(tables),
            analysis_unit_count=len(rows),
            explicit_zero_count=sum(zero_counts.values()),
            status=self.STATUS,
            receipt_path=self.output_root / "fulltext_gold_source_audit_receipt.json",
        )

    def verify(self) -> FulltextGoldSourceAuditSummary:
        report_path = self.output_root / "fulltext_gold_source_audit_report.json"
        receipt_path = self.output_root / "fulltext_gold_source_audit_receipt.json"
        report = self._json(report_path, "R3 full-text gold report")
        receipt = self._json(receipt_path, "R3 full-text gold receipt")
        source_map = _mapping(report.get("source_to_cell_map"), "R3 full-text gold source map")
        source_map_path = self.assets_root / _string(source_map.get("location"), "R3 source map location")
        if (
            report.get("audit_id") != self.AUDIT_ID
            or receipt.get("audit_id") != self.AUDIT_ID
            or report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report_sha256") != _sha256(report_path)
            or report.get("target_status") != "SOURCE_NATIVE_RANK_ONLY"
            or report.get("model_fitted") is not False
            or report.get("scientific_submission_ready") is not False
            or not source_map_path.is_file()
            or source_map.get("sha256") != _sha256(source_map_path)
        ):
            raise FulltextGoldSourceAuditError("R3 full-text gold source receipt is invalid")
        return FulltextGoldSourceAuditSummary(
            source_asset_count=int(receipt["verified_source_asset_count"]),
            table_count=int(receipt["table_count"]),
            analysis_unit_count=int(receipt["analysis_unit_count"]),
            explicit_zero_count=int(receipt["explicit_zero_count"]),
            status=self.STATUS,
            receipt_path=receipt_path,
        )
