"""Audit a CC-BY human-plasma protein-corona source for a future frozen R4 study.

This source was discovered after the R3 protocol and results were frozen.  It
is deliberately retained in a separate R4 ledger: it cannot be concatenated
with the R3 target, used for R3 model selection, or counted as an independent
laboratory because it shares the Michigan State source lineage.
"""

from __future__ import annotations

import csv
import json
import math
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R4SmallMoleculeCoronaSourceAuditError(RuntimeError):
    """Raised when a source cell cannot be bound to the official CC-BY assets."""


@dataclass(frozen=True)
class R4SmallMoleculeCoronaSourceAuditSummary:
    """Compact accounting for the separately scoped R4 source candidate."""

    source_asset_count: int
    protein_row_count: int
    all_measurement_batch_count: int
    corona_measurement_batch_count: int
    rank_qualified_measurement_batch_count: int
    shared_canonical_protein_count: int
    source_cell_count: int
    candidate_positive_source_cell_count: int
    receipt_path: Path


class R4SmallMoleculeCoronaSourceAuditWorkflow:
    """Create a no-imputation cell map for PMC11544298 without retrofitting R3."""

    AUDIT_ID = "bioif-r4-small-molecule-corona-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R4_T157_SMALL_MOLECULE_CORONA_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_4/small_molecule_corona_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R4_PMC11544298_small_molecule_corona_source_cell_map.csv"
    STATUS = "ADMITTED_SEPARATE_R4_SAME_LINEAGE_CANDIDATE_PENDING_NEW_PROTOCOL"
    SOURCE_CELL_FIELDS = (
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "canonical_accession",
        "measurement_batch_id",
        "biological_unit_id",
        "condition_label",
        "analysis_candidate_eligible",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    )
    REQUIRED_TOP_LEVEL = {
        "schema_version",
        "audit_id",
        "evaluated_at",
        "evidence_class",
        "allowed_claim_level",
        "article",
        "source_scope",
        "source_assets",
        "r3_reference_asset",
        "worksheet_contracts",
        "quantification_contract",
        "admission_minimums",
        "claim_boundary",
    }

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise R4SmallMoleculeCoronaSourceAuditError(f"cannot parse {label}") from exc

    @staticmethod
    def _numeric(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, str) and value.strip().upper() == "NA":
            return None
        if isinstance(value, bool) or not isinstance(value, int | float):
            raise R4SmallMoleculeCoronaSourceAuditError("author quantity must be numeric or blank")
        numeric = float(value)
        if not math.isfinite(numeric):
            raise R4SmallMoleculeCoronaSourceAuditError("author quantity must be finite")
        return numeric

    def _under(self, root: Path, relative_path: str, label: str) -> Path:
        if "\\" in relative_path:
            raise R4SmallMoleculeCoronaSourceAuditError(f"{label} must use POSIX separators")
        pure = PurePosixPath(relative_path)
        if pure.is_absolute() or not pure.parts or ".." in pure.parts:
            raise R4SmallMoleculeCoronaSourceAuditError(f"{label} escapes its root")
        path = (root / Path(*pure.parts)).resolve(strict=False)
        if not path.is_relative_to(root) or not path.is_file():
            raise R4SmallMoleculeCoronaSourceAuditError(f"{label} is missing or outside its root")
        return path

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path], Path]:
        registry = self._json(self.registry_path, "R4 small-molecule source registry")
        if set(registry) != self.REQUIRED_TOP_LEVEL or registry.get("schema_version") != 1:
            raise R4SmallMoleculeCoronaSourceAuditError("registry fields are invalid")
        if registry.get("audit_id") != self.AUDIT_ID:
            raise R4SmallMoleculeCoronaSourceAuditError("registry audit ID is invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise R4SmallMoleculeCoronaSourceAuditError("registry evidence boundary is invalid")
        _string(registry.get("evaluated_at"), "evaluated_at")
        _string(registry.get("claim_boundary"), "claim_boundary")
        if _mapping(registry.get("article"), "article") != {
            "pmcid": "PMC11544298",
            "doi": "10.1038/s41467-024-53966-z",
            "title": "Small molecule modulation of protein corona for deep plasma proteome profiling",
            "publication_year": 2024,
            "license": "CC-BY-4.0",
            "full_text_locator": "https://europepmc.org/articles/PMC11544298",
            "supplementary_locator": "https://www.ebi.ac.uk/europepmc/webservices/rest/PMC11544298/supplementaryFiles",
        }:
            raise R4SmallMoleculeCoronaSourceAuditError("article declaration is invalid")
        scope = _mapping(registry.get("source_scope"), "source scope")
        expected_scope_keys = {
            "source_id",
            "laboratory_anchor",
            "source_lineage",
            "biofluid",
            "nanoparticle",
            "analysis_role",
            "prohibited_interpretations",
        }
        if set(scope) != expected_scope_keys or (
            scope.get("source_id") != "PMC11544298_SMALL_MOLECULE_HUMAN_PLASMA_CORONA"
            or scope.get("laboratory_anchor") != "Michigan State University-led small-molecule protein-corona study"
            or scope.get("source_lineage") != "NOT_INDEPENDENT_OF_EXISTING_R3_MICHIGAN_STATE_LINEAGE"
            or scope.get("biofluid") != "commercial pooled healthy human plasma plus four-donor plasma panel"
            or scope.get("nanoparticle") != "80 nm polystyrene nanoparticles"
            or scope.get("analysis_role") != "SEPARATE_R4_NEW_PROTOCOL_CANDIDATE_ONLY"
            or not isinstance(scope.get("prohibited_interpretations"), list)
            or len(scope["prohibited_interpretations"]) != 5
        ):
            raise R4SmallMoleculeCoronaSourceAuditError("source scope is invalid")

        assets = registry.get("source_assets")
        if not isinstance(assets, list) or len(assets) != 4:
            raise R4SmallMoleculeCoronaSourceAuditError("source assets are invalid")
        paths: dict[str, Path] = {}
        for item in assets:
            item = _mapping(item, "source asset")
            if set(item) != {"asset_id", "relative_path", "sha256", "expected_bytes"}:
                raise R4SmallMoleculeCoronaSourceAuditError("source asset fields are invalid")
            asset_id = _string(item.get("asset_id"), "source asset ID")
            path = self._under(self.assets_root, _string(item.get("relative_path"), asset_id), asset_id)
            expected_bytes = item.get("expected_bytes")
            if (
                asset_id in paths
                or isinstance(expected_bytes, bool)
                or not isinstance(expected_bytes, int)
                or path.stat().st_size != expected_bytes
                or _sha256(path) != _checksum(item.get("sha256"), asset_id)
            ):
                raise R4SmallMoleculeCoronaSourceAuditError("source asset checksum differs")
            paths[asset_id] = path
        required_assets = {"supplementary_package_zip", "data_1", "data_3", "data_5"}
        if set(paths) != required_assets:
            raise R4SmallMoleculeCoronaSourceAuditError("source assets are incomplete")
        with zipfile.ZipFile(paths["supplementary_package_zip"]) as archive:
            archive_names = set(archive.namelist())
            for asset_id in ("data_1", "data_3", "data_5"):
                name = paths[asset_id].name
                if name not in archive_names or archive.read(name) != paths[asset_id].read_bytes():
                    raise R4SmallMoleculeCoronaSourceAuditError(
                        "official extraction differs from supplementary package"
                    )

        reference = _mapping(registry.get("r3_reference_asset"), "R3 reference asset")
        if set(reference) != {"relative_path", "sha256"}:
            raise R4SmallMoleculeCoronaSourceAuditError("R3 reference asset fields are invalid")
        feature_path = self._under(
            self.root,
            _string(reference.get("relative_path"), "R3 reference asset"),
            "R3 reference asset",
        )
        if _sha256(feature_path) != _checksum(reference.get("sha256"), "R3 reference asset"):
            raise R4SmallMoleculeCoronaSourceAuditError("R3 reference asset checksum differs")

        contracts = registry.get("worksheet_contracts")
        expected_contracts = {
            "data_1": ("Supp Data 1", 20592, 12, 1609, 90, 31),
            "data_3": ("Supp Data 3", 5080, 12, 379, 89, 5),
            "data_5": ("20240825_ProteinData_Choline_PS", 1260, 54, 75, 75, 28),
        }
        if not isinstance(contracts, list) or len(contracts) != len(expected_contracts):
            raise R4SmallMoleculeCoronaSourceAuditError("worksheet contracts are invalid")
        for item in contracts:
            item = _mapping(item, "worksheet contract")
            fields = {
                "asset_id",
                "worksheet",
                "header_row",
                "first_data_row",
                "expected_rows",
                "expected_columns",
                "expected_direct_r3_rows",
                "expected_direct_r3_accessions",
                "candidate_condition_count",
            }
            if set(item) != fields:
                raise R4SmallMoleculeCoronaSourceAuditError("worksheet contract fields are invalid")
            asset_id = _string(item.get("asset_id"), "worksheet contract asset ID")
            expected = expected_contracts.pop(asset_id, None)
            values = tuple(
                item[key]
                for key in (
                    "worksheet",
                    "expected_rows",
                    "expected_columns",
                    "expected_direct_r3_rows",
                    "expected_direct_r3_accessions",
                    "candidate_condition_count",
                )
            )
            expected_header = 3 if item.get("asset_id") in {"data_1", "data_3"} else 4
            expected_first = expected_header + 1
            if (
                expected is None
                or values != expected
                or item.get("header_row") != expected_header
                or item.get("first_data_row") != expected_first
            ):
                raise R4SmallMoleculeCoronaSourceAuditError("worksheet contract differs")
        if expected_contracts:
            raise R4SmallMoleculeCoronaSourceAuditError("worksheet contracts are incomplete")
        quantification = _mapping(registry.get("quantification_contract"), "quantification contract")
        if quantification != {
            "source_accession_policy": "map only a non-empty direct UniProt accession that is exactly present in the frozen R3 feature table; never collapse semicolon-delimited source groups",  # noqa: E501
            "author_quantity_type_data_1_and_3": "NORMALIZED_INTENSITY",
            "author_quantity_type_data_5": "AUTHOR_REPORTED_ABUNDANCE",
            "rank_eligibility": "candidate-eligible source condition plus strictly positive finite author-reported value",  # noqa: E501
            "numeric_zero_policy": "retain as NUMERIC_ZERO and exclude from rank; never impute",
            "source_na_policy": "retain literal author NA markers as SOURCE_NA and exclude from rank; never coerce to zero or impute",  # noqa: E501
            "blank_policy": "retain as SOURCE_BLANK and exclude from rank; never impute",
            "plasma_alone_policy": "retain in source cell map with analysis_candidate_eligible=false; never treat as a corona measurement batch",  # noqa: E501
            "raw_scale_cross_study_use": "PROHIBITED",
        }:
            raise R4SmallMoleculeCoronaSourceAuditError("quantification contract is invalid")
        if _mapping(registry.get("admission_minimums"), "admission minimums") != {
            "corona_measurement_batch_count": 136,
            "rank_qualified_measurement_batch_count": 134,
            "minimum_positive_shared_proteins_per_candidate_measurement_batch": 10,
            "independent_laboratory_anchor_count_contributed": 0,
        }:
            raise R4SmallMoleculeCoronaSourceAuditError("admission minimums are invalid")
        return registry, paths, feature_path

    @staticmethod
    def _feature_accessions(feature_path: Path) -> set[str]:
        with feature_path.open(newline="", encoding="utf-8") as stream:
            accessions = {row["canonical_accession"] for row in csv.DictReader(stream)}
        if len(accessions) != 99:
            raise R4SmallMoleculeCoronaSourceAuditError("frozen R3 feature population differs")
        return accessions

    @staticmethod
    def _state(value: float | None, raw_value: Any, candidate: bool) -> tuple[str, str, str]:
        if value is None:
            if isinstance(raw_value, str) and raw_value.strip().upper() == "NA":
                return "SOURCE_NA", "", "false"
            return "SOURCE_BLANK", "", "false"
        rendered = format(value, ".17g")
        if value == 0.0:
            return "NUMERIC_ZERO", rendered, "false"
        if value < 0.0:
            return "NEGATIVE_FINITE", rendered, "false"
        return "POSITIVE_FINITE", rendered, "true" if candidate else "false"

    @staticmethod
    def _slug(value: Any) -> str:
        source = str(value).strip()
        return "".join(character if character.isalnum() else "_" for character in source).strip("_")

    @staticmethod
    def _direct_accession(value: Any, features: set[str]) -> str | None:
        accession = str(value or "").strip()
        if ";" in accession or accession not in features:
            return None
        return accession

    def _data_1_or_3(
        self,
        path: Path,
        asset_id: str,
        expected_sheet: str,
        expected_rows: int,
        expected_columns: int,
        features: set[str],
    ) -> list[dict[str, str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if workbook.sheetnames != [expected_sheet]:
            raise R4SmallMoleculeCoronaSourceAuditError("supplementary worksheet differs")
        sheet = workbook.active
        if sheet.max_row != expected_rows or sheet.max_column != expected_columns:
            raise R4SmallMoleculeCoronaSourceAuditError("supplementary worksheet dimensions differ")
        header = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
        expected_header = (
            None,
            "Accession",
            "Description",
            "Coverage [%]",
            "# Peptides",
            "# Unique Peptides",
            "Group",
            "Concentration (µg/µl)",
            "Rep_1",
            "Rep_2",
            "Rep_3",
        )
        if header[:11] != expected_header:
            raise R4SmallMoleculeCoronaSourceAuditError("supplementary worksheet header differs")
        rows: list[dict[str, str]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            accession = self._direct_accession(row[1], features)
            if accession is None:
                continue
            group = str(row[6] or "").strip()
            concentration = str(row[7] or "").strip()
            if not group or not concentration:
                raise R4SmallMoleculeCoronaSourceAuditError("supplementary condition metadata is empty")
            candidate = group != "Plasma"
            condition = f"{group}_{concentration}"
            for column_index in range(8, 11):
                replicate = header[column_index]
                value = self._numeric(row[column_index])
                state, rendered, rank_eligible = self._state(value, row[column_index], candidate)
                rows.append(
                    {
                        "source_id": "PMC11544298_SMALL_MOLECULE_HUMAN_PLASMA_CORONA",
                        "laboratory_anchor": "Michigan State University-led small-molecule protein-corona study",
                        "source_asset_id": asset_id,
                        "source_worksheet": expected_sheet,
                        "source_row": str(row_number),
                        "source_coordinate": f"{get_column_letter(column_index + 1)}{row_number}",
                        "source_identifier": str(row[1]),
                        "canonical_accession": accession,
                        "measurement_batch_id": f"R4_PMC11544298_{asset_id}_{self._slug(condition)}_{replicate}",
                        "biological_unit_id": "POOLED_HUMAN_PLASMA",
                        "condition_label": condition,
                        "analysis_candidate_eligible": "true" if candidate else "false",
                        "author_quantity_type": "NORMALIZED_INTENSITY",
                        "author_numeric_value": rendered,
                        "author_value_state": state,
                        "rank_target_eligible": rank_eligible,
                    }
                )
        return rows

    def _data_5(self, path: Path, features: set[str]) -> list[dict[str, str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        expected_sheet = "20240825_ProteinData_Choline_PS"
        if workbook.sheetnames != [expected_sheet]:
            raise R4SmallMoleculeCoronaSourceAuditError("donor worksheet differs")
        sheet = workbook.active
        if sheet.max_row != 1260 or sheet.max_column != 54:
            raise R4SmallMoleculeCoronaSourceAuditError("donor worksheet dimensions differ")
        header = next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
        if header[1:5] != (
            "Protein Groups",
            "Genes",
            "Protein Descriptions",
            "Number of modified sequences identified (Experiment-wide)",
        ):
            raise R4SmallMoleculeCoronaSourceAuditError("donor worksheet header differs")
        sample_columns = [
            (index, label)
            for index, label in enumerate(header)
            if index >= 5 and isinstance(label, str) and not label.startswith(("Mean_", "SD_", "NACount_"))
        ]
        if len(sample_columns) != 28 or any("_donor" not in label for _, label in sample_columns):
            raise R4SmallMoleculeCoronaSourceAuditError("donor sample columns differ")
        rows: list[dict[str, str]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            accession = self._direct_accession(row[1], features)
            if accession is None:
                continue
            for column_index, label in sample_columns:
                value = self._numeric(row[column_index])
                state, rendered, rank_eligible = self._state(value, row[column_index], True)
                donor = label.rsplit("_donor", 1)[1]
                rows.append(
                    {
                        "source_id": "PMC11544298_SMALL_MOLECULE_HUMAN_PLASMA_CORONA",
                        "laboratory_anchor": "Michigan State University-led small-molecule protein-corona study",
                        "source_asset_id": "data_5",
                        "source_worksheet": expected_sheet,
                        "source_row": str(row_number),
                        "source_coordinate": f"{get_column_letter(column_index + 1)}{row_number}",
                        "source_identifier": str(row[1]),
                        "canonical_accession": accession,
                        "measurement_batch_id": f"R4_PMC11544298_data_5_{label}",
                        "biological_unit_id": f"DONOR_{donor}",
                        "condition_label": label.rsplit("_donor", 1)[0],
                        "analysis_candidate_eligible": "true",
                        "author_quantity_type": "AUTHOR_REPORTED_ABUNDANCE",
                        "author_numeric_value": rendered,
                        "author_value_state": state,
                        "rank_target_eligible": rank_eligible,
                    }
                )
        return rows

    def _cells(self, paths: dict[str, Path], feature_path: Path) -> list[dict[str, str]]:
        features = self._feature_accessions(feature_path)
        rows = [
            *self._data_1_or_3(paths["data_1"], "data_1", "Supp Data 1", 20592, 12, features),
            *self._data_1_or_3(paths["data_3"], "data_3", "Supp Data 3", 5080, 12, features),
            *self._data_5(paths["data_5"], features),
        ]
        if len(rows) != 8064:
            raise R4SmallMoleculeCoronaSourceAuditError("source-cell accounting differs")
        by_batch: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in rows:
            by_batch[row["measurement_batch_id"]].append(row)
        candidate_batches = {
            batch: values for batch, values in by_batch.items() if values[0]["analysis_candidate_eligible"] == "true"
        }
        if len(by_batch) != 142 or len(candidate_batches) != 136:
            raise R4SmallMoleculeCoronaSourceAuditError("measurement batch accounting differs")
        rank_qualified_batches: dict[str, list[dict[str, str]]] = {}
        for batch, values in candidate_batches.items():
            accessions = [row["canonical_accession"] for row in values]
            positive = [row for row in values if row["rank_target_eligible"] == "true"]
            if len(accessions) != len(set(accessions)):
                raise R4SmallMoleculeCoronaSourceAuditError(f"candidate measurement batch {batch} is invalid")
            if len(positive) >= 10:
                rank_qualified_batches[batch] = values
        if len(rank_qualified_batches) != 134:
            raise R4SmallMoleculeCoronaSourceAuditError("rank-qualified measurement batch accounting differs")
        # The frozen R3 feature table maps 90, 89, and 75 direct accessions in
        # the three workbooks respectively.  Their union is 97 (not the sum or
        # the intersection), which is the population represented in this
        # source-cell ledger.
        if len({row["canonical_accession"] for row in rows}) != 97:
            raise R4SmallMoleculeCoronaSourceAuditError("shared canonical protein population differs")
        if sum(row["rank_target_eligible"] == "true" for row in rows) != 7075:
            raise R4SmallMoleculeCoronaSourceAuditError("positive source-cell accounting differs")
        return rows

    def run(self, *, strict: bool = False) -> R4SmallMoleculeCoronaSourceAuditSummary:
        if not strict:
            raise R4SmallMoleculeCoronaSourceAuditError("R4 small-molecule source audit requires --strict")
        if self.output_root.exists():
            raise R4SmallMoleculeCoronaSourceAuditError("R4 small-molecule source audit already executed")
        registry, paths, feature_path = self._registry()
        rows = self._cells(paths, feature_path)
        derived = self.assets_root / self.DERIVED_RELATIVE
        if derived.exists():
            raise R4SmallMoleculeCoronaSourceAuditError("derived source cell map already exists")
        derived.parent.mkdir(parents=True, exist_ok=True)
        with derived.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=self.SOURCE_CELL_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
        self.output_root.mkdir(parents=True, exist_ok=False)
        all_batches = {row["measurement_batch_id"] for row in rows}
        candidate_batches = {
            row["measurement_batch_id"] for row in rows if row["analysis_candidate_eligible"] == "true"
        }
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "source_assets": {
                item["asset_id"]: {
                    "relative_path": item["relative_path"],
                    "sha256": _sha256(paths[item["asset_id"]]),
                }
                for item in registry["source_assets"]
            },
            "r3_reference_asset": {
                "relative_path": registry["r3_reference_asset"]["relative_path"],
                "sha256": _sha256(feature_path),
            },
            "source_cell_map": {"relative_path": self.DERIVED_RELATIVE, "sha256": _sha256(derived)},
            "protein_row_count": 2168,
            "all_measurement_batch_count": len(all_batches),
            "corona_measurement_batch_count": len(candidate_batches),
            "rank_qualified_measurement_batch_count": sum(
                sum(row["rank_target_eligible"] == "true" for row in rows if row["measurement_batch_id"] == batch) >= 10
                for batch in candidate_batches
            ),
            "shared_canonical_protein_count": len({row["canonical_accession"] for row in rows}),
            "source_cell_count": len(rows),
            "candidate_positive_source_cell_count": sum(row["rank_target_eligible"] == "true" for row in rows),
            "same_lineage_independent_laboratory_anchor_count_contributed": 0,
            "claim_boundary": registry["claim_boundary"],
        }
        report_path = self.output_root / "small_molecule_corona_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report": {"relative_path": report_path.name, "sha256": _sha256(report_path)},
            "source_cell_map": report["source_cell_map"],
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "small_molecule_corona_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R4SmallMoleculeCoronaSourceAuditSummary(
            len(paths),
            report["protein_row_count"],
            report["all_measurement_batch_count"],
            report["corona_measurement_batch_count"],
            report["rank_qualified_measurement_batch_count"],
            report["shared_canonical_protein_count"],
            report["source_cell_count"],
            report["candidate_positive_source_cell_count"],
            receipt_path,
        )

    def verify(self) -> R4SmallMoleculeCoronaSourceAuditSummary:
        _, paths, feature_path = self._registry()
        report_path = self.output_root / "small_molecule_corona_source_audit_report.json"
        receipt_path = self.output_root / "small_molecule_corona_source_audit_receipt.json"
        report = self._json(report_path, "R4 small-molecule source audit report")
        receipt = self._json(receipt_path, "R4 small-molecule source audit receipt")
        if (
            report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report", {}).get("sha256") != _sha256(report_path)
        ):
            raise R4SmallMoleculeCoronaSourceAuditError("audit receipt differs")
        cell_map = self._under(
            self.assets_root,
            _string(report.get("source_cell_map", {}).get("relative_path"), "source cell map path"),
            "source cell map",
        )
        if report["source_cell_map"].get("sha256") != _sha256(cell_map):
            raise R4SmallMoleculeCoronaSourceAuditError("source cell map checksum differs")
        with cell_map.open(newline="", encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        if (set(rows[0]) if rows else set()) != set(self.SOURCE_CELL_FIELDS):
            raise R4SmallMoleculeCoronaSourceAuditError("source cell map fields differ")
        summary = R4SmallMoleculeCoronaSourceAuditSummary(
            len(paths),
            int(report["protein_row_count"]),
            int(report["all_measurement_batch_count"]),
            int(report["corona_measurement_batch_count"]),
            int(report["rank_qualified_measurement_batch_count"]),
            int(report["shared_canonical_protein_count"]),
            int(report["source_cell_count"]),
            int(report["candidate_positive_source_cell_count"]),
            receipt_path,
        )
        if (
            len(rows) != summary.source_cell_count
            or sum(row["rank_target_eligible"] == "true" for row in rows)
            != summary.candidate_positive_source_cell_count
            or report.get("r3_reference_asset", {}).get("sha256") != _sha256(feature_path)
        ):
            raise R4SmallMoleculeCoronaSourceAuditError("audit accounting differs")
        return summary
