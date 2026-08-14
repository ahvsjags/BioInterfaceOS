"""Audit the CC0 paper-attached PXD068107 technical source data.

The source is the public source-data workbook for the Westlake University
study on blood contamination in nanoparticle plasma proteomics.  The 21 rows
in the primary heatmap are nanoparticle/technical conditions, not 21 donors.
This audit therefore creates a source-local technical OOD ledger with one
pooled biological unit and preserves the donor-level claim boundary.
"""

from __future__ import annotations

import csv
import json
import math
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from biointerfaceos.r3_uniprot_mapping import _canonical, _checksum, _mapping, _sha256, _string


class R4PXD068107SourceAuditError(RuntimeError):
    """Raised when the public PXD068107 source cannot be audited safely."""


@dataclass(frozen=True)
class R4PXD068107SourceAuditSummary:
    source_asset_count: int
    protein_row_count: int
    measurement_batch_count: int
    rank_qualified_measurement_batch_count: int
    shared_canonical_protein_count: int
    source_cell_count: int
    positive_source_cell_count: int
    biological_unit_count: int
    laboratory_anchor_count: int
    receipt_path: Path


class R4PXD068107SourceAuditWorkflow:
    """Create and verify the byte-traceable PXD068107 source-cell map."""

    AUDIT_ID = "bioif-r4-pxd068107-technical-source-audit-v1.0.0"
    REGISTRY_RELATIVE = "docs/data/R4_T264_PXD068107_SOURCE_REGISTRY.json"
    OUTPUT_RELATIVE = "reports/review_round_4/pxd068107_source_audit/v1.0.0"
    DERIVED_RELATIVE = "derived/R4_PXD068107_technical_source_cell_map.csv"
    STATUS = "R4_CC0_PAPER_SOURCE_AUDITED_TECHNICAL_OOD_EXPLORATORY"
    SOURCE_ID = "PXD068107_WESTLAKE_OMNIPROT_TECHNICAL"
    LABORATORY = "Westlake University"
    HEATMAP_ASSET = "heatmap_2b"
    SOURCE_CELLS = (
        "source_id",
        "laboratory_anchor",
        "source_asset_id",
        "source_worksheet",
        "source_row",
        "source_coordinate",
        "source_identifier",
        "canonical_accession",
        "measurement_batch_id",
        "biological_unit_id",
        "condition_label",
        "analysis_candidate_eligible",
        "author_quantity_type",
        "author_numeric_value",
        "author_value_state",
        "rank_target_eligible",
    )

    def __init__(
        self,
        root: Path,
        assets_root: Path,
        *,
        registry_path: Path | None = None,
        output_root: Path | None = None,
    ) -> None:
        self.root = root.resolve(strict=True)
        self.assets_root = assets_root.resolve(strict=False)
        self.registry_path = registry_path or self.root / self.REGISTRY_RELATIVE
        self.output_root = output_root or self.root / self.OUTPUT_RELATIVE

    @staticmethod
    def _json(path: Path, label: str) -> dict[str, Any]:
        try:
            return _mapping(json.loads(path.read_text(encoding="utf-8")), label)
        except (OSError, UnicodeError, json.JSONDecodeError, TypeError, ValueError) as exc:
            raise R4PXD068107SourceAuditError(f"cannot parse {label}") from exc

    @staticmethod
    def _write(path: Path, value: Any) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(_canonical(value))

    @staticmethod
    def _number(value: Any) -> tuple[str, str, bool]:
        if value is None or not str(value).strip() or str(value).strip().upper() == "NA":
            return ("SOURCE_NA", "", False)
        try:
            numeric = float(value)
        except (TypeError, ValueError) as exc:
            raise R4PXD068107SourceAuditError("paper intensity must be numeric or NA") from exc
        if not math.isfinite(numeric):
            raise R4PXD068107SourceAuditError("paper intensity must be finite")
        rendered = format(numeric, ".17g")
        if numeric > 0:
            return ("POSITIVE_FINITE", rendered, True)
        if numeric == 0:
            return ("NUMERIC_ZERO", rendered, False)
        return ("NEGATIVE_FINITE", rendered, False)

    def _under(self, root: Path, relative: str, label: str) -> Path:
        if "\\" in relative:
            raise R4PXD068107SourceAuditError(f"{label} must use POSIX path separators")
        pure = PurePosixPath(relative)
        if pure.is_absolute() or not pure.parts or ".." in pure.parts:
            raise R4PXD068107SourceAuditError(f"{label} escapes its root")
        path = (root / Path(*pure.parts)).resolve(strict=False)
        if not path.is_relative_to(root) or not path.is_file():
            raise R4PXD068107SourceAuditError(f"{label} is missing")
        return path

    def _registry(self) -> tuple[dict[str, Any], dict[str, Path], Path]:
        registry = self._json(self.registry_path, "PXD068107 source registry")
        expected = {
            "schema_version",
            "audit_id",
            "evaluated_at",
            "evidence_class",
            "allowed_claim_level",
            "article",
            "dataset",
            "source_scope",
            "source_assets",
            "r3_reference_asset",
            "table_contract",
            "quantification_contract",
            "admission_minimums",
            "claim_boundary",
        }
        if (
            set(registry) != expected
            or registry.get("schema_version") != 1
            or registry.get("audit_id") != self.AUDIT_ID
        ):
            raise R4PXD068107SourceAuditError("registry identity or fields are invalid")
        if (
            registry.get("evidence_class") != "DEVELOPMENT_OBSERVATION"
            or registry.get("allowed_claim_level") != "EXPLORATORY"
        ):
            raise R4PXD068107SourceAuditError("registry evidence boundary is invalid")
        article = _mapping(registry.get("article"), "article")
        if article.get("pmcid") != "PMC12808129" or article.get("license") != "CC0":
            raise R4PXD068107SourceAuditError("article declaration is invalid")
        dataset = _mapping(registry.get("dataset"), "dataset")
        if dataset.get("accession") != "PXD068107" or dataset.get("license") != "CC0":
            raise R4PXD068107SourceAuditError("dataset declaration is invalid")
        scope = _mapping(registry.get("source_scope"), "source scope")
        if scope.get("source_id") != self.SOURCE_ID or scope.get("laboratory_anchor") != self.LABORATORY:
            raise R4PXD068107SourceAuditError("source scope identity is invalid")
        if scope.get("biological_unit_count") != 1 or scope.get("claim_level") != "TECHNICAL_OOD_ONLY":
            raise R4PXD068107SourceAuditError("source biological scope is invalid")
        assets = registry.get("source_assets")
        if not isinstance(assets, list) or len(assets) != 6:
            raise R4PXD068107SourceAuditError("source asset declaration is invalid")
        paths: dict[str, Path] = {}
        for item in assets:
            item = _mapping(item, "source asset")
            asset_id = _string(item.get("asset_id"), "source asset ID")
            if asset_id in paths:
                raise R4PXD068107SourceAuditError("source asset IDs are duplicated")
            path = self._under(self.assets_root, _string(item.get("relative_path"), asset_id), asset_id)
            if not isinstance(item.get("expected_bytes"), int) or path.stat().st_size != item["expected_bytes"]:
                raise R4PXD068107SourceAuditError("source asset byte count differs")
            if _sha256(path) != _checksum(item.get("sha256"), asset_id):
                raise R4PXD068107SourceAuditError("source asset checksum differs")
            paths[asset_id] = path
        if self.HEATMAP_ASSET not in paths:
            raise R4PXD068107SourceAuditError("primary heatmap asset is missing")
        reference = _mapping(registry.get("r3_reference_asset"), "R3 reference asset")
        feature_path = self._under(
            self.root, _string(reference.get("relative_path"), "R3 reference asset"), "R3 reference asset"
        )
        if _sha256(feature_path) != _checksum(reference.get("sha256"), "R3 reference asset"):
            raise R4PXD068107SourceAuditError("R3 reference checksum differs")
        contract = _mapping(registry.get("table_contract"), "table contract")
        if (
            contract.get("worksheet") != "2b_heatmap"
            or contract.get("expected_row_count") != 22
            or contract.get("expected_column_count") != 7820
        ):
            raise R4PXD068107SourceAuditError("heatmap table contract is invalid")
        if contract.get("expected_measurement_batch_count") != 21 or contract.get("minimum_targets_per_batch") != 10:
            raise R4PXD068107SourceAuditError("heatmap batch contract is invalid")
        return registry, paths, feature_path

    @staticmethod
    def _feature_accessions(path: Path) -> set[str]:
        with path.open(newline="", encoding="utf-8") as stream:
            reader = csv.DictReader(stream)
            if not reader.fieldnames or "canonical_accession" not in reader.fieldnames:
                raise R4PXD068107SourceAuditError("R3 feature header is missing")
            accessions = {row["canonical_accession"] for row in reader}
        if len(accessions) != 99:
            raise R4PXD068107SourceAuditError("R3 feature population differs")
        return accessions

    def _cells(self, heatmap: Path, feature_path: Path) -> tuple[list[dict[str, str]], dict[str, Any]]:
        features = self._feature_accessions(feature_path)
        workbook = load_workbook(heatmap, read_only=True, data_only=True)
        try:
            if "2b_heatmap" not in workbook.sheetnames:
                raise R4PXD068107SourceAuditError("primary worksheet is missing")
            sheet = workbook["2b_heatmap"]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if len(rows) != 22 or len(rows[0]) != 7820 or rows[0][-1] != "NP_ID":
            raise R4PXD068107SourceAuditError("primary heatmap dimensions differ")
        header = list(rows[0][:-1])
        if len(set(header)) != len(header):
            raise R4PXD068107SourceAuditError("primary heatmap protein headers are duplicated")
        target_columns = [(index + 1, value) for index, value in enumerate(header) if value in features]
        if len(target_columns) != 98:
            raise R4PXD068107SourceAuditError("primary heatmap target intersection differs")
        if any(not str(row[-1]).strip() for row in rows[1:]) or len({str(row[-1]) for row in rows[1:]}) != 21:
            raise R4PXD068107SourceAuditError("technical condition identifiers differ")
        cells: list[dict[str, str]] = []
        positive_by_batch: dict[str, int] = defaultdict(int)
        for row_index, row in enumerate(rows[1:], start=2):
            condition = str(row[-1]).strip()
            batch = f"PXD068107_{condition}"
            for column_index, accession in target_columns:
                state, rendered, eligible = self._number(row[column_index - 1])
                cells.append(
                    {
                        "source_id": self.SOURCE_ID,
                        "laboratory_anchor": self.LABORATORY,
                        "source_asset_id": self.HEATMAP_ASSET,
                        "source_worksheet": "2b_heatmap",
                        "source_row": str(row_index),
                        "source_coordinate": f"{get_column_letter(column_index)}{row_index}",
                        "source_identifier": accession,
                        "canonical_accession": accession,
                        "measurement_batch_id": batch,
                        "biological_unit_id": "POOLED_TECHNICAL_PLASMA_SOURCE",
                        "condition_label": condition,
                        "analysis_candidate_eligible": "true",
                        "author_quantity_type": "SOURCE_DATA_INTENSITY",
                        "author_numeric_value": rendered,
                        "author_value_state": state,
                        "rank_target_eligible": "true" if eligible else "false",
                    }
                )
                positive_by_batch[batch] += int(eligible)
        qualified = sum(count >= 10 for count in positive_by_batch.values())
        if (
            len(cells) != 2058
            or len(positive_by_batch) != 21
            or qualified != 21
            or min(positive_by_batch.values()) < 10
        ):
            raise R4PXD068107SourceAuditError("PXD068107 source accounting differs")
        totals = {
            "protein_row_count": 21,
            "measurement_batch_count": 21,
            "rank_qualified_measurement_batch_count": qualified,
            "shared_canonical_protein_count": 98,
            "source_cell_count": len(cells),
            "positive_source_cell_count": sum(positive_by_batch.values()),
            "positive_by_batch": dict(sorted(positive_by_batch.items())),
        }
        return cells, totals

    def run(self, *, strict: bool = False) -> R4PXD068107SourceAuditSummary:
        if not strict:
            raise R4PXD068107SourceAuditError("PXD068107 source audit requires --strict")
        if self.output_root.exists():
            raise R4PXD068107SourceAuditError("PXD068107 source audit already executed")
        registry, paths, feature_path = self._registry()
        cells, totals = self._cells(paths[self.HEATMAP_ASSET], feature_path)
        derived = self.assets_root / self.DERIVED_RELATIVE
        if derived.exists():
            raise R4PXD068107SourceAuditError("derived source cell map already exists")
        derived.parent.mkdir(parents=True, exist_ok=True)
        with derived.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=self.SOURCE_CELLS)
            writer.writeheader()
            writer.writerows(cells)
        self.output_root.mkdir(parents=True, exist_ok=False)
        report = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "evidence_class": "DEVELOPMENT_OBSERVATION",
            "allowed_claim_level": "EXPLORATORY",
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
            "source_assets": {
                item["asset_id"]: {"relative_path": item["relative_path"], "sha256": _sha256(paths[item["asset_id"]])}
                for item in registry["source_assets"]
            },
            "r3_reference_asset": {
                "relative_path": registry["r3_reference_asset"]["relative_path"],
                "sha256": _sha256(feature_path),
            },
            "source_cell_map": {"relative_path": self.DERIVED_RELATIVE, "sha256": _sha256(derived)},
            **totals,
            "biological_unit_count": 1,
            "laboratory_anchor_count": 1,
            "claim_boundary": registry["claim_boundary"],
        }
        report_path = self.output_root / "pxd068107_source_audit_report.json"
        self._write(report_path, report)
        receipt = {
            "schema_version": 1,
            "audit_id": self.AUDIT_ID,
            "status": self.STATUS,
            "report": {"relative_path": report_path.name, "sha256": _sha256(report_path)},
            "source_cell_map": report["source_cell_map"],
            "model_fitted": False,
            "independent_validation": False,
            "external_scientific_reproduction": False,
            "scientific_submission_ready": False,
        }
        receipt_path = self.output_root / "pxd068107_source_audit_receipt.json"
        self._write(receipt_path, receipt)
        return R4PXD068107SourceAuditSummary(
            len(paths),
            totals["protein_row_count"],
            totals["measurement_batch_count"],
            totals["rank_qualified_measurement_batch_count"],
            totals["shared_canonical_protein_count"],
            totals["source_cell_count"],
            totals["positive_source_cell_count"],
            1,
            1,
            receipt_path,
        )

    def verify(self) -> R4PXD068107SourceAuditSummary:
        registry, paths, feature_path = self._registry()
        report_path = self.output_root / "pxd068107_source_audit_report.json"
        receipt_path = self.output_root / "pxd068107_source_audit_receipt.json"
        report = self._json(report_path, "PXD068107 source audit report")
        receipt = self._json(receipt_path, "PXD068107 source audit receipt")
        if (
            report.get("status") != self.STATUS
            or receipt.get("status") != self.STATUS
            or receipt.get("report", {}).get("sha256") != _sha256(report_path)
        ):
            raise R4PXD068107SourceAuditError("PXD068107 audit receipt differs")
        cell_map_info = _mapping(report.get("source_cell_map"), "source cell map")
        cell_map = self._under(
            self.assets_root, _string(cell_map_info.get("relative_path"), "source cell map"), "source cell map"
        )
        if cell_map_info.get("sha256") != _sha256(cell_map):
            raise R4PXD068107SourceAuditError("source cell map checksum differs")
        with cell_map.open(newline="", encoding="utf-8") as stream:
            rows = list(csv.DictReader(stream))
        if not rows or set(rows[0]) != set(self.SOURCE_CELLS):
            raise R4PXD068107SourceAuditError("source cell map schema differs")
        _, totals = self._cells(paths[self.HEATMAP_ASSET], feature_path)
        if any(report.get(key) != value for key, value in totals.items() if key != "positive_by_batch"):
            raise R4PXD068107SourceAuditError("PXD068107 accounting differs")
        if report.get("positive_by_batch") != totals["positive_by_batch"] or len(rows) != totals["source_cell_count"]:
            raise R4PXD068107SourceAuditError("PXD068107 cell accounting differs")
        return R4PXD068107SourceAuditSummary(
            len(paths),
            totals["protein_row_count"],
            totals["measurement_batch_count"],
            totals["rank_qualified_measurement_batch_count"],
            totals["shared_canonical_protein_count"],
            totals["source_cell_count"],
            totals["positive_source_cell_count"],
            int(report["biological_unit_count"]),
            int(report["laboratory_anchor_count"]),
            receipt_path,
        )
